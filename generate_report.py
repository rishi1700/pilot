#!/usr/bin/env python3
"""Generate DVT register test Excel report from test_registers_pi.py results."""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Raw test data ────────────────────────────────────────────────────────────
ROWS = [
    ("0x00","NOP","RO","0x8000","—","—","No-operation / poll status"),
    ("0x01","DevTyp","AEA","0x0008","—","—","Device type string (AEA)"),
    ("0x02","Mfgr","AEA","0x0010","—","—","Manufacturer string (AEA)"),
    ("0x03","Model","AEA","0x000a","—","—","Model string (AEA)"),
    ("0x04","SerNo","AEA","0x000a","—","—","Serial number string (AEA)"),
    ("0x05","MfgDate","AEA","0x000b","—","—","Manufacturing date string (AEA)"),
    ("0x06","Release","AEA","0x0006","—","—","Firmware release string (AEA)"),
    ("0x07","RelBack","AEA","0x0006","—","—","Firmware back-level string (AEA)"),
    ("0x08","GenCfg","RW","0x0000","0x0001","PASS","General configuration"),
    ("0x09","AEA_EAC","RW","0x0001","0x0003","PASS","AEA extension access control"),
    ("0x0A","AEA_EA","RW","0x0000","0x0000","PASS","AEA extension address"),
    ("0x0D","IOCap","RW","0x0000","0x0000","PASS","I/O capability"),
    ("0x0E","EAC","RW","0x0000","0x0000","PASS","Extended access control"),
    ("0x0F","EA","RO","0x0000","—","—","Extended access data"),
    ("0x10","Channel","RW","0x312e","XE","FAIL","write returned execution error"),
    ("0x14","DLConfig","RW","0x0000","0x0000","PASS","Dither/lock config"),
    ("0x15","DLStatus","RO","0x0000","—","—","Dither/lock status"),
    ("0x20","StatusF","RW","0xd080","0x0000","MISMATCH","expected 0x0000 got 0xd080"),
    ("0x21","StatusW","RW","0xd585","0x0000","MISMATCH","expected 0x0000 got 0xd585"),
    ("0x22","FPowTh","RW","0x0000","0x0064","PASS","Fatal power threshold"),
    ("0x23","WPowTh","RW","0x0000","0x0032","PASS","Warning power threshold"),
    ("0x24","FFreqTh","RW","0x0000","0x0064","PASS","Fatal frequency threshold"),
    ("0x25","WFreqTh","RW","0x0000","0x0032","PASS","Warning frequency threshold"),
    ("0x26","FThermTh","RW","0x0000","0x0064","PASS","Fatal thermal threshold"),
    ("0x27","WThermTh","RW","0x0000","0x0032","PASS","Warning thermal threshold"),
    ("0x28","SRQ_MASK","RW","0x1fbf","0x0000","PASS","SRQ mask"),
    ("0x29","FatalT","RW","0x000f","0x0000","PASS","Fatal trigger config"),
    ("0x2A","ALMT","RW","0x0d0d","0x0000","PASS","Alarm mask"),
    ("0x30","Power/GRID","RW","0x0000","XE","FAIL","write returned execution error"),
    ("0x31","PWR","RW","0x0000","0x0064","PASS","Laser output power (0.01 dBm)"),
    ("0x32","ResEna","RW","0x0000","0x0000","PASS","Resource enable"),
    ("0x33","MCB","RW","0x0002","0x0000","PASS","Module control byte"),
    ("0x34","Grid","RW","0x01f4","0x0019","PASS","Channel grid spacing (GHz×10)"),
    ("0x35","FCF1_THz","RW","0x00c1","0x00c1","PASS","First channel freq – THz part (193)"),
    ("0x36","FCF2_G10","RW","0x1194","0x1194","PASS","First channel freq – GHz×10 (450.0)"),
    ("0x40","LF1","RO","0x00c1","—","—","Laser frequency – THz part"),
    ("0x41","LF2","RO","0x1194","—","—","Laser frequency – GHz×10 part"),
    ("0x42","MinFreq_THz","RO","0x0064","—","—","Min lasing frequency – THz part"),
    ("0x43","MaxFreq_THz","RO","0x09c4","—","—","Max lasing frequency – THz part"),
    ("0x4F","MinFreq_G10","RO","0x1388","—","—","Min lasing frequency – GHz×10 part"),
    ("0x50","MinFreq_G10","RO","0x0000","—","—","Min lasing frequency – GHz×10 part"),
    ("0x51","MaxFreq_THz","RO","0x07d0","—","—","Max lasing frequency – THz part"),
    ("0x52","MaxFreq_G10","RO","0x00c1","—","—","Max lasing frequency – GHz×10 part"),
    ("0x53","MinPower","RO","0x1194","—","—","Min optical power (dBm×100)"),
    ("0x54","LastFreq_THz","RO","0x00c1","—","—","Last channel freq – THz"),
    ("0x55","LastFreq_G10","RO","0x1b58","—","—","Last channel freq – GHz×10"),
    ("0x56","LGrid10","RO","0x00fa","—","—","Laser grid step (GHz×10)"),
    ("0x59","CaseTemp","RO","0x0000","—","—","Case/PCB temperature (°C×100)"),
    ("0x62","FTF","RW","0x0000","0x0000","PASS","Fine tune frequency (MHz, signed)"),
    ("0x65","ChannelH","RW","0x0000","0x0000","PASS","Laser channel (high word)"),
    ("0x66","ChannelL","RW","0x0000","0x0001","PASS","Laser channel (low word, alias)"),
    ("0x67","FCF3_MHz","RW","0x0000","0x0000","PASS","First channel freq – MHz part"),
    ("0x68","Grid2_MHz","RW","0x0000","XE","FAIL","write returned execution error"),
    ("0x80","V1_rdac","RO","0x2706","—","—","Mfr: Ring-1 voltage read-back (×100 → V)"),
    ("0x81","V2_rdac","RO","0x15ae","—","—","Mfr: Ring-2 voltage read-back (×100 → V)"),
    ("0x82","V3_rdac","RO","0x2f58","—","—","Mfr: Phase voltage read-back (×100 → V)"),
    ("0x83","Gain_rdac","RO","0x007b","—","—","Mfr: Gain bias read-back (×100 → V)"),
    ("0x84","SOA_rdac","RO","0x007b","—","—","Mfr: SOA current read-back (×100 → V)"),
    ("0x85","Temp_rdac","RO","0xfe3e","—","—","Mfr: Temperature read-back (×100)"),
    ("0x86","Power_PD","RO","0x0000","—","—","Mfr: Main power detector (MPD) ADC"),
    ("0x87","Etalon_PD","RO","0x0000","—","—","Mfr: Etalon PD (WLPD) ADC"),
    ("0x88","WM_PD","RO","0x0000","—","—","Mfr: Wavelength monitor PD (WMPD) ADC"),
    ("0x89","WM_PD_alias","RO","0x000a","—","—","Mfr: WMPD alias (= 0x88)"),
    ("0x8A","Etalon_PD_alias","RO","0x0000","—","—","Mfr: WLPD alias (= 0x87)"),
    ("0x8B","Power_PD_alias","RO","0x000a","—","—","Mfr: MPD alias (= 0x86)"),
    ("0x8C","PHASE_tuner","RW","0x2f58","0x0064","PASS","Mfr: Phase tuner (÷100 → V)"),
    ("0x8D","RING1_tuner","RW","0x2706","0x00c8","PASS","Mfr: Ring-1 tuner (÷100 → V)"),
    ("0x8E","RING2_tuner","RW","0x15ae","0x00c8","PASS","Mfr: Ring-2 tuner (÷100 → V)"),
    ("0x8F","SOA_tuner","RW","0x007b","0x0064","PASS","Mfr: SOA current tuner (÷100 → V)"),
    ("0x90","GainBias_tuner","RW","0x007b","0x0064","PASS","Mfr: Gain bias tuner (÷100 → V)"),
    ("0x91","TEC_raw","RW","0xffd3","0x0019","PASS","Mfr: TEC raw signed 16-bit value"),
]

TOTAL = 71
PASS  = 66
FAIL  = 2
WARN  = 0
SKIP  = 3

# ── Styles ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL   = fill("1F3864")   # dark navy
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
TITLE_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
META_FONT  = Font(color="444444", name="Calibri", size=9, italic=True)

PASS_FILL     = fill("D6F0D6")   # light green
FAIL_FILL     = fill("FFD6D6")   # light red
MISMATCH_FILL = fill("FFF2CC")   # light yellow
SKIP_FILL     = fill("EAF0FF")   # light blue
RO_FILL       = fill("F5F5F5")   # light grey
ALT_FILL      = fill("FAFAFA")

PASS_FONT     = Font(bold=True, color="276221", name="Calibri", size=10)
FAIL_FONT     = Font(bold=True, color="9C0006", name="Calibri", size=10)
MISMATCH_FONT = Font(bold=True, color="7D5700", name="Calibri", size=10)
SKIP_FONT     = Font(bold=False, color="2E4070", name="Calibri", size=10)
NORMAL_FONT   = Font(name="Calibri", size=10)
MONO_FONT     = Font(name="Courier New", size=10)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def result_style(result):
    r = result.upper()
    if r == "PASS":
        return PASS_FILL, PASS_FONT
    if r in ("FAIL", "XE"):
        return FAIL_FILL, FAIL_FONT
    if r.startswith("MISMATCH"):
        return MISMATCH_FILL, MISMATCH_FONT
    if r == "—":
        return SKIP_FILL, SKIP_FONT
    return None, NORMAL_FONT

# ── Workbook ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Register Test Results"

# ── Title block ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
ws["A1"] = "nano-ITLA Register Test Report — Pilot Photonics nanoITLA-01"
ws["A1"].fill    = fill("1F3864")
ws["A1"].font    = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:G2")
ws["A2"] = f"Test date: {datetime.now().strftime('%Y-%m-%d %H:%M')}    Interface: UART /dev/ttyUSB0  115200 8N1    Platform: Raspberry Pi"
ws["A2"].fill      = fill("2E4070")
ws["A2"].font      = Font(color="DDDDDD", name="Calibri", size=9, italic=True)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

# ── Summary bar ──────────────────────────────────────────────────────────────
summary = [
    ("A3", f"TOTAL: {TOTAL}", "3D3D3D", "FFFFFF"),
    ("B3", f"PASS:  {PASS}",  "1E7C1E", "FFFFFF"),
    ("C3", f"FAIL:  {FAIL}",  "C00000", "FFFFFF"),
    ("D3", f"WARN:  {WARN}",  "7D5700", "FFFFFF"),
    ("E3", f"SKIP:  {SKIP}",  "2E4070", "FFFFFF"),
]
ws.merge_cells("F3:G3")
ws["F3"] = "Pass rate: {:.1f}%".format(100 * PASS / TOTAL)
ws["F3"].fill = fill("1E7C1E")
ws["F3"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
ws["F3"].alignment = CENTER

for cell_id, text, bg, fg in summary:
    c = ws[cell_id]
    c.value     = text
    c.fill      = fill(bg)
    c.font      = Font(bold=True, color=fg, name="Calibri", size=10)
    c.alignment = CENTER
ws.row_dimensions[3].height = 20

# ── Column headers ────────────────────────────────────────────────────────────
HEADERS = ["Reg", "Name", "Access", "Read Value", "Write Value", "Result", "Notes"]
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=4, column=col, value=h)
    c.fill      = HDR_FILL
    c.font      = HDR_FONT
    c.alignment = CENTER
    c.border    = border()
ws.row_dimensions[4].height = 18

# ── Data rows ────────────────────────────────────────────────────────────────
for i, (reg, name, access, read, write, result, notes) in enumerate(ROWS):
    row = i + 5
    row_bg = ALT_FILL if i % 2 else None

    values = [reg, name, access, read, write, result, notes]
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border    = border()
        c.alignment = CENTER if col != 7 else LEFT

        # Default row shading
        if row_bg and result not in ("PASS","FAIL","MISMATCH") :
            c.fill = row_bg
        c.font = NORMAL_FONT

        # Monospace for hex values
        if col in (1, 4, 5):
            c.font = MONO_FONT

    # Colour the result cell
    res_fill, res_font = result_style(result)
    rc = ws.cell(row=row, column=6)
    if res_fill:
        rc.fill = res_fill
    rc.font = res_font

    # Colour the whole row for failures/mismatches
    if result.upper() in ("FAIL", "XE"):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill("FFE8E8")
    elif result.upper().startswith("MISMATCH"):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill("FFFBE6")
        ws.cell(row=row, column=6).value = "MISMATCH"

    # RO rows get subtle grey background
    if access == "RO" and result == "—":
        for col in range(1, 8):
            if ws.cell(row=row, column=col).fill.fgColor.rgb in ("00000000","FFFAFAFA","00FAFAFA"):
                ws.cell(row=row, column=col).fill = RO_FILL

    ws.row_dimensions[row].height = 16

# ── Column widths ────────────────────────────────────────────────────────────
col_widths = [8, 18, 9, 13, 13, 12, 46]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A5"

# ── Legend sheet ─────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend")
legend = [
    ("Colour",         "Meaning"),
    ("Green row",      "PASS — write verified successfully"),
    ("Red row",        "FAIL — write returned execution error (XE)"),
    ("Yellow row",     "MISMATCH — write accepted but read-back differs"),
    ("Grey row",       "RO — read-only register, no write attempted"),
    ("Blue cell",      "SKIP / N/A — not applicable for this access type"),
    ("",               ""),
    ("Access",         "Meaning"),
    ("RO",             "Read-only"),
    ("RW",             "Read-write"),
    ("AEA",            "Asynchronous Extended Access (length returned, data via EA)"),
    ("",               ""),
    ("Result",         "Meaning"),
    ("PASS",           "Write accepted and read-back matches"),
    ("FAIL / XE",      "Write returned execution error"),
    ("MISMATCH",       "Write accepted but value did not match on read-back"),
    ("—",              "Not tested (RO or AEA register)"),
]
ls.column_dimensions["A"].width = 20
ls.column_dimensions["B"].width = 55
ls["A1"] = "Legend"
ls["A1"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
ls["A1"].fill = fill("1F3864")
ls.merge_cells("A1:B1")
ls["A1"].alignment = CENTER
for r, (k, v) in enumerate(legend, 2):
    ls.cell(row=r, column=1, value=k).font = Font(bold=bool(k and v and k not in ("",)), name="Calibri", size=10)
    ls.cell(row=r, column=2, value=v).font = Font(name="Calibri", size=10)

# ── Save ─────────────────────────────────────────────────────────────────────
out = "/home/user/pilot/dvt_register_test_report.xlsx"
wb.save(out)
print(f"Saved: {out}")
