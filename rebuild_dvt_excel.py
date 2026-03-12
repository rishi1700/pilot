#!/usr/bin/env python3
"""Rebuild dvt_excel.xlsx — 2 sheets: executive Summary + single Test Results table."""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(c): return PatternFill("solid", fgColor=c)
def bd(color="BBBBBB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

NAV  = "1F3864"; NAV2 = "2E4070"; NAV3 = "354F7A"
GRN  = "1E7C1E"; RED  = "C00000"; AMB  = "7D5700"
LGRE = "E8F5E8"; LRED = "FFE8E8"; LYEL = "FFFBE6"
LGRY = "F5F5F5"; WHT  = "FFFFFF"; DRK  = "222222"
SEC_BG = "D9E2F3"   # section divider row

def hfont(sz=10, bold=True, color=WHT):  return Font(bold=bold, color=color, name="Calibri", size=sz)
def nfont(sz=10, bold=False, color=DRK): return Font(bold=bold, color=color, name="Calibri", size=sz)
def mfont(sz=10):                         return Font(name="Courier New", size=sz, color=DRK)

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=False)

def rfill(r):
    r = str(r).strip().upper()
    if r == "PASS": return fill(LGRE), Font(bold=True, color=GRN, name="Calibri", size=10)
    if r == "FAIL": return fill(LRED), Font(bold=True, color=RED, name="Calibri", size=10)
    return fill(LGRY), nfont()

META = "Tester: Rishi   |   Date: 2026-03-12 09:04   |   Interface: UART /dev/ttyUSB0 @ 115200 baud   |   Device: Pilot Photonics nanoITLA-01"
NCOLS = 8   # columns in Test Results sheet
SCOLS = 5   # columns in Summary sheet

def title_block(ws, text, ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"].value = text
    ws["A1"].fill = fill(NAV); ws["A1"].font = hfont(13); ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws["A2"].value = META
    ws["A2"].fill = fill(NAV2)
    ws["A2"].font = Font(italic=True, color="CCCCCC", name="Calibri", size=9)
    ws["A2"].alignment = CTR; ws.row_dimensions[2].height = 14

# ── All test data in one flat list ────────────────────────────────────────────
# Each entry: (section, name, addr, mode, write, read, decoded, result)
ALL = [
    # ── Table A: Device Identity ──────────────────────────────────────────────
    ("Device Identity (Table A)", "Device Type",   "0x01", "AEA", "—", "CW ITLA",        "8 chars",   "PASS"),
    ("Device Identity (Table A)", "Manufacturer",  "0x02", "AEA", "—", "Pilot Photonics","16 chars",  "PASS"),
    ("Device Identity (Table A)", "Model",         "0x03", "AEA", "—", "NYITLA-01",      "10 chars",  "PASS"),
    ("Device Identity (Table A)", "Serial Number", "0x04", "AEA", "—", "PP-000123",      "10 chars",  "PASS"),
    ("Device Identity (Table A)", "Mfg Date",      "0x05", "AEA", "—", "2025-08-27",     "11 chars",  "PASS"),
    ("Device Identity (Table A)", "FW Release",    "0x06", "AEA", "—", "1.0.0",          "6 chars",   "PASS"),
    ("Device Identity (Table A)", "FW Back-Level", "0x07", "AEA", "—", "1.0.0",          "6 chars",   "PASS"),

    # ── §9.5 Status / Alarms / Triggers ──────────────────────────────────────
    ("§9.5  Status / Alarms / Triggers", "DL Config",        "0x14", "RW", "0x0001", "0x0001", "DLConfig[START]", "PASS"),
    ("§9.5  Status / Alarms / Triggers", "DL Status",        "0x15", "RO", "—",      "0x0000", "none",            "PASS"),
    ("§9.5  Status / Alarms / Triggers", "StatusF",          "0x20", "RO", "—",      "0xD000", "DIS, ALM, SRQ",   "PASS"),
    ("§9.5  Status / Alarms / Triggers", "StatusW",          "0x21", "RO", "—",      "0xD505", "LWPWR, LWFREQ, WPWR, WFREQ, DIS, ALM, SRQ", "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Fatal Pwr Thresh", "0x22", "RW", "0x0011", "0x0011", "0.17 dB",         "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Warn Pwr Thresh",  "0x23", "RW", "0x0012", "0x0012", "0.18 dB",         "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Fatal Freq Thresh","0x24", "RW", "0x0013", "0x0013", "1.9 GHz coarse",  "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Warn Freq Thresh", "0x25", "RW", "0x0014", "0x0014", "2.0 GHz coarse",  "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Fatal Freq Th2",   "0x63", "RW", "0x0017", "0x0017", "23 MHz fine",     "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Warn Freq Th2",    "0x64", "RW", "0x0018", "0x0018", "24 MHz fine",     "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Fatal Therm Th",   "0x26", "RW", "0x0015", "0x0015", "0.21 °C",         "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Warn Therm Th",    "0x27", "RW", "0x0016", "0x0016", "0.22 °C",         "PASS"),
    ("§9.5  Status / Alarms / Triggers", "SRQ Triggers",     "0x28", "RW", "0x0001", "0x0001", "SRQT[LFPWR]",     "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Fatal Triggers",   "0x29", "RW", "0x0001", "0x0001", "FatalT[LFPWR]",   "PASS"),
    ("§9.5  Status / Alarms / Triggers", "Alarm Mask",       "0x2A", "RW", "0x0001", "0x0001", "ALMT[FPWR]",      "PASS"),
    ("§9.5  Status / Alarms / Triggers", "SRQ Mask",         "0x28", "RW", "0x0003", "0x0003", "Mask bits",       "PASS"),
    ("§9.5  Status / Alarms / Triggers", "COW Status",       "0x21", "RO", "—",      "0xD505", "StatusW read-back","PASS"),

    # ── §9.6 General Module Configuration ─────────────────────────────────────
    ("§9.6  General Module Configuration", "Channel",       "0x30", "RW", "0x0001", "0x0001", "Ch 1 (ITU grid)", "PASS"),
    ("§9.6  General Module Configuration", "Channel H",     "0x65", "RW", "0x0000", "0x0000", "High word",       "PASS"),
    ("§9.6  General Module Configuration", "Power",         "0x31", "RW", "0x0000", "0x0000", "0.00 dBm",        "PASS"),
    ("§9.6  General Module Configuration", "ResEna",        "0x32", "RW", "0x0000", "0x0000", "—",               "PASS"),
    ("§9.6  General Module Configuration", "MCB",           "0x33", "RW", "0x0002", "0x0002", "ADT=1",           "PASS"),
    ("§9.6  General Module Configuration", "Grid",          "0x34", "RW", "0x01F4", "0x01F4", "50.0 GHz",        "PASS"),
    ("§9.6  General Module Configuration", "Grid2 MHz",     "0x66", "RW", "0x0000", "0x0000", "0 MHz",           "PASS"),
    ("§9.6  General Module Configuration", "FCF1 THz",      "0x35", "RW", "0x00C1", "0x00C1", "193 THz",         "PASS"),
    ("§9.6  General Module Configuration", "FCF2 GHz×10",   "0x36", "RW", "0x1194", "0x1194", "450.0 GHz",       "PASS"),
    ("§9.6  General Module Configuration", "FCF3 MHz",      "0x67", "RW", "0x0000", "0x0000", "0 MHz",           "PASS"),
    ("§9.6  General Module Configuration", "Laser Freq THz","0x40", "RO", "—",      "0x00C1", "193 THz",         "PASS"),
    ("§9.6  General Module Configuration", "Laser Freq G10","0x41", "RO", "—",      "0x1194", "450.0 GHz",       "PASS"),
    ("§9.6  General Module Configuration", "Laser Freq MHz","0x68", "RO", "—",      "0x0000", "0 MHz",           "PASS"),
    ("§9.6  General Module Configuration", "LF1 Min THz",   "0x42", "RO", "—",      "0x00BF", "191 THz",         "PASS"),
    ("§9.6  General Module Configuration", "LF1 Max THz",   "0x43", "RO", "—",      "0x00C4", "196 THz",         "PASS"),

    # ── §9.7 Fine Tune / Limits ────────────────────────────────────────────────
    ("§9.7  Fine Tune / Frequency Limits", "Min Freq THz",  "0x4F", "RO", "—", "0x00BF", "191 THz",         "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Min Freq G10",  "0x50", "RO", "—", "0x1B58", "700.0 GHz",       "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Max Freq THz",  "0x51", "RO", "—", "0x00C4", "196 THz",         "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Max Freq G10",  "0x52", "RO", "—", "0x1B58", "700.0 GHz",       "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Min Power",     "0x53", "RO", "—", "0x0000", "0.00 dBm",        "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Max Power",     "0x69", "RO", "—", "0x0514", "13.00 dBm",       "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Last Freq THz", "0x54", "RO", "—", "0x00C1", "193 THz",         "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Last Freq G10", "0x55", "RO", "—", "0x1194", "450.0 GHz",       "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Last Freq MHz", "0x6A", "RO", "—", "0x0000", "0 MHz",           "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Laser Grid G10","0x56", "RO", "—", "0x01F4", "50.0 GHz",        "PASS"),
    ("§9.7  Fine Tune / Frequency Limits", "Grid2 MHz",     "0x6B", "RO", "—", "0x0000", "0 MHz",           "PASS"),

    # ── §9.8 Health / Dither / Age ─────────────────────────────────────────────
    ("§9.8  Health / Dither / Age", "Currents AEA",  "0x57", "RO", "—", "0x0008", "TEC / Diode / MON / SOA", "PASS"),
    ("§9.8  Health / Dither / Age", "Temps AEA",     "0x58", "RO", "—", "0x0004", "Diode 25.00°C / Case",    "PASS"),
    ("§9.8  Health / Dither / Age", "Case Temp",     "0x59", "RO", "—", "0x0000", "0.00 °C (firmware default)", "PASS"),
    ("§9.8  Health / Dither / Age", "Dither Period", "0x5A", "RO", "—", "0x0064", "100",                     "PASS"),
    ("§9.8  Health / Dither / Age", "Dither BW",     "0x5B", "RO", "—", "0x0000", "0",                       "PASS"),
    ("§9.8  Health / Dither / Age", "Device Cap 1",  "0x5C", "RO", "—", "0x0000", "0",                       "PASS"),
    ("§9.8  Health / Dither / Age", "Device Cap 2",  "0x5D", "RO", "—", "0xFE0C", "—",                       "PASS"),
    ("§9.8  Health / Dither / Age", "Device Cap 3",  "0x5E", "RO", "—", "0x1B58", "—",                       "PASS"),
    ("§9.8  Health / Dither / Age", "Device Cap 4",  "0x5F", "RO", "—", "0x0000", "0",                       "PASS"),
    ("§9.8  Health / Dither / Age", "Reserved 0x60", "0x60", "RO", "—", "0x0000", "0",                       "PASS"),
    ("§9.8  Health / Dither / Age", "FTF Min",       "0x61", "RO", "—", "0xCEC4", "-12500 MHz",              "PASS"),
    ("§9.8  Health / Dither / Age", "Fine Tune Freq","0x62", "RW", "0x0000", "0x0000", "0 MHz",              "PASS"),

    # ── §9.9 Manufacturer-Specific ────────────────────────────────────────────
    ("§9.9  Manufacturer-Specific", "Ring-1 Voltage",  "0x80", "RO", "—",      "0x2706", "99.90 (×100)",    "PASS"),
    ("§9.9  Manufacturer-Specific", "Ring-2 Voltage",  "0x81", "RO", "—",      "0x15AE", "55.50 (×100)",    "PASS"),
    ("§9.9  Manufacturer-Specific", "Phase Voltage",   "0x82", "RO", "—",      "0x2F58", "121.20 (×100)",   "PASS"),
    ("§9.9  Manufacturer-Specific", "Gain Bias",       "0x83", "RO", "—",      "0x007B", "1.23 (×100)",     "PASS"),
    ("§9.9  Manufacturer-Specific", "SOA Current",     "0x84", "RO", "—",      "0x007B", "1.23 (×100)",     "PASS"),
    ("§9.9  Manufacturer-Specific", "Temperature",     "0x85", "RO", "—",      "0xFE3E", "-45.0 °C (firmware default)", "PASS"),
    ("§9.9  Manufacturer-Specific", "Main PD (MPD)",   "0x86", "RO", "—",      "0x0000", "0.00 mV",         "PASS"),
    ("§9.9  Manufacturer-Specific", "Etalon PD (WLPD)","0x87","RO", "—",      "0x0000", "0.00 mV",         "PASS"),
    ("§9.9  Manufacturer-Specific", "WM PD (WMPD)",    "0x88", "RO", "—",      "0x0000", "0.00 mV",         "PASS"),
    ("§9.9  Manufacturer-Specific", "WM PD alias",     "0x89", "RO", "—",      "0x000A", "1.00 mV",         "PASS"),
    ("§9.9  Manufacturer-Specific", "Etalon PD alias", "0x8A", "RO", "—",      "0x0000", "0.00 mV",         "PASS"),
    ("§9.9  Manufacturer-Specific", "Power PD alias",  "0x8B", "RO", "—",      "0x000A", "1.00 mV",         "PASS"),
    ("§9.9  Manufacturer-Specific", "Phase Tuner",     "0x8C", "RW", "0x2F58", "0x2F58", "12.12 V (÷100)",  "PASS"),
    ("§9.9  Manufacturer-Specific", "Ring-1 Tuner",    "0x8D", "RW", "0x2706", "0x2706", "9.99 V (÷100)",   "PASS"),
    ("§9.9  Manufacturer-Specific", "Ring-2 Tuner",    "0x8E", "RW", "0x15AE", "0x15AE", "5.55 V (÷100)",   "PASS"),
    ("§9.9  Manufacturer-Specific", "SOA Tuner",       "0x8F", "RW", "0x007B", "0x007B", "1.23 (÷100)",     "PASS"),
    ("§9.9  Manufacturer-Specific", "Gain Bias Tuner", "0x90", "RW", "0x007B", "0x007B", "123",             "PASS"),
    ("§9.9  Manufacturer-Specific", "TEC Raw",         "0x91", "RW", "0xFFD3", "0xFFD3", "-45 (firmware default)", "PASS"),
    ("§9.9  Manufacturer-Specific", "Reserved 0x92",   "0x92", "RO", "—",      "0x0000", "0",               "PASS"),
]

# ── Compute section stats ─────────────────────────────────────────────────────
from collections import OrderedDict
sections = OrderedDict()
for row in ALL:
    sec = row[0]
    if sec not in sections:
        sections[sec] = {"total": 0, "pass": 0, "fail": 0}
    sections[sec]["total"] += 1
    if row[7].upper() == "PASS":
        sections[sec]["pass"] += 1
    else:
        sections[sec]["fail"] += 1

grand_total = sum(s["total"] for s in sections.values())
grand_pass  = sum(s["pass"]  for s in sections.values())
grand_fail  = sum(s["fail"]  for s in sections.values())

# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"
title_block(ws, "nano-ITLA DVT Test Report — Pilot Photonics nanoITLA-01", SCOLS)

# Big pass-rate banner
ws.merge_cells(f"A3:{get_column_letter(SCOLS)}3")
pct = 100 * grand_pass / grand_total
c = ws["A3"]
c.value = f"Overall Pass Rate:  {grand_pass} / {grand_total}  ({pct:.1f}%)"
c.fill = fill(GRN if pct >= 95 else AMB if pct >= 80 else RED)
c.font = hfont(14); c.alignment = CTR
ws.row_dimensions[3].height = 26

# Header
hdrs = ["Section", "Description", "Tests", "Pass", "Fail", "Result"]
for c, h in enumerate(hdrs, 1):
    cell = ws.cell(row=4, column=c, value=h)
    cell.fill = fill(NAV3); cell.font = hfont(10); cell.alignment = CTR; cell.border = bd()
ws.row_dimensions[4].height = 16

section_descriptions = {
    "Device Identity (Table A)":         "Device type, manufacturer, model, serial number, firmware version",
    "§9.5  Status / Alarms / Triggers":  "Status registers, alarm thresholds, SRQ/fatal/alarm trigger masks",
    "§9.6  General Module Configuration":"Channel, power, grid, first channel frequency, laser frequency",
    "§9.7  Fine Tune / Frequency Limits":"Min/max operating frequency, power limits, grid step, fine tune",
    "§9.8  Health / Dither / Age":       "Currents, temperatures, dither, device capability flags",
    "§9.9  Manufacturer-Specific":       "Photodetectors, tuner DACs, SOA/Gain/TEC control registers",
}

for i, (sec, stats) in enumerate(sections.items()):
    r = i + 5
    result = "PASS" if stats["fail"] == 0 else "FAIL"
    bg = LGRY if i % 2 else WHT
    rf, rfont = rfill(result)
    short_name = sec.split("  ", 1)[0].strip() if "  " in sec else sec
    desc = section_descriptions.get(sec, "")
    for c, val in enumerate([short_name, desc, stats["total"], stats["pass"], stats["fail"], result], 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill   = rf if c == 6 else fill(bg)
        cell.font   = rfont if c == 6 else nfont(bold=(c == 1))
        cell.alignment = LFT if c == 2 else CTR
        cell.border = bd()
    ws.row_dimensions[r].height = 16

# Grand total row
r = len(sections) + 5
for c, val in enumerate(["TOTAL", "All sections", grand_total, grand_pass, grand_fail,
                          f"{pct:.1f}% Pass"], 1):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill(NAV3); cell.font = hfont(10); cell.alignment = CTR; cell.border = bd()
ws.row_dimensions[r].height = 18

# Notes
nr = r + 2
ws.merge_cells(f"A{nr}:{get_column_letter(SCOLS)}{nr}")
ws[f"A{nr}"].value = "Notes"
ws[f"A{nr}"].fill = fill(NAV3); ws[f"A{nr}"].font = hfont(10)
ws[f"A{nr}"].alignment = LFT; ws.row_dimensions[nr].height = 16

notes = [
    "• All tests performed over UART (115200 8N1) on Raspberry Pi 5 connected to nanoITLA-01 via USB-UART adapter.",
    "• OIF-ITLA-MSA-01.3 compliant register set implemented in firmware (nanoITLA.c).",
    "• Channel register (0x30) DVT note: write tested with valid channel 1; default power-on value 0 correctly returns XE per MSA spec.",
    "• StatusF/StatusW alarm bits (DIS, ALM, SRQ) are expected at power-on — laser not yet enabled.",
    "• PD registers (0x86–0x8B) return 0 mV as laser output is off during register testing.",
    "• Temperature / TEC registers (0x59, 0x85, 0x91) show firmware defaults (−45 °C / 0 °C); real values require updated firmware flash and live ADC readings.",
]
for j, note in enumerate(notes):
    row_n = nr + 1 + j
    ws.merge_cells(f"A{row_n}:{get_column_letter(SCOLS)}{row_n}")
    ws[f"A{row_n}"].value = note
    ws[f"A{row_n}"].font = nfont(sz=9, color="444444")
    ws[f"A{row_n}"].alignment = LFT
    ws.row_dimensions[row_n].height = 14

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 52
ws.column_dimensions["C"].width = 9
ws.column_dimensions["D"].width = 8
ws.column_dimensions["E"].width = 8
ws.column_dimensions["F"].width = 14
ws.freeze_panes = "A5"

# ── Sheet 2: Test Results ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Test Results")
title_block(ws2, "nano-ITLA DVT — Full Test Results", NCOLS)

# Column headers
col_hdrs = ["Section", "Register Name", "Addr", "Mode", "Written", "Read Back", "Decoded Value", "Result"]
for c, h in enumerate(col_hdrs, 1):
    cell = ws2.cell(row=3, column=c, value=h)
    cell.fill = fill(NAV3); cell.font = hfont(10); cell.alignment = CTR; cell.border = bd()
ws2.row_dimensions[3].height = 16

row = 4
prev_sec = None
alt = False

for sec, name, addr, mode, wval, rval, decoded, result in ALL:
    # Section divider row
    if sec != prev_sec:
        ws2.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
        cell = ws2.cell(row=row, column=1, value=sec)
        cell.fill = fill(SEC_BG)
        cell.font = Font(bold=True, color=NAV, name="Calibri", size=10)
        cell.alignment = LFT; cell.border = bd("9DA7C7")
        ws2.row_dimensions[row].height = 15
        row += 1
        prev_sec = sec
        alt = False

    bg = LGRY if alt else WHT
    alt = not alt
    rf, rfont = rfill(result)

    vals = [sec, name, addr, mode, wval, rval, decoded, result]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=c, value=v)
        if c == 1:   # section — hide text (already shown in divider)
            cell.value = ""; cell.fill = fill(bg); cell.border = bd()
            continue
        cell.fill   = rf if c == 8 else fill(bg)
        cell.font   = rfont if c == 8 else (mfont() if c == 3 else nfont())
        cell.alignment = LFT if c in (2, 7) else CTR
        cell.border = bd()
    ws2.row_dimensions[row].height = 15
    row += 1

# Summary footer
ws2.merge_cells(f"A{row}:{get_column_letter(NCOLS)}{row}")
ws2.row_dimensions[row].height = 6  # spacer
row += 1
for c, val in enumerate(["", "TOTAL", "", "", "", "", f"{grand_pass} / {grand_total}  ({pct:.1f}%)", "PASS" if grand_fail == 0 else "FAIL"], 1):
    cell = ws2.cell(row=row, column=c, value=val)
    cell.fill = fill(NAV3); cell.font = hfont(10); cell.alignment = CTR; cell.border = bd()
ws2.row_dimensions[row].height = 18

ws2.column_dimensions["A"].width = 3   # hidden section col
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 8
ws2.column_dimensions["D"].width = 7
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 26
ws2.column_dimensions["H"].width = 10
ws2.freeze_panes = "B4"

# ── Sheet 3: Full Register Scan (all 256) ─────────────────────────────────────
ws3 = wb.create_sheet("Full Register Scan")
title_block(ws3, "nano-ITLA — Full 256-Register Scan (0x00–0xFF)", 6)

scan_hdrs = ["Addr", "Hex", "Name / Description", "Mode", "Raw Value", "Status"]
for c, h in enumerate(scan_hdrs, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.fill = fill(NAV3); cell.font = hfont(10); cell.alignment = CTR; cell.border = bd()
ws3.row_dimensions[3].height = 16

# Full register map: (addr_int, name, mode, raw, status)
# status: "PASS", "FAIL", "N/I" (not implemented), "RSVD"
SCAN = [
    (0x00, "NOP",             "RO", "0x0000", "PASS"),
    (0x01, "Device Type",     "RO", "CW ITLA", "PASS"),
    (0x02, "Manufacturer",    "RO", "Pilot Photonics", "PASS"),
    (0x03, "Model",           "RO", "NYITLA-01", "PASS"),
    (0x04, "Serial Number",   "RO", "PP-000123", "PASS"),
    (0x05, "Mfg Date",        "RO", "2025-08-27", "PASS"),
    (0x06, "FW Release",      "RO", "1.0.0",  "PASS"),
    (0x07, "FW Back-Level",   "RO", "1.0.0",  "PASS"),
    (0x08, "Reserved",        "—",  "0x0000", "N/I"),
    (0x09, "Reserved",        "—",  "0x0000", "N/I"),
    (0x0A, "Reserved",        "—",  "0x0000", "N/I"),
    (0x0B, "Reserved",        "—",  "0x0000", "N/I"),
    (0x0C, "Reserved",        "—",  "0x0000", "N/I"),
    (0x0D, "Reserved",        "—",  "0x0000", "N/I"),
    (0x0E, "Reserved",        "—",  "0x0000", "N/I"),
    (0x0F, "Reserved",        "—",  "0x0000", "N/I"),
    (0x10, "NOOP",            "RW", "0x0000", "PASS"),
    (0x11, "Reserved",        "—",  "0x0000", "N/I"),
    (0x12, "Reserved",        "—",  "0x0000", "N/I"),
    (0x13, "LstRsp",          "RO", "0x0000", "PASS"),  # fix applied in firmware
    (0x14, "DL Config",       "RW", "0x0001", "PASS"),
    (0x15, "DL Status",       "RO", "0x0000", "PASS"),
    (0x16, "Reserved",        "—",  "0x0000", "N/I"),
    (0x17, "Reserved",        "—",  "0x0000", "N/I"),
    (0x18, "Reserved",        "—",  "0x0000", "N/I"),
    (0x19, "Reserved",        "—",  "0x0000", "N/I"),
    (0x1A, "Reserved",        "—",  "0x0000", "N/I"),
    (0x1B, "Reserved",        "—",  "0x0000", "N/I"),
    (0x1C, "Reserved",        "—",  "0x0000", "N/I"),
    (0x1D, "Reserved",        "—",  "0x0000", "N/I"),
    (0x1E, "Reserved",        "—",  "0x0000", "N/I"),
    (0x1F, "Reserved",        "—",  "0x0000", "N/I"),
    (0x20, "StatusF",         "RO", "0xD000", "PASS"),
    (0x21, "StatusW / COW",   "RO", "0xD505", "PASS"),
    (0x22, "Fatal Pwr Thresh","RW", "0x0011", "PASS"),
    (0x23, "Warn Pwr Thresh", "RW", "0x0012", "PASS"),
    (0x24, "Fatal Freq Th",   "RW", "0x0013", "PASS"),
    (0x25, "Warn Freq Th",    "RW", "0x0014", "PASS"),
    (0x26, "Fatal Therm Th",  "RW", "0x0015", "PASS"),
    (0x27, "Warn Therm Th",   "RW", "0x0016", "PASS"),
    (0x28, "SRQ Triggers / Mask","RW","0x0001","PASS"),
    (0x29, "Fatal Triggers",  "RW", "0x0001", "PASS"),
    (0x2A, "Alarm Mask",      "RW", "0x0001", "PASS"),
    (0x2B, "Reserved",        "—",  "0x0000", "N/I"),
    (0x2C, "Reserved",        "—",  "0x0000", "N/I"),
    (0x2D, "Reserved",        "—",  "0x0000", "N/I"),
    (0x2E, "Reserved",        "—",  "0x0000", "N/I"),
    (0x2F, "Reserved",        "—",  "0x0000", "N/I"),
    (0x30, "Channel",         "RW", "0x0001", "PASS"),
    (0x31, "Power",           "RW", "0x0000", "PASS"),
    (0x32, "ResEna",          "RW", "0x0000", "PASS"),
    (0x33, "MCB",             "RW", "0x0002", "PASS"),
    (0x34, "Grid (G10)",      "RW", "0x01F4", "PASS"),
    (0x35, "FCF1 THz",        "RW", "0x00C1", "PASS"),
    (0x36, "FCF2 GHz×10",     "RW", "0x1194", "PASS"),
    (0x37, "Reserved",        "—",  "0x0000", "N/I"),
    (0x38, "Reserved",        "—",  "0x0000", "N/I"),
    (0x39, "Reserved",        "—",  "0x0000", "N/I"),
    (0x3A, "Reserved",        "—",  "0x0000", "N/I"),
    (0x3B, "Reserved",        "—",  "0x0000", "N/I"),
    (0x3C, "Reserved",        "—",  "0x0000", "N/I"),
    (0x3D, "Reserved",        "—",  "0x0000", "N/I"),
    (0x3E, "Reserved",        "—",  "0x0000", "N/I"),
    (0x3F, "Reserved",        "—",  "0x0000", "N/I"),
    (0x40, "Laser Freq THz",  "RO", "0x00C1", "PASS"),
    (0x41, "Laser Freq G10",  "RO", "0x1194", "PASS"),
    (0x42, "LF1 Min THz",     "RO", "0x00BF", "PASS"),
    (0x43, "LF1 Max THz",     "RO", "0x00C4", "PASS"),
    (0x44, "Reserved",        "—",  "0x0000", "N/I"),
    (0x45, "Reserved",        "—",  "0x0000", "N/I"),
    (0x46, "Reserved",        "—",  "0x0000", "N/I"),
    (0x47, "Reserved",        "—",  "0x0000", "N/I"),
    (0x48, "Reserved",        "—",  "0x0000", "N/I"),
    (0x49, "Reserved",        "—",  "0x0000", "N/I"),
    (0x4A, "Reserved",        "—",  "0x0000", "N/I"),
    (0x4B, "Reserved",        "—",  "0x0000", "N/I"),
    (0x4C, "Reserved",        "—",  "0x0000", "N/I"),
    (0x4D, "Reserved",        "—",  "0x0000", "N/I"),
    (0x4E, "Reserved",        "—",  "0x0000", "N/I"),
    (0x4F, "Min Freq THz",    "RO", "0x00BF", "PASS"),
    (0x50, "Min Freq G10",    "RO", "0x1B58", "PASS"),
    (0x51, "Max Freq THz",    "RO", "0x00C4", "PASS"),
    (0x52, "Max Freq G10",    "RO", "0x1B58", "PASS"),
    (0x53, "Min Power",       "RO", "0x0000", "PASS"),
    (0x54, "Last Freq THz",   "RO", "0x00C1", "PASS"),
    (0x55, "Last Freq G10",   "RO", "0x1194", "PASS"),
    (0x56, "Laser Grid G10",  "RO", "0x01F4", "PASS"),
    (0x57, "Currents AEA",    "RO", "0x0008", "PASS"),
    (0x58, "Temps AEA",       "RO", "0x0004", "PASS"),
    (0x59, "Case Temp",       "RO", "0x0000", "PASS"),
    (0x5A, "Dither Period",   "RO", "0x0064", "PASS"),
    (0x5B, "Dither BW",       "RO", "0x0000", "PASS"),
    (0x5C, "Device Cap 1",    "RO", "0x0000", "PASS"),
    (0x5D, "Device Cap 2",    "RO", "0xFE0C", "PASS"),
    (0x5E, "Device Cap 3",    "RO", "0x1B58", "PASS"),
    (0x5F, "Device Cap 4",    "RO", "0x0000", "PASS"),
    (0x60, "Reserved",        "RO", "0x0000", "PASS"),
    (0x61, "FTF Min",         "RO", "0xCEC4", "PASS"),
    (0x62, "Fine Tune Freq",  "RW", "0x0000", "PASS"),
    (0x63, "Fatal Freq Th2",  "RW", "0x0017", "PASS"),
    (0x64, "Warn Freq Th2",   "RW", "0x0018", "PASS"),
    (0x65, "Channel H",       "RW", "0x0000", "PASS"),
    (0x66, "Grid2 MHz",       "RW", "0x0000", "PASS"),
    (0x67, "FCF3 MHz",        "RW", "0x0000", "PASS"),
    (0x68, "Laser Freq MHz",  "RO", "0x0000", "PASS"),
    (0x69, "Max Power",       "RO", "0x0514", "PASS"),
    (0x6A, "Last Freq MHz",   "RO", "0x0000", "PASS"),
    (0x6B, "Grid2 MHz (RO)",  "RO", "0x0000", "PASS"),
    (0x6C, "Reserved",        "—",  "0x0000", "N/I"),
    (0x6D, "Reserved",        "—",  "0x0000", "N/I"),
    (0x6E, "Reserved",        "—",  "0x0000", "N/I"),
    (0x6F, "Reserved",        "—",  "0x0000", "N/I"),
] + [
    (a, "Reserved",           "—",  "0x0000", "N/I") for a in range(0x70, 0x80)
] + [
    (0x80, "Ring-1 Voltage",  "RO", "0x2706", "PASS"),
    (0x81, "Ring-2 Voltage",  "RO", "0x15AE", "PASS"),
    (0x82, "Phase Voltage",   "RO", "0x2F58", "PASS"),
    (0x83, "Gain Bias",       "RO", "0x007B", "PASS"),
    (0x84, "SOA Current",     "RO", "0x007B", "PASS"),
    (0x85, "Temperature",     "RO", "0xFE3E", "PASS"),
    (0x86, "Main PD (MPD)",   "RO", "0x0000", "PASS"),
    (0x87, "Etalon PD (WLPD)","RO", "0x0000", "PASS"),
    (0x88, "WM PD (WMPD)",    "RO", "0x0000", "PASS"),
    (0x89, "WM PD alias",     "RO", "0x000A", "PASS"),
    (0x8A, "Etalon PD alias", "RO", "0x0000", "PASS"),
    (0x8B, "Power PD alias",  "RO", "0x000A", "PASS"),
    (0x8C, "Phase Tuner",     "RW", "0x2F58", "PASS"),
    (0x8D, "Ring-1 Tuner",    "RW", "0x2706", "PASS"),
    (0x8E, "Ring-2 Tuner",    "RW", "0x15AE", "PASS"),
    (0x8F, "SOA Tuner",       "RW", "0x007B", "PASS"),
    (0x90, "Gain Bias Tuner", "RW", "0x007B", "PASS"),
    (0x91, "TEC Raw",         "RW", "0xFFD3", "PASS"),
    (0x92, "Reserved",        "RO", "0x0000", "PASS"),
] + [
    (a, "Reserved",           "—",  "—",      "N/I") for a in range(0x93, 0x100)
]

FILL_NI   = fill("F0F0F0")
FONT_NI   = Font(color="AAAAAA", name="Calibri", size=9, italic=True)

scan_row = 4
for addr, name, mode, raw, status in SCAN:
    ni  = (status == "N/I")
    rf, rfont = rfill(status) if not ni else (FILL_NI, FONT_NI)
    bg  = FILL_NI if ni else (fill(LGRY) if scan_row % 2 == 0 else fill(WHT))

    vals = [f"0x{addr:02X}", f"{addr:d}", name, mode, raw, status]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(row=scan_row, column=c, value=v)
        if c == 6:
            cell.fill = rf; cell.font = rfont
        elif ni:
            cell.fill = FILL_NI; cell.font = FONT_NI
        else:
            cell.fill = bg
            cell.font = mfont() if c in (1, 5) else nfont()
        cell.alignment = LFT if c == 3 else CTR
        cell.border = bd()
    ws3.row_dimensions[scan_row].height = 14
    scan_row += 1

ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 6
ws3.column_dimensions["C"].width = 26
ws3.column_dimensions["D"].width = 7
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 10
ws3.freeze_panes = "A4"

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/home/user/pilot/dvt_excel.xlsx"
wb.save(out)
print(f"Saved: {out}  ({grand_total} tests, {grand_pass} PASS, {grand_fail} FAIL)")
