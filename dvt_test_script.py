#!/usr/bin/env python3
# dvt_test_script.py — Cross-platform DVT harness (Windows/Linux/macOS)

import logging
from logging.handlers import RotatingFileHandler
import subprocess
import sys
import os
import ctypes
from datetime import datetime
from pathlib import Path



import time
import math

# ===============================
# Strict validation (NO "magic")
# ===============================
# When enabled, we do NOT:
#   - apply any per-unit wavemeter offset/correction
#   - bias peak selection toward the expected wavelength
# We only measure what the instrument reports and compare directly to CSV.
STRICT_NO_MAGIC = True  # Always enforce direct measurement vs expected (no corrections/biasing)

# --- Optional: Web GUI (Flask) ---
try:
    from flask import Flask, jsonify, request
except Exception:
    Flask = None
    jsonify = None
    request = None
# ---------------- Web GUI (Flask) ----------------

def _require_flask():
    if Flask is None:
        raise RuntimeError(
            "Flask is not installed. Install with: python3 -m pip install flask"
        )

def _require_uart_open():
    if not IS_UART:
        # We allow ctypes mode too, but in the lab you likely want UART.
        return
    if ser is None:
        raise RuntimeError("UART is not open. Ensure ITLA_MODE=uart and the script opened /dev/ttyUSB0.")


def start_web_gui(host="0.0.0.0", port=5000):
    """Start a simple browser GUI on the Pi.

    Endpoints:
      GET  /                 -> HTML dashboard
      GET  /api/ping          -> basic health
      GET  /api/statusw       -> StatusW and BUSY bit
      POST /api/channel       -> set channel (json: {"channel":6})
      GET  /api/pd            -> read PD regs (0x89..0x8B and 0x86..0x88)
      GET  /api/reg/<reg>     -> read register (reg can be hex like 0x89 or decimal)
      POST /api/reg           -> write register (json: {"reg":"0x30","data":6})
      POST /api/run           -> run a test (json: {"name":"tableA"|"tableB"|"pd"|"fullscan"})

    Notes:
      - For long tests, the endpoint returns immediately with a message.
      - Logs/Excel continue to be produced by the existing harness.
    """
    _require_flask()
    _require_uart_open()

    app = Flask(__name__)

    DASH_HTML = """<!doctype html>
<html>
<head>
  <meta charset='utf-8'/>
  <meta name='viewport' content='width=device-width, initial-scale=1'/>
  <title>nITLA DVT GUI</title>
  <style>
    body{font-family:system-ui, -apple-system, Segoe UI, Roboto, Arial; margin:20px;}
    .row{display:flex; gap:16px; flex-wrap:wrap;}
    .card{border:1px solid #ddd; border-radius:10px; padding:14px; min-width:280px;}
    button{padding:10px 14px; border-radius:10px; border:1px solid #333; background:#111; color:#fff; cursor:pointer;}
    input{padding:10px; border-radius:10px; border:1px solid #ccc; width:120px;}
    pre{background:#0b0b0b; color:#d7ffd7; padding:12px; border-radius:10px; overflow:auto; max-height:260px;}
    .ok{color:#0a7f2e; font-weight:600;}
    .busy{color:#b45309; font-weight:700;}
  </style>
</head>
<body>
  <h2>nITLA DVT GUI</h2>
  <p>Mode: <b id='mode'></b> | UART: <b id='uart'></b> | StatusW: <b id='sw'></b></p>

  <div class='row'>
    <div class='card'>
      <h3>Set Channel</h3>
      <input id='ch' type='number' min='1' max='2000' value='6'/>
      <button onclick='setChannel()'>Set</button>
      <p id='busy' class='busy'></p>
    </div>

    <div class='card'>
      <h3>PD Registers</h3>
      <button onclick='refreshPD()'>Refresh</button>
      <pre id='pd'></pre>
    </div>

    <div class='card'>
      <h3>Register I/O</h3>
      <div>
        <input id='reg' placeholder='0x89' value='0x89'/>
        <button onclick='readReg()'>Read</button>
      </div>
      <div style='margin-top:10px;'>
        <input id='wreg' placeholder='0x30' value='0x30'/>
        <input id='wdata' placeholder='6' value='6'/>
        <button onclick='writeReg()'>Write</button>
      </div>
      <pre id='rio'></pre>
    </div>

    <div class='card'>
      <h3>Run Tests</h3>
      <button onclick='runTest("tableA")'>Table A</button>
      <button onclick='runTest("tableB")'>Table B Snapshot</button>
      <button onclick='runTest("pd")'>PD Test</button>
      <button onclick='runTest("fullscan")'>Full Scan</button>
      <pre id='run'></pre>
    </div>
  </div>

  <script>
    async function api(path, opts){
      const r = await fetch(path, Object.assign({headers:{'Content-Type':'application/json'}}, opts||{}));
      const t = await r.text();
      try { return JSON.parse(t); } catch(e){ return {raw:t}; }
    }

    async function refreshStatus(){
      const st = await api('/api/statusw');
      document.getElementById('mode').innerText = st.mode;
      document.getElementById('uart').innerText = st.uart;
      document.getElementById('sw').innerText = st.statusw_hex + ' (BUSY=' + (st.busy?1:0) + ')';
      document.getElementById('busy').innerText = st.busy ? 'BUSY… waiting to clear' : '';
    }

    async function setChannel(){
      const ch = parseInt(document.getElementById('ch').value||'0',10);
      const res = await api('/api/channel', {method:'POST', body: JSON.stringify({channel: ch})});
      document.getElementById('rio').innerText = JSON.stringify(res, null, 2);
      // poll busy until clear
      for (let i=0;i<30;i++){
        await refreshStatus();
        const st = await api('/api/statusw');
        if (!st.busy) break;
        await new Promise(r=>setTimeout(r, 200));
      }
      await refreshPD();
    }

    async function refreshPD(){
      const res = await api('/api/pd');
      document.getElementById('pd').innerText = JSON.stringify(res, null, 2);
    }

    async function readReg(){
      const reg = document.getElementById('reg').value.trim();
      const res = await api('/api/reg/' + encodeURIComponent(reg));
      document.getElementById('rio').innerText = JSON.stringify(res, null, 2);
    }

    async function writeReg(){
      const reg = document.getElementById('wreg').value.trim();
      const data = document.getElementById('wdata').value.trim();
      const res = await api('/api/reg', {method:'POST', body: JSON.stringify({reg, data})});
      document.getElementById('rio').innerText = JSON.stringify(res, null, 2);
    }

    async function runTest(name){
      const res = await api('/api/run', {method:'POST', body: JSON.stringify({name})});
      document.getElementById('run').innerText = JSON.stringify(res, null, 2);
    }

    refreshStatus();
    refreshPD();
    setInterval(refreshStatus, 1500);
  </script>
</body>
</html>"""

    def _parse_reg_str(s: str) -> int:
        return int(str(s).strip(), 0)

    @app.get("/")
    def home():
        return DASH_HTML

    @app.get("/api/ping")
    def ping():
        return jsonify({"ok": True, "mode": MODE, "uart": bool(IS_UART)})

    @app.get("/api/statusw")
    def statusw():
        try:
            ce, status, sw = reg_read(0x21)
            busy = (sw & 0x0002) != 0
            return jsonify({
                "mode": MODE,
                "uart": bool(IS_UART),
                "ce": ce,
                "status": status,
                "status_text": _status_name(status),
                "statusw": sw,
                "statusw_hex": f"0x{sw:04X}",
                "busy": bool(busy),
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.post("/api/channel")
    def set_channel():
        try:
            payload = request.get_json(force=True) or {}
            ch = int(payload.get("channel"))

            # MSA-correct: write ChannelH first, then Channel. Tuning starts on 0x30 write.
            w65, w30 = _set_channel_32(ch)
            ce65, st65, in65, out65 = w65
            ce30, st30, in30, out30 = w30

            tune_ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)

            return jsonify({
                "channel": ch,
                "write_0x65": {
                    "ce": ce65,
                    "status": st65,
                    "status_text": _status_name(st65),
                    "in": in65,
                    "in_hex": f"0x{in65:04X}",
                    "out": out65,
                    "out_hex": f"0x{out65:04X}",
                },
                "write_0x30": {
                    "ce": ce30,
                    "status": st30,
                    "status_text": _status_name(st30),
                    "in": in30,
                    "in_hex": f"0x{in30:04X}",
                    "out": out30,
                    "out_hex": f"0x{out30:04X}",
                },
                "tune_complete": bool(tune_ok),
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @app.get("/api/pd")
    def pd():
        try:
            def _read_scaled(reg, scale=10.0):
                ce, st, dout = reg_read(reg)
                return {
                    "reg": f"0x{reg:02X}",
                    "ce": ce,
                    "status": st,
                    "status_text": _status_name(st),
                    "raw": dout,
                    "raw_hex": f"0x{dout:04X}",
                    "value": float(dout) / float(scale),
                    "scale": scale,
                }

            return jsonify({
                "WM_PD": _read_scaled(0x89, 10.0),
                "ETALON_PD": _read_scaled(0x8A, 10.0),
                "Power_PD": _read_scaled(0x8B, 10.0),
                "MPD": _read_scaled(0x86, 10.0),
                "WLPD": _read_scaled(0x87, 10.0),
                "WMPD": _read_scaled(0x88, 10.0),
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.get("/api/reg/<path:reg_s>")
    def api_reg_read(reg_s):
        try:
            reg = _parse_reg_str(reg_s)
            if not (0 <= reg <= 0xFF):
                return jsonify({"error": "reg out of range 0..0xFF"}), 400
            ce, st, dout = reg_read(reg)
            return jsonify({
                "reg": f"0x{reg:02X}",
                "ce": ce,
                "status": st,
                "status_text": _status_name(st),
                "out": dout,
                "out_hex": f"0x{dout:04X}",
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @app.post("/api/reg")
    def api_reg_write():
        try:
            payload = request.get_json(force=True) or {}
            reg = _parse_reg_str(payload.get("reg"))
            data = int(str(payload.get("data")).strip(), 0)
            if not (0 <= reg <= 0xFF):
                return jsonify({"error": "reg out of range 0..0xFF"}), 400
            if not (0 <= data <= 0xFFFF):
                return jsonify({"error": "data out of range 0..0xFFFF"}), 400
            ce, st, din, dout = reg_write(reg, data)
            return jsonify({
                "reg": f"0x{reg:02X}",
                "in": din,
                "in_hex": f"0x{din:04X}",
                "out": dout,
                "out_hex": f"0x{dout:04X}",
                "ce": ce,
                "status": st,
                "status_text": _status_name(st),
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 400

    @app.post("/api/run")
    def api_run():
        try:
            payload = request.get_json(force=True) or {}
            name = str(payload.get("name") or "").strip().lower()

            # Run synchronously for short tests; return a helpful message.
            if name == "tablea":
                run_supervisory_table()
                return jsonify({"started": True, "name": "tableA", "note": "Completed. Check Excel/log."})
            if name in ("tableb", "snapshot", "tableb_snapshot"):
                run_9_6_snapshot_table()
                return jsonify({"started": True, "name": "tableB_snapshot", "note": "Completed. Check Excel/log."})
            if name == "pd":
                run_pd_test()
                return jsonify({"started": True, "name": "pd", "note": "Completed. Check Excel/log."})
            if name == "fullscan":
                # Full scan is long; still run, but warn.
                run_full_register_scan()
                return jsonify({"started": True, "name": "fullscan", "note": "Completed. Check Excel/log."})

            return jsonify({"error": "Unknown test name. Use tableA, tableB, pd, fullscan."}), 400
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    print(f"[INFO] Web GUI starting on http://{host}:{port}")
    print("[INFO] Open in a browser on the same network, e.g. http://raspberrypi.local:5000")
    app.run(host=host, port=int(port), debug=False)


# --- Optional: readPeaks import for OSA test ---
READPEAKS_IMPORT_ERROR = None
READPEAKS_SNAPSHOT_IMPORT_ERROR = None
try:
    from readPeaks import readPeaks
except Exception as e:
    readPeaks = None
    READPEAKS_IMPORT_ERROR = str(e)

# Optional: some lab versions provide a snapshot function that returns ALL peaks
# from a single acquisition (preferred for debug). If missing, we fall back to
# logging whatever the primary readPeaks() returns.
try:
    from readPeaks import readPeaks_snapshot  # type: ignore
except Exception as e:
    readPeaks_snapshot = None
    READPEAKS_SNAPSHOT_IMPORT_ERROR = str(e)

# --- Optional: direct GPIB wavemeter backend (HP/Agilent) ---
# This matches the class shown by Chris (returns wavelength_nm, power_dBm, smsr_dB).
try:
    import Gpib  # linux-gpib python bindings
except Exception:
    Gpib = None

import re as _re

class HP_wavemeter(object):
    """HP/Agilent 86120C (or compatible) wavemeter over linux-gpib.

    Usage (linux-gpib):
      wm = HP_wavemeter(gpib_addr=18, board=0)

    Optional VISA-style string convenience:
      wm = HP_wavemeter(gpib="GPIB0::18::INSTR")
    """

    def __init__(self, gpib_addr=18, board=0, gpib=None):
        if Gpib is None:
            raise ImportError("Gpib module not available (install linux-gpib python bindings)")

        # Allow VISA-like strings e.g. "GPIB0::18::INSTR"
        if gpib is not None:
            s = str(gpib)
            m = _re.search(r"GPIB(\d+)::(\d+)::INSTR", s, flags=_re.IGNORECASE)
            if m:
                board = int(m.group(1))
                gpib_addr = int(m.group(2))
            else:
                raise ValueError(f"Unsupported GPIB string: {gpib!r}. Expected like 'GPIB0::18::INSTR'")

        self.board = int(board)
        self.gpib_addr = int(gpib_addr)

        print(f"[INFO] Trying to connect to wavemeter GPIB addr {self.gpib_addr} on board {self.board}")
        try:
            self.dev = Gpib.Gpib(self.board, self.gpib_addr)
        except Exception as e:
            raise RuntimeError(
                "Failed to open GPIB device. "
                f"board={self.board} addr={self.gpib_addr}. "
                "Check linux-gpib is configured and the instrument/interface is present (e.g., /dev/gpib0, ibtest/gpib_config). "
                f"Details: {e}"
            )

        # Basic link check: many failures surface on the first write (ENODEV / errno 19)
        try:
            time.sleep(0.05)
            self._write("*CLS")
            # Leave continuous OFF by default for deterministic snapshots.
            # We'll explicitly trigger a measurement before each read.
            self.continuous_sweep_off()
        except Exception as e:
            raise RuntimeError(
                "GPIB link check failed while talking to the wavemeter. "
                f"board={self.board} addr={self.gpib_addr}. "
                "Most common causes: wrong GPIB address, missing/incorrect linux-gpib configuration, or the USB-GPIB interface/instrument is not connected/powered. "
                f"Details: {e}"
            )

    def _write(self, cmd: str):
        """Write a SCPI command, ensuring it is properly terminated."""
        s = str(cmd)
        if not s.endswith("\n"):
            s += "\n"
        try:
            self.dev.write(s)
        except Exception as e:
            raise RuntimeError(f"GPIB write failed for command {cmd!r}: {e}")

    def _read_text(self, n=1024) -> str:
        """Read from device and always return a text string."""
        raw = self.dev.read(int(n))
        if isinstance(raw, bytes):
            return raw.decode("utf-8", errors="ignore")
        return str(raw)

    def _query(self, cmd: str, n: int = 2048) -> str:
        """Write a query command and read the response."""
        self._write(cmd)
        return self._read_text(n).strip()

    def _init_and_wait(self):
        """Trigger a new measurement sweep and wait for completion."""
      # Many 86120C setups require an explicit INIT to refresh the measurement buffers.
      # Best-effort: INIT + *WAI (wait) to ensure ARR/SCAL data updates.
        try:
            self._write(":INIT")
            self._write("*WAI")
        except Exception:
            pass
      
    def idn(self) -> str:
        """Return *IDN? response (quick connectivity check)."""
        self._write("*IDN?")
        return self._read_text(256).strip()

    def continuous_sweep_on(self):
        self._write(":INITiate:CONTinuous ON")

    def continuous_sweep_off(self):
        self._write(":INIT:CONT OFF")

    def read_wavelength_nm_max(self) -> float:
        """Return peak/max wavelength in nm using scalar query.

        Note: Some 86120C firmwares support :MEAS:SCAL:WAV?; others use the
        older :MEAS:SCAL:POW:WAV? form. We try the more standard command first,
        and fall back if needed.
        """
        self._init_and_wait()
        # Try standard scalar wavelength query first
        try:
            self._write(":MEAS:SCAL:WAV? MAX")
            txt = self._read_text(256).strip()
            val_m = float(txt)
            return val_m * 1e9
        except Exception:
            # Fallback used in some examples
            self._write(":MEAS:SCAL:POW:WAV? MAX")
            txt = self._read_text(256).strip()
            val_m = float(txt)
            return val_m * 1e9

    # Backwards-compatible helpers (Chris' one-liners)
    def read_wavelength(self) -> float:
        """Return a single-shot wavelength in nm.

        This is a compatibility alias used by older scripts/one-liners.
        We do **not** apply any offsets/corrections. We simply return what the
        instrument reports for the MAX/peak wavelength.
        """
        return float(self.read_wavelength_nm_max())

    def read_wavelength_nm(self) -> float:
        """Alias for read_wavelength(), explicit nm naming."""
        return float(self.read_wavelength_nm_max())

    def read_power_dbm_max(self) -> float:
        """Return peak/max power in dBm using scalar query.

        Instruments vary: some return watts, some return dBm. We:
          - read the raw scalar value
          - if it looks like watts (>~1e-6), convert W->dBm
          - if it looks like dBm (typical range -120..+30), return as-is
        """
        self._init_and_wait()
        self._write(":MEAS:SCAL:POW? MAX")
        txt = self._read_text(256).strip()
        raw = float(txt)

        # Heuristic: if within typical dBm range, assume it's already dBm
        if -200.0 <= raw <= 50.0:
            return raw

        # Otherwise assume watts and convert
        val_w = raw
        if val_w <= 0:
            return -99.0
        return 10.0 * math.log10(val_w / 1e-3)

    def read_peaks(self, target_wl_nm: float | None = None):
        """Return (peak_wavelength_nm, peak_power_dBm, smsr_dB).

        The 86120C can return multiple peaks. There are two common use-cases:
          1) Pick the **highest-power** peak (default behaviour)
          2) Pick the peak **closest to a target wavelength** (useful when spurious peaks exist)

        If `target_wl_nm` is provided, we pick the wavelength closest to that target.
        """
        self._init_and_wait()

        # Enforce STRICT_NO_MAGIC: never bias selection toward an expected/target wavelength.
        if STRICT_NO_MAGIC:
            target_wl_nm = None

        # Query power array (count, p1, p2, ...)
        power_txt = self._query("MEAS:ARR:POW?", 4096)
        power_parts = [p for p in _re.split(",|\n", power_txt) if p != ""]

      # Query wavelength array corresponding to the same peaks (count, w1, w2, ...)
        wav_txt = self._query("FETCh:ARR:POW:WAV?", 4096)
        wav_parts = [w for w in _re.split(",|\n", wav_txt) if w != ""]

        if not power_parts or not wav_parts:
            return 0.0, -99.0, 0.0

        try:
            n = int(float(power_parts[0]))
        except Exception:
            n = 0

        if n <= 0:
            return 0.0, -99.0, 0.0

        # Defensive: make sure we have enough entries
        n = min(n, len(power_parts) - 1, len(wav_parts) - 1)
        if n <= 0:
            return 0.0, -99.0, 0.0

        # Parse powers
        powers = []
        for i in range(1, n + 1):
            try:
                powers.append(float(power_parts[i]))
            except Exception:
                powers.append(-99.0)

      # Parse wavelengths (meters -> nm)
        wls_nm = []
        for i in range(1, n + 1):
            try:
                wls_nm.append(float(wav_parts[i]) * 1e9)
            except Exception:
                wls_nm.append(0.0)

      # Choose index
        if target_wl_nm is not None and target_wl_nm > 0:
          # pick closest wavelength to the target
            tgt = float(target_wl_nm)
            max_i0 = min(range(len(wls_nm)), key=lambda i: abs(wls_nm[i] - tgt))
        else:
          # pick maximum-power peak
            max_i0 = max(range(len(powers)), key=lambda i: powers[i])

        peak_pw = float(powers[max_i0])
        peak_wl_nm = float(wls_nm[max_i0])

      # SMSR estimate: difference between top and runner-up powers
        if len(powers) >= 2:
            sorted_pw = sorted(powers, reverse=True)
            smsr = float(sorted_pw[0]) - float(sorted_pw[1])
        else:
            smsr = 40.0

        return peak_wl_nm, peak_pw, smsr


    def read_peaks_snapshot(self):
        """Take ONE wavemeter sweep and return all peaks from the same snapshot.

        Returns:
            list[dict]: [{"wl_nm": float, "pwr_dbm": float}, ...]
        """
        self._init_and_wait()

        p_txt = self._query("MEAS:ARR:POW?", 4096)
        w_txt = self._query("FETCh:ARR:POW:WAV?", 4096)

        p = [x for x in _re.split(",|\n", p_txt) if x != ""]
        w = [x for x in _re.split(",|\n", w_txt) if x != ""]

        if not p or not w:
            return []

        try:
            n = int(float(p[0]))
        except Exception:
            return []

        n = min(n, len(p) - 1, len(w) - 1)
        peaks = []

        def _to_dbm(raw: float) -> float:
          # If the instrument already returns dBm, it is typically within this range.
            if -200.0 <= raw <= 50.0:
                return raw
          # Otherwise assume watts and convert W -> dBm.
            if raw <= 0:
                return -99.0
            return 10.0 * math.log10(raw / 1e-3)

        for i in range(1, n + 1):
            try:
                wl_nm = float(w[i]) * 1e9
                p_raw = float(p[i])
                peaks.append({
                    "wl_nm": wl_nm,
                    "pwr_dbm": _to_dbm(p_raw),
                })
            except Exception:
                pass

        return peaks

# ---------------- HP wavemeter singleton (avoid reconnect per channel) ----------------
_HP_WM_SINGLETON = None

def _get_hp_wavemeter():
    """Return a cached HP_wavemeter instance (GPIB) if available.

    Accepts multiple env var spellings so lab runs don't accidentally fall back to addr=18.

    String forms (preferred):
      - HP_WM_GPIB_STR
      - HP_WM_GPIB
      - WM_GPIB
      - WAVEMETER_GPIB

    Numeric forms:
      - HP_WM_GPIB_ADDR
      - WM_GPIB_ADDR
      - WAVEMETER_GPIB_ADDR

    Board:
      - HP_WM_GPIB_BOARD
      - WM_GPIB_BOARD
      - WAVEMETER_GPIB_BOARD

    Defaults: addr=4, board=0
    """
    global _HP_WM_SINGLETON
    if _HP_WM_SINGLETON is not None:
        return _HP_WM_SINGLETON

    # 1) Prefer explicit VISA-like strings
    for k in ("HP_WM_GPIB_STR", "HP_WM_GPIB", "WM_GPIB", "WAVEMETER_GPIB"):
        gpib_str = (os.environ.get(k, "") or "").strip()
        if gpib_str:
            _HP_WM_SINGLETON = HP_wavemeter(gpib=gpib_str)
            return _HP_WM_SINGLETON

    # 2) Numeric address + board fallbacks
    def _first_int(keys, default: int) -> int:
        for kk in keys:
            vv = (os.environ.get(kk, "") or "").strip()
            if vv:
                try:
                    return int(vv, 0)
                except Exception:
                    pass
        return int(default)

    addr = _first_int(("HP_WM_GPIB_ADDR", "WM_GPIB_ADDR", "WAVEMETER_GPIB_ADDR"), 4)
    board = _first_int(("HP_WM_GPIB_BOARD", "WM_GPIB_BOARD", "WAVEMETER_GPIB_BOARD"), 0)

    _HP_WM_SINGLETON = HP_wavemeter(gpib_addr=addr, board=board)
    return _HP_WM_SINGLETON

def _wm_read_peaks_stable(wm: "HP_wavemeter", discard_first: bool = True, settle_s: float = 0.0):
    """Read wavemeter peaks in a stable, spec-auditable way.

    Rationale:
      - After a SetChannel/tune event, the first sweep/buffer can be stale or transient.
      - We discard ONE measurement (no bias/offset), then take the next measurement.

    This is NOT "magic" or target-biasing; we still select the max-power peak.

    Returns:
        (wl_nm, p_dbm, smsr_db)
    """
    if discard_first:
        try:
            wm.read_peaks(target_wl_nm=None)  # throw-away sweep
        except Exception:
            # Even if the throw-away read fails, proceed to the real read.
            pass
        if float(settle_s) > 0:
            time.sleep(float(settle_s))

    return wm.read_peaks(target_wl_nm=None)
def _nm_to_ghz(wl_nm: float) -> float:
    # c (nm*THz) = 299792.458
    # GHz = THz * 1000
    if wl_nm <= 0:
        return 0.0
    return (299792.458 * 1000.0) / wl_nm

def _ghz_to_nm(f_ghz: float) -> float:
    # c (nm*THz) = 299792.458 ; GHz = THz*1000
    if f_ghz <= 0:
        return 0.0
    return (299792.458 * 1000.0) / f_ghz

import csv as _csv

def _norm_header(h: str) -> str:
    return " ".join(str(h or "").strip().lower().replace("_", " ").split())


def load_expected_wavelengths_csv(csv_path: str) -> dict:
    """Load expected wavelength (nm) per channel from a CSV.

    The loader is robust to header naming. It will try to detect:
      - channel column: Channel, CH, channel
      - wavelength column: Real Wavelength, Real_Wavelength, Wavelength, wl_nm, etc.

    Returns:
        dict[int, float]: {channel:int -> expected_wl_nm:float}
    """
    path = str(csv_path or "").strip()
    if not path:
        raise ValueError("CSV path is empty")
    if not os.path.exists(path):
        raise FileNotFoundError(f"CSV not found: {path}")

    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        # Sniff delimiter using sample
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t")
        except Exception:
            dialect = _csv.get_dialect("excel")
        reader = _csv.DictReader(f, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row")

        # Normalize headers
        hdr_map = { _norm_header(h): h for h in reader.fieldnames }

        # Channel header candidates
        ch_key = None
        for cand in ("channel", "ch", "channel number", "chn"):
            if cand in hdr_map:
                ch_key = hdr_map[cand]
                break

        # Wavelength header candidates
        wl_key = None
        for cand in (
            "real wavelength",
            "real wavelength nm",
            "real wl nm",
            "real wl",
            "wavelength",
            "wavelength nm",
            "wl nm",
            "wl_nm",
            "real_wavelength",
        ):
            if _norm_header(cand) in hdr_map:
                wl_key = hdr_map[_norm_header(cand)]
                break

        # If not found, try fuzzy contains-match
        if ch_key is None:
            for nh, orig in hdr_map.items():
                if "channel" in nh or nh == "ch":
                    ch_key = orig
                    break
        if wl_key is None:
            for nh, orig in hdr_map.items():
                if "wavelength" in nh and ("real" in nh or "nm" in nh or nh == "wavelength"):
                    wl_key = orig
                    break

        if ch_key is None or wl_key is None:
            raise ValueError(
                "Could not detect required columns in CSV. "
                f"Found headers: {reader.fieldnames}. "
                "Need a Channel/CH column and a Real Wavelength (nm) column."
            )

        out = {}
        for row in reader:
            if not row:
                continue
            ch_raw = (row.get(ch_key) or "").strip()
            wl_raw = (row.get(wl_key) or "").strip()
            if not ch_raw or not wl_raw:
                continue
            try:
                ch = int(float(ch_raw))
                wl_nm = float(wl_raw)
            except Exception:
                continue
            if ch <= 0:
                continue
            out[ch] = wl_nm

        if not out:
            raise ValueError("No usable channel/wavelength rows found in CSV")
        return out


def _get_expected_map() -> dict:
    """Load expected wavelengths from CSV path provided via env var.

    Env vars:
      - LUT_CSV_PATH: path to your exported LUT CSV containing Real Wavelength.
    """
    csv_path = os.environ.get("LUT_CSV_PATH", "").strip()
    if not csv_path:
        raise RuntimeError(
            "LUT_CSV_PATH is not set. Example:\n"
            "  export LUT_CSV_PATH=/home/pilot/nITLA_test/Unit_1_50GHz_LUT.csv"
        )
    return load_expected_wavelengths_csv(csv_path)

# ---------------- LUT CSV loader (full row) ----------------

def _norm_col(s: str) -> str:
    return str(s or "").strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def load_lut_rows_csv(csv_path: str) -> dict:
    """Load full LUT rows keyed by channel.

    Supports two CSV styles:
      A) CSV has an explicit Channel/CH column -> rows keyed by that channel.
      B) CSV has NO Channel column (like your Debug_unit2_50GHz_SOA200mA_LUT.csv)
         -> rows are assigned sequential channels starting at 1 for **each physical row**.

    Usable row rule:
      - Validated is TRUE (if column exists) AND
      - Real Wavelength (nm) > 0 (or Real Frequency > 0)

    IMPORTANT:
      - Even if a row is not usable (e.g., all zeros), we still KEEP it under its
        sequential channel number and mark it as usable=False. This prevents the
        channel numbering from shifting when some rows are empty/missing.

    Returns:
      dict[int, dict] where inner dict contains normalized keys:
        ch, v1, v2, v3, gain, soa, temp,
        pd1, pd2, pd3, real_wl_nm, real_freq_thz, validated
    """
    csv_path = str(csv_path or "").strip()
    if not csv_path:
        raise RuntimeError("CSV path is empty")

    with open(csv_path, "r", newline="", encoding="utf-8-sig") as f:
        # Sniff delimiter using sample
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t")
        except Exception:
            dialect = _csv.get_dialect("excel")
        reader = _csv.DictReader(f, dialect=dialect)
        # Safety check for semicolon files not sniffed correctly
        if reader.fieldnames and len(reader.fieldnames) == 1 and ";" in reader.fieldnames[0]:
            f.seek(0)
            reader = _csv.DictReader(f, delimiter=';')
        if not reader.fieldnames:
            raise ValueError("CSV has no header row")

        # Map raw headers -> normalized
        hdr_map = {h: _norm_col(h) for h in reader.fieldnames}

        # Detect whether CSV provides an explicit channel column
        has_channel_col = any(norm in ("channel", "ch") for norm in hdr_map.values())
        seq_ch = 0  # sequential channel assignment when no channel column exists (physical row index)

        def _truthy(v) -> bool:
            s = str(v or "").strip().lower()
            return s in ("1", "true", "yes", "y", "t")

        def _get(row, *names, default=""):
            # names are normalized forms
            for raw_h, norm_h in hdr_map.items():
                if norm_h in names:
                    return row.get(raw_h, default)
            return default

        out = {}
        for row in reader:
            # Channel selection: explicit column if present, otherwise sequential.
            ch = None

            if has_channel_col:
                ch_raw = _get(row, "channel", "ch")
                if not str(ch_raw or "").strip():
                    continue
                try:
                    ch = int(float(str(ch_raw).strip()))
                except Exception:
                    continue
                if ch <= 0:
                    continue
            else:
                # IMPORTANT: assign channel by physical row index so numbering never shifts
                seq_ch += 1
                ch = seq_ch

            def _f(v, default=0.0):
                try:
                    if v is None:
                        return float(default)
                    s = str(v).strip()
                    if s == "":
                        return float(default)
                    return float(s)
                except Exception:
                    return float(default)

            # V1/V2/V3 sometimes labelled B1/B2/B3 in emails
            v1 = _f(_get(row, "v1", "b1"))
            v2 = _f(_get(row, "v2", "b2"))
            v3 = _f(_get(row, "v3", "b3"))

            # CSV headers often include units, e.g. "Gain (mA)", "SOA (mA)", "Temperature (degC)"
            gain = _f(_get(row, "gain", "gain(ma)", "gainma"))
            soa  = _f(_get(row, "soa", "soa(ma)", "soama"))
            temp = _f(_get(row, "temp", "temperature", "temperature(degc)", "temperaturedegc"))

            # Expected real frequency / wavelength
            real_freq_thz = _f(_get(row, "realfrequency(thz)", "realfrequencythz", "realfrequency", "realfreqthz"))
            real_wl_nm = _f(_get(row, "realwavelength(nm)", "realwavelengthnm", "realwavelength", "realwl(nm)", "realwlnm", "wavelength(nm)", "wavelengthnm", "wavelength"))

            # PD columns in your CSV are PD1/PD2/PD3
            pd1 = _f(_get(row, "pd1"))
            pd2 = _f(_get(row, "pd2"))
            pd3 = _f(_get(row, "pd3"))

            # Validated column (optional)
            validated = _truthy(_get(row, "validated")) if any("validated" in n for n in hdr_map.values()) else True

            # Decide whether this row is usable (empty operating points stay mapped but marked unusable)
            usable = True
            if not validated:
                usable = False
            if real_wl_nm <= 0.0 and real_freq_thz <= 0.0:
                usable = False

            out[ch] = {
                "ch": ch,
                "v1": v1,
                "v2": v2,
                "v3": v3,
                "gain": gain,
                "soa": soa,
                "temp": temp,
                "pd1": pd1,
                "pd2": pd2,
                "pd3": pd3,
                "real_wl_nm": real_wl_nm,
                "real_freq_thz": real_freq_thz,
                "validated": bool(validated),
                "usable": bool(usable),
            }

        if not out:
            raise ValueError("No LUT rows found in CSV")
        return out


def _get_lut_rows() -> dict:
    """Load LUT rows from LUT_CSV_PATH."""
    csv_path = os.environ.get("LUT_CSV_PATH", "").strip()
    if not csv_path:
        raise RuntimeError(
            "LUT_CSV_PATH is not set. Example:\n"
            "  export LUT_CSV_PATH=/home/pilot/nITLA_test/Unit_1_50GHz_LUT.csv"
        )
    return load_lut_rows_csv(csv_path)

def _lut_operating_point_missing(lut_row: dict | None) -> bool:
    """Return True when the LUT row represents a missing operating point.

    Per lab note: some LUT rows are all zeros (operating points not found).
    The explicit rule given: if Gain and SOA are 0, do not test; auto FAIL.

    We implement:
      - Missing if row is None
      - Missing if (gain==0 and soa==0)
      - Missing if ALL key numeric fields are 0
    """
    if not lut_row:
        return True
    try:
        gain = float(lut_row.get("gain", 0.0) or 0.0)
        soa  = float(lut_row.get("soa", 0.0) or 0.0)
    except Exception:
        gain, soa = 0.0, 0.0

    if gain == 0.0 and soa == 0.0:
        return True

    # All-zeros fallback (robust)
    keys = ("v1", "v2", "v3", "gain", "soa", "temp", "pd1", "pd2", "pd3", "real_wl_nm", "real_freq_thz")
    vals = []
    for k in keys:
        try:
            vals.append(float(lut_row.get(k, 0.0) or 0.0))
        except Exception:
            vals.append(0.0)
    return all(v == 0.0 for v in vals)
    
def _ensure_sheet_wm_compare():
    """Ensure the WM_Compare_CSV sheet exists with a custom header."""
    ws = get_sheet("WM_Compare_CSV")
    # If sheet is new, get_sheet will have written the default EXCEL_HEADERS. Replace with our custom header.
    if ws.max_row == 1 and (ws.cell(1, 1).value == EXCEL_HEADERS[0]):
        ws.delete_rows(1, 1)
        ws.append([
            "CH",
            "Expected_WL_nm",
            "Measured_WL_nm",
            "Expected_GHz",
            "Measured_GHz",
            "Err_MHz",
            "Tol_MHz",
            "Power_dBm",
            "SMSR_dB",
            "Result",
            "Tester",
            "DateTime",
        ])
        _excel_save()
    return ws

def _coerce_readpeaks_to_ghz(ret) -> float:
    """Accepts several possible readPeaks/read_peaks return shapes and converts to GHz.

    Supported:
      - float/int: assumed already in GHz
      - (wl_nm, pw_dbm, smsr): convert wl_nm -> GHz
      - dict with 'ghz' or 'wl_nm'
    """
    if ret is None:
        raise ValueError("readPeaks returned None")

    # Simple number
    if isinstance(ret, (int, float)):
        return float(ret)

    # Tuple/list
    if isinstance(ret, (tuple, list)):
        if len(ret) >= 1 and isinstance(ret[0], (int, float)):
            # Heuristic: if first item looks like wavelength in nm (1000..2000), convert.
            first = float(ret[0])
            if 1000.0 <= first <= 2000.0:
                return _nm_to_ghz(first)
            # Otherwise assume it's already GHz
            return first
        raise ValueError(f"Unsupported readPeaks tuple/list: {ret}")

    # Dict
    if isinstance(ret, dict):
        if 'ghz' in ret:
            return float(ret['ghz'])
        if 'wl_nm' in ret:
            return _nm_to_ghz(float(ret['wl_nm']))

    raise ValueError(f"Unsupported readPeaks return type: {type(ret)}")

def _format_peak_snapshot(peaks, selected_wl_nm: float | None = None, max_n: int = 5) -> str:
    """Return a compact, single-line string describing the peak snapshot.

    `peaks` may be:
      - list[dict] with keys like wl_nm / pwr_dbm
      - list/tuple of tuples
      - dict containing a 'peaks' list

    We keep this compact because it gets appended into the Notes column.
    """
    try:
        # If dict with embedded peaks
        if isinstance(peaks, dict) and "peaks" in peaks:
            peaks = peaks.get("peaks")

        items = []
        if isinstance(peaks, (list, tuple)):
            for p in peaks:
                if isinstance(p, dict):
                    wl = p.get("wl_nm")
                    pw = p.get("pwr_dbm")
                    if wl is None and "wavelength_nm" in p:
                        wl = p.get("wavelength_nm")
                    if pw is None and "power_dbm" in p:
                        pw = p.get("power_dbm")
                    if wl is None:
                        continue
                    items.append((float(wl), float(pw) if pw is not None else None))
                elif isinstance(p, (list, tuple)) and len(p) >= 1:
                    wl = float(p[0])
                    pw = float(p[1]) if len(p) >= 2 and p[1] is not None else None
                    items.append((wl, pw))

        # Sort by power desc when available
        if items and any(pw is not None for _, pw in items):
            items.sort(key=lambda t: (-1e9 if t[1] is None else -t[1]))

        items = items[:max_n]
        parts = []
        for wl, pw in items:
            if pw is None:
                parts.append(f"{wl:.3f}nm")
            else:
                parts.append(f"{wl:.3f}nm@{pw:.2f}dBm")

        sel = ""
        if selected_wl_nm is not None and selected_wl_nm > 0:
            sel = f" sel={float(selected_wl_nm):.3f}nm"

        if not parts:
            return "peaks=[]" + sel
        return "peaks=[" + ",".join(parts) + "]" + sel
    except Exception as e:
        return f"peaks=<snapshot_error:{e}>"
    
def _env_int(name: str, default: int) -> int:
    s = os.environ.get(name)
    if not s:
        return default
    try:
        return int(s, 0)
    except Exception:
        return default

def _env_float(name: str, default: float) -> float:
    s = os.environ.get(name)
    if not s:
        return float(default)
    try:
        return float(s)
    except Exception:
        return float(default)

def _env_bool(name: str, default: bool = False) -> bool:
    s = os.environ.get(name)
    if s is None or s == "":
        return bool(default)
    return str(s).strip().lower() in ("1", "true", "yes", "y", "on")

#
# Lab update: external WM power correction
# Option 9 fix: Use +10 dB correction, not +6 dB
#ATTENUATOR_DB = 6.0
# Lab update: external WM power correction
EXT_WM_POWER_OFFSET_DB = _env_float("EXT_WM_POWER_OFFSET_DB", 6.0)

# Limit channel sweep based on LUT size
MAX_LUT_CHANNELS = _env_int("MAX_LUT_CHANNELS", 52)
def _set_channel_32(channel: int):
    """MSA-correct 32-bit SetChannel.

    Per OIF-ITLA MSA §9.6.1: write ChannelH (0x65) first, then Channel (0x30). Tuning starts on the write to 0x30.

    Returns:
        tuple: (w65, w30) where each is (ce, status, din, dout)
    """
    ch = int(channel)
    ch_h = (ch >> 16) & 0xFFFF
    ch_l = ch & 0xFFFF

    w65 = reg_write(0x65, ch_h)
    w30 = reg_write(0x30, ch_l)
    return w65, w30


def _wait_tune_complete(timeout_s: float = 6.0, poll_s: float = 0.1):
    """Wait for tune completion.

    Prefer MSA NOP(0x00) pending-bits clear if possible; otherwise fall back
    to StatusW(0x21) BUSY bit clear.
    """
    t0 = time.time()
    # Try NOP pending bits first
    try:
        while (time.time() - t0) < float(timeout_s):
            _, _, nop = reg_read(0x00)
            pending = (nop >> 8) & 0xFF
            if pending == 0:
                return True
            time.sleep(float(poll_s))
        # Timed out; print a last snapshot for debugging
        try:
            _, _, nop_last = reg_read(0x00)
            _, _, sw_last = reg_read(0x21)
            print(f"[DEBUG] Tune wait timeout: NOP=0x{nop_last:04X} pending=0x{(nop_last>>8)&0xFF:02X} | StatusW=0x{sw_last:04X} BUSY={(sw_last & 0x0002)!=0}")
        except Exception:
            pass
        return False
    except Exception:
        # Fall back to StatusW BUSY
        while (time.time() - t0) < float(timeout_s):
            _, _, sw = reg_read(0x21)
            busy = (sw & 0x0002) != 0
            if not busy:
                return True
            time.sleep(float(poll_s))
        return False
# ---------------- Mode selection (set early!) ----------------
# ITLA_MODE: "uart" (real hardware) or "ctypes" (DLL/sim)
MODE = (os.environ.get("ITLA_MODE", "ctypes") or "ctypes").strip().lower()
IS_UART = (MODE == "uart")

# ---------------- Dependencies ----------------
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter



# ---------------- App output dir (store alongside this script) ----------------
# Requirement: Excel/log should be written into the same folder as this script.
# This avoids /root paths when running with sudo and avoids env-var dependence.
SCRIPT_DIR = Path(__file__).resolve().parent
APPDIR = SCRIPT_DIR
ConsoleLogPath = str(APPDIR / "console.log")
ExcelFileName  = str(APPDIR / "dvt_excel.xlsx")

# ---------------- Global Excel workbook handle ----------------
# Workbook is lazily created/loaded inside get_sheet().
workbook = None
EXCEL_HEADERS = [
    "Test",                  # A  - Test name / operation
    "Reg",                   # B  - Register (e.g. 0x01)
    "Length",                # C  - Actual length (OUT)
    "ExpectedLen",           # D  - Expected length
    "LenResult",             # E  - PASS / MISMATCH
    "ID_String",             # F  - Actual identity string
    "Expected_ID_Substr",    # G  - Expected substring
    "ID_Result",             # H  - PASS / FAIL for string
    "CE",                    # I  - CE bit
    "STATUS",                # J  - Status code
    "STATUS_Text",           # K  - Status text (OK/AEA/…)
    "OverallResult",         # L  - Pass / Fail / Info
    "Tester",                # M  - Tester name
    "DateTime",              # N  - Timestamp
    "Test Suite",            # O  - Table name / suite
    "Test Description"       # P  - Full response text
]

 


def get_sheet(sheet_name: str):
    """
    Return a worksheet with the standard DVT header.
    - If the Excel file exists, load it.
    - If the sheet exists, reuse it.
    - If the sheet is new, create it and add the header row.
    """
    global workbook

    # Initialise / load workbook if needed
    if workbook is None:
        if os.path.exists(ExcelFileName):
            workbook = load_workbook(ExcelFileName)
        else:
            workbook = Workbook()
            # Remove default empty sheet header later if unused
    if sheet_name == "SetChannel_Compare":
        # Reuse existing sheet if present; otherwise create it.
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(title=sheet_name)
            ws.append(SETCHANNEL_COMPARE_HEADERS)
            _excel_save()
        return ws
    # Get or create sheet with special handling for FullScan
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        # Special handling for FullScan: always reset contents and header
        if sheet_name == "FullScan":
            # Clear all existing rows (old headers + data)
            ws.delete_rows(1, ws.max_row)
            # Recreate compact header for full register scan
            ws.append([
                "Test", "Reg", "CE", "STATUS", "STATUS_Text", "Implemented", "OverallResult",
                "Tester", "DateTime", "Test Suite", "Test Description"
            ])
        return ws
    else:
        if sheet_name == "FullScan":
            ws = workbook.create_sheet(title=sheet_name)
            ws.append([
                "Test", "Reg", "CE", "STATUS", "STATUS_Text", "Implemented", "OverallResult",
                "Tester", "DateTime", "Test Suite", "Test Description"
            ])
            return ws
        if sheet_name == "SetChannelFull":
            ws = workbook.create_sheet(title=sheet_name)
            ws.append([
                "Test", "Result", "OverallResult", "Tester", "DateTime", "Test Suite"
            ])
            return ws
        if sheet_name == "MSA_Extensions":
            ws = workbook.create_sheet(title=sheet_name)
            ws.append(["Test", "Result", "OverallResult", "Tester", "DateTime", "Test Suite"])
            return ws
    if sheet_name == "WM_Compare_CSV":
        ws = workbook.create_sheet(title=sheet_name)
        ws.append([
            "CH",
            "Expected_WL_nm",
            "Measured_WL_nm",
            "Expected_GHz",
            "Measured_GHz",
            "Err_MHz",
            "Tol_MHz",
            "Power_dBm",
            "SMSR_dB",
            "Result",
            "Tester",
            "DateTime",
        ])
        return ws

    # Requirement Option 9 output (clean header; NO Notes column)
    if sheet_name == "Requirement_SetChannel":
        return _ensure_sheet_requirement_setchannel()

    # Default: create a generic DVT sheet with the standard headers
    ws = workbook.create_sheet(title=sheet_name)
    ws.append(EXCEL_HEADERS)
    return ws
# NOTE: Match the original LUT CSV unit naming for these columns (except Tolerance which is in GHz).
REQUIREMENT_SETCHANNEL_HEADERS = [
    "Channel",
    "Target Frequency (THz)",
    "Real Frequency (THz)",
    "External Wavemeter Frequency (THz)",
    "Real Wavelength (nm)",
    "External Wavemeter Wavelength (nm)",
    "Frequency Error (GHz)",
    "Tolerance (GHz)",
    "External Wavemeter Power (dBm)",
    "External Wavemeter SMSR (dB)",
    "Result",
    "Tester",
    "DateTime",
    "Notes",
]

# ---------------- Requirement_SetChannel Excel formatting ----------------

# Map long headers -> multi-line versions (Excel wrap)
_REQUIREMENT_HEADER_DISPLAY = {
    "Target Frequency (THz)": "Target\nFrequency\n(THz)",
    "Real Frequency (THz)": "Real\nFrequency\n(THz)",
    "External Wavemeter Frequency (THz)": "External\nWavemeter\nFrequency (THz)",
    "Real Wavelength (nm)": "Real\nWavelength\n(nm)",
    "External Wavemeter Wavelength (nm)": "External\nWavemeter\nWavelength (nm)",
    "Frequency Error (GHz)": "Frequency\nError\n(GHz)",
    "Tolerance (GHz)": "Tolerance\n(GHz)",
    "External Wavemeter Power (dBm)": "External\nWavemeter\nPower (dBm)",
    "External Wavemeter SMSR (dB)": "External\nWavemeter\nSMSR (dB)",
    "DateTime": "Date\nTime",
    "Notes": "Notes",
}

# Reasonable default column widths (tweakable later)
_REQUIREMENT_COL_WIDTHS = {
    "Channel": 9,
    "Target Frequency (THz)": 18,
    "Real Frequency (THz)": 18,
    "External Wavemeter Frequency (THz)": 22,
    "Real Wavelength (nm)": 18,
    "External Wavemeter Wavelength (nm)": 26,
    "Frequency Error (GHz)": 18,
    "Tolerance (GHz)": 14,
    "External Wavemeter Power (dBm)": 20,
    "External Wavemeter SMSR (dB)": 18,
    "Result": 10,
    "Tester": 12,
    "DateTime": 18,
    "Notes": 42,
}

def _format_requirement_setchannel_sheet(ws):
    """Apply consistent formatting to the Requirement_SetChannel sheet."""
    try:
        # Freeze header row
        ws.freeze_panes = "A2"

        # Autofilter
        max_col = ws.max_column
        if max_col and max_col > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"

        # Tab color (optional)
        try:
            ws.sheet_properties.tabColor = "1F4E79"
        except Exception:
            pass

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 54

        # Borders
        try:
            from openpyxl.styles import Border, Side
            thin_hdr = Side(style="thin", color="A6A6A6")
            header_border = Border(left=thin_hdr, right=thin_hdr, top=thin_hdr, bottom=thin_hdr)
            thin_data = Side(style="thin", color="D9D9D9")
            data_border = Border(left=thin_data, right=thin_data, top=thin_data, bottom=thin_data)
        except Exception:
            header_border = None
            data_border = None

        desired = REQUIREMENT_SETCHANNEL_HEADERS

        # Write wrapped display headers + style
        for col_idx, header in enumerate(desired, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = _REQUIREMENT_HEADER_DISPLAY.get(header, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            if header_border is not None:
                cell.border = header_border

        # Set column widths
        for col_idx, header in enumerate(desired, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = float(_REQUIREMENT_COL_WIDTHS.get(header, 14))

        # Data formatting
        data_align = Alignment(horizontal="center", vertical="center")

        num_fmt = {
            1: "0",           # Channel
            2: "0.00",        # Target THz (2 decimals)
            3: "0.000000",    # Real THz
            4: "0.0000000",   # External WM Frequency THz (7 decimals)
            5: "0.00000",     # Real Wavelength nm (5 decimals)
            6: "0.00000",     # External WM Wavelength nm (5 decimals)
            7: "0.000",       # Frequency Error GHz
            8: "0.##",        # Tolerance GHz
            9: "0.00",        # Power dBm
            10: "0.00",       # SMSR dB
        }

        if ws.max_row >= 2:
            for r in range(2, ws.max_row + 1):
                ws.row_dimensions[r].height = 18
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(r, c)
                    cell.alignment = data_align
                    if data_border is not None:
                        cell.border = data_border
                    if c in num_fmt and isinstance(cell.value, (int, float)):
                        cell.number_format = num_fmt[c]

    except Exception:
        # Formatting must never break tests
        pass

def _ensure_sheet_requirement_setchannel():
    """Ensure the Requirement_SetChannel sheet exists and has the expected header.

    IMPORTANT:
      - This function must NOT call get_sheet("Requirement_SetChannel"), otherwise it will recurse.
      - It works directly with the global `workbook`.
    """
    global workbook

    name = "Requirement_SetChannel"

    # Ensure workbook is initialized (safe even if called before get_sheet())
    if workbook is None:
        if os.path.exists(ExcelFileName):
            workbook = load_workbook(ExcelFileName)
        else:
            workbook = Workbook()

    # Get or create the sheet directly (NO get_sheet() calls)
    if name in workbook.sheetnames:
        ws = workbook[name]
    else:
        ws = workbook.create_sheet(title=name)

    desired = REQUIREMENT_SETCHANNEL_HEADERS

    # If sheet is empty, write header
    if ws.max_row < 1 or (ws.max_row == 1 and ws.cell(1, 1).value in (None, "")):
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(desired)
        _format_requirement_setchannel_sheet(ws)
        _excel_save()
        return ws

    # Compare current header cells (same width as desired)
    def _norm_hdr_cell(v) -> str:
        # Normalize headers so we can compare raw vs display (with newlines)
        s = "" if v is None else str(v)
        return " ".join(s.replace("\n", " ").split()).strip()

    current = [ws.cell(1, c).value for c in range(1, len(desired) + 1)]

    if [_norm_hdr_cell(x) for x in current] != [_norm_hdr_cell(x) for x in desired]:
        # Clear sheet and rewrite header
        ws.delete_rows(1, ws.max_row)
        ws.append(desired)
        _format_requirement_setchannel_sheet(ws)
        _excel_save()

    _format_requirement_setchannel_sheet(ws)
    return ws
# ---------------- Excel save helper ----------------
# Avoid saving on every row (slow on Pi / network filesystems).
# Use periodic saves for long loops and a final save at the end.

def _excel_save():
    """Save the workbook to disk (best-effort)."""
    global workbook
    try:
        if workbook is not None:
            workbook.save(ExcelFileName)
            print(f"[INFO] Excel saved: {ExcelFileName}")
    except Exception:
        # Don't crash tests due to transient IO / file lock issues.
        pass
# ---------------- DLL loading helpers (used only in ctypes mode) ----------------
def _candidate_lib_names():
    base_names = ["nanoitla", "libitla"]  # try both names
    if sys.platform.startswith("win"):
        exts = [".dll"]
    elif sys.platform == "darwin":
        exts = [".dylib", ".so"]
    else:
        exts = [".so"]
    for b in base_names:
        for e in exts:
            yield b + e

def _load_itla_lib():
    # 1) explicit override
    override = os.getenv("ITLA_DLL")
    if override:
        return ctypes.CDLL(override)

    # 2) optional DLL dir (Windows 3.8+)
    dll_dir = os.getenv("ITLA_DLL_DIR")
    if dll_dir and hasattr(os, "add_dll_directory") and sys.platform.startswith("win"):
        os.add_dll_directory(dll_dir)

    # 3) try dll_dir, then CWD
    search_dirs = []
    if dll_dir:
        search_dirs.append(Path(dll_dir))
    search_dirs.append(Path.cwd())

    for d in search_dirs:
        for name in _candidate_lib_names():
            p = d / name
            if p.exists():
                return ctypes.CDLL(str(p))

    # 4) let OS PATH handle it
    for name in _candidate_lib_names():
        try:
            return ctypes.CDLL(name)
        except OSError:
            pass

    raise OSError("Could not find ITLA library (set ITLA_DLL or ITLA_DLL_DIR).")

# ---------------- Conditional DLL Loading (lazy) ----------------
# IMPORTANT:
#   - Do NOT hard-exit on import. This file is imported for tools like HP_wavemeter.
#   - Only load the ctypes library when a ctypes register I/O path is actually used.
LIB = None
_LIB_READY = False

if IS_UART:
    print("[INFO] ITLA Mode selected: UART")
else:
    print("[INFO] ITLA Mode selected: CTYPES")


def print_env_snapshot():
    """Print only relevant env vars that are actually set.

    This does NOT set or modify any environment variables. It only reports what
    is already present in the process environment.
    """
    prefixes = ("ITLA_", "LUT_", "EXPECT_", "OPTICAL_", "STRICT_", "TUNING_", "HP_", "WM_", "WAVEMETER_")
    keys = sorted([k for k in os.environ.keys() if k.startswith(prefixes)])
    if not keys:
        print("[ENV] No ITLA/LUT/EXPECT/OPTICAL/STRICT/TUNING/HP_/WM_/WAVEMETER_ environment variables are set.")
        return

    print("[ENV] Environment variables currently set (filtered):")
    for k in keys:
        v = os.environ.get(k, "")
        print(f"  {k}={v}")


# Print env snapshot once at startup (helps detect accidental use of -E / inherited env)
print_env_snapshot()

def _ensure_ctypes_lib():
    """Lazy-load the ITLA shared library and set ctypes prototypes.

    This allows importing this module on systems that only need the wavemeter
    (GPIB) functionality, without requiring the ITLA library to be present.
    """
    global LIB, _LIB_READY
    if IS_UART:
        return
    if _LIB_READY and LIB is not None:
        return

    print("[INFO] Loading ITLA library for ctypes mode...")
    try:
        LIB = _load_itla_lib()
    except OSError as e:
        raise RuntimeError(f"Could not find ITLA library (set ITLA_DLL or ITLA_DLL_DIR). Details: {e}")

    # Optional embedded-python init
    if hasattr(LIB, "itla_py_init"):
        try:
            LIB.itla_py_init()
        except Exception as e:
            print(f"[WARN] itla_py_init() failed: {e}")

    # Prototypes (only when LIB is valid)
    if hasattr(LIB, "itla_is_python"):
        LIB.itla_is_python.restype = ctypes.c_int
    if hasattr(LIB, "itla_use_python"):
        LIB.itla_use_python.argtypes = [ctypes.c_int]
    if hasattr(LIB, "itla_py_init"):
        LIB.itla_py_init.restype = ctypes.c_int
    if hasattr(LIB, "itla_py_fini"):
        LIB.itla_py_fini.restype = None
    if hasattr(LIB, "itla_process"):
        LIB.itla_process.argtypes = [
            ctypes.c_uint8, ctypes.c_uint8, ctypes.c_uint8, ctypes.c_uint16,
            ctypes.POINTER(ctypes.c_uint8), ctypes.POINTER(ctypes.c_uint8),
            ctypes.POINTER(ctypes.c_uint8), ctypes.POINTER(ctypes.c_uint16),
            ctypes.POINTER(ctypes.c_uint32)
        ]
        LIB.itla_process.restype = ctypes.c_uint32

    _LIB_READY = True

# ---------------- Transport selection ----------------
SER_AVAILABLE = False
ser = None
if IS_UART:
    try:
        import serial  # pyserial
        SER_AVAILABLE = True
    except Exception:
        # try auto-install once
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "pyserial"])
            import serial
            SER_AVAILABLE = True
        except Exception as e:
            print(f"[WARN] UART mode requested but pyserial not available: {e}")
            SER_AVAILABLE = False

#
# ---------------- UART auto-open (for import/use in one-liners) ----------------

def _uart_open_if_needed():
    """Ensure UART `ser` is open when running in UART mode.

    This module is often imported from one-liners (e.g. python3 -c / heredoc)
    where the interactive menu isn't run. In that case `ser` remains None and
    reg_read/reg_write will crash. This helper makes reg_read/reg_write usable
    in both interactive and imported contexts.

    Env vars:
      - ITLA_UART_PORT (default /dev/ttyUSB0)
      - ITLA_UART_BAUD (default 115200)
      - ITLA_UART_TIMEOUT (default 0.5)
    """
    global ser
    if not IS_UART:
        return
    if ser is not None:
        return
    if not SER_AVAILABLE:
        raise RuntimeError("UART mode requested but pyserial is not available")

    port = os.environ.get("ITLA_UART_PORT", "/dev/ttyUSB0")
    baud = int(os.environ.get("ITLA_UART_BAUD", "115200"), 0)
    timeout = float(os.environ.get("ITLA_UART_TIMEOUT", "0.5"))

    try:
        ser = serial.Serial(port, baudrate=baud, timeout=timeout)
        try:
            # best-effort flush
            ser.reset_input_buffer()
            ser.reset_output_buffer()
        except Exception:
            pass
        print(f"[INFO] Opened {port} @{baud}")
    except Exception as e:
        raise RuntimeError(f"Failed to open UART {port} @{baud}: {e}")

# ---------------- UART helpers ----------------
def _bip4(b0, b1, b2, b3):
    bip8 = (b0 & 0x0F) ^ b1 ^ b2 ^ b3
    return ((bip8 >> 4) & 0x0F) ^ (bip8 & 0x0F)

def build_inbound_frame(lstRsp, is_write, reg, data):
    app = ((is_write & 1) << 26) | ((reg & 0xFF) << 18) | (data & 0xFFFF)
    tmp = ((lstRsp & 1) << 27) | (app & 0x07FFFFFF)
    b0 = (tmp >> 24) & 0xFF; b1 = (tmp >> 16) & 0xFF; b2 = (tmp >> 8) & 0xFF; b3 = tmp & 0xFF
    csum = _bip4(b0 & 0x0F, b1, b2, b3) & 0x0F
    return tmp | (csum << 28)

def send_frame(ser, frame_u32):
    ser.write(bytes([(frame_u32 >> 24) & 0xFF, (frame_u32 >> 16) & 0xFF,
                     (frame_u32 >> 8) & 0xFF, frame_u32 & 0xFF]))

def recv_frame(ser, attempts=6):
    # attempts * ser.timeout = total wait (e.g., 6 * 0.5s = 3s)
    for _ in range(attempts):
        buf = bytearray()
        while len(buf) < 4:
            chunk = ser.read(4 - len(buf))
            if not chunk:
                break  # this attempt timed out; try again
            buf.extend(chunk)
        if len(buf) == 4:
            return (buf[0] << 24) | (buf[1] << 16) | (buf[2] << 8) | buf[3]
    raise IOError("Timeout or short read from UART (no 4-byte frame)")

def parse_outbound(frame_u32):
    ce  = (frame_u32 >> 27) & 0x1
    xe  = (frame_u32 >> 25) & 0x1
    reg = (frame_u32 >> 17) & 0xFF
    data= (frame_u32 >> 1)  & 0xFFFF
    status = 1 if xe else 0   # UART path only exposes XE bit
    return ce, status, reg, data, frame_u32

def _bits32(v: int) -> str:
    s = "".join("1" if (v & (1<<i)) else "0" for i in range(31,-1,-1))
    return f"{s[:8]} {s[8:16]} {s[16:24]} {s[24:]}"

def _status_name(s: int) -> str:
    # In ctypes mode: 0 OK, 1 XE, 2 AEA, 3 CP
    # In UART mode:   0 OK, 1 XE  (we may annotate AEA hint for identity reads)
    if IS_UART:
        return "OK" if s == 0 else "XE"
    return "OK" if s == 0 else "XE" if s == 1 else "AEA" if s == 2 else "CP"


# --- Status code acceptability helper ---
def _status_ok_for_test(status: int, allow_aea: bool = False) -> bool:
    """Return True if STATUS is acceptable for the current transport.

    UART mode exposes only OK(0) vs XE(1). CTYPES mode can return OK(0), XE(1), AEA(2), CP(3).
    Some tests (identity / EA) may accept AEA in ctypes mode.
    """
    if IS_UART:
        return status == 0
    if allow_aea:
        return status in (0, 2)
    return status == 0


# ---------------- 9.5 human-readable decoders ----------------
_STATUSF_BITS = {
    0: "LFPWR", 1: "LFTHERM", 2: "LFFREQ", 3: "LFVSF",
    4: "CRL",   5: "MRL",     6: "CEL",    7: "XEL",
    8: "FPWR",  9: "FTHERM", 10: "FFREQ", 11: "FVSF",
    12: "DIS", 13: "FATAL",  14: "ALM",   15: "SRQ",
}

_STATUSW_BITS = {
    0: "LWPWR", 1: "LWTHERM", 2: "LWFREQ", 3: "LWVSF",
    4: "CRL",   5: "MRL",     6: "CEL",    7: "XEL",
    8: "WPWR",  9: "WTHERM", 10: "WFREQ", 11: "WVSF",
    12: "DIS", 13: "FATAL",  14: "ALM",   15: "SRQ",
}

_SRQT_BITS = {
    0: "LFPWR", 1: "LFTHERM", 2: "LFFREQ", 3: "LFVSF",
    4: "CRL",   5: "MRL",     6: "CEL",    7: "XEL",
    8: "LWPWR", 9: "LWTHERM", 10: "LWFREQ", 11: "LWVSF",
    12: "DIS",
}

_FATALT_BITS = {
    0: "LFPWR", 1: "LFTHERM", 2: "LFFREQ", 3: "LFVSF",
    5: "MRL",
    8: "LWPWR", 9: "LWTHERM", 10: "LWFREQ", 11: "LWVSF",
}

_ALMT_BITS = {
    0: "FPWR", 1: "FTHERM", 2: "FFREQ", 3: "FVSF",
    8: "WPWR", 9: "WTHERM", 10: "WFREQ", 11: "WVSF",
}

def _decode_bits16(val: int, bit_names: dict) -> str:
    names = []
    for bit in range(16):
        if (val >> bit) & 0x1:
            names.append(bit_names.get(bit, f"b{bit}"))
    return ",".join(names) if names else "none"

def _decode_set_bits(val: int) -> str:
    bits = [str(bit) for bit in range(16) if ((val >> bit) & 0x1)]
    return ",".join(bits) if bits else "none"

def _decode_9_5_value(reg: int, val: int) -> str:
    val = int(val) & 0xFFFF

    if reg == 0x14:
        names = []
        if val & 0x0001:
            names.append("START")
        if val & 0x0002:
            names.append("ABORT")
        if val & ~(0x0003):
            names.append(f"raw=0x{val:04X}")
        return f"DLConfig[{','.join(names) if names else 'none'}]"
    if reg == 0x15:
        names = []
        if val & 0x0001:
            names.append("BUSY")
        if val & 0x0002:
            names.append("DONE")
        if val & 0x0004:
            names.append("FAIL")
        if val & 0x0008:
            names.append("ABORT")
        if val & ~(0x000F):
            names.append(f"raw=0x{val:04X}")
        return f"DLStatus[{','.join(names) if names else 'none'}]"

    if reg == 0x20:
        return f"StatusF[{_decode_bits16(val, _STATUSF_BITS)}]"
    if reg == 0x21:
        return f"StatusW[{_decode_bits16(val, _STATUSW_BITS)}]"
    if reg == 0x28:
        return f"SRQT[{_decode_bits16(val, _SRQT_BITS)}]"
    if reg == 0x29:
        return f"FatalT[{_decode_bits16(val, _FATALT_BITS)}]"
    if reg == 0x2A:
        return f"ALMT[{_decode_bits16(val, _ALMT_BITS)}]"

    if reg in (0x22, 0x23):
        return f"{val / 100.0:.2f} dB"
    if reg in (0x24, 0x25):
        return f"{val / 10.0:.1f} GHz coarse"
    if reg in (0x63, 0x64):
        return f"{val} MHz fine"
    if reg in (0x26, 0x27):
        return f"{val / 100.0:.2f} C threshold"

    return ""

# ---------------- CTYPES Register I/O ----------------
def _reg_read_ctypes(reg: int):
    _ensure_ctypes_lib()
    ce = ctypes.c_uint8(0)
    status = ctypes.c_uint8(0)
    rout = ctypes.c_uint8(0)
    dout = ctypes.c_uint16(0)
    inframe = ctypes.c_uint32(0)
    out = LIB.itla_process(0, 0, reg & 0xFF, 0,
                           ctypes.byref(ce), ctypes.byref(status),
                           ctypes.byref(rout), ctypes.byref(dout),
                           ctypes.byref(inframe))
    print(f"Inbound:  0x{inframe.value:08X}  [{_bits32(inframe.value)}]")
    print(f"Outbound: 0x{out:08X}        [{_bits32(out)}]")
    print(f"  CE={ce.value}, STATUS={status.value}, REG=0x{rout.value:02X}, OUT=0x{dout.value:04X}")
    return ce.value, status.value, dout.value

def _reg_write_ctypes(reg: int, data: int):
    _ensure_ctypes_lib()
    ce = ctypes.c_uint8(0)
    status = ctypes.c_uint8(0)
    rout = ctypes.c_uint8(0)
    dout = ctypes.c_uint16(0)
    inframe = ctypes.c_uint32(0)
    out = LIB.itla_process(0, 1, reg & 0xFF, data & 0xFFFF,
                           ctypes.byref(ce), ctypes.byref(status),
                           ctypes.byref(rout), ctypes.byref(dout),
                           ctypes.byref(inframe))
    print(f"Inbound:  0x{inframe.value:08X}  [{_bits32(inframe.value)}]")
    print(f"Outbound: 0x{out:08X}        [{_bits32(out)}]")
    print(f"  CE={ce.value}, STATUS={status.value}, REG=0x{rout.value:02X}, IN=0x{data & 0xFFFF:04X}, OUT=0x{dout.value:04X}")
    return ce.value, status.value, data & 0xFFFF, dout.value

# ---------------- UART Register I/O ----------------
def _reg_read_uart(reg: int):
    _uart_open_if_needed()
    in_frame = build_inbound_frame(0, 0, reg & 0xFF, 0)
    send_frame(ser, in_frame)
    out = recv_frame(ser)
    ce, status, rout, dout, _ = parse_outbound(out)
    ui_status = "OK/AEA" if (reg in range(0x01, 0x08) and status == 0) else _status_name(status)
    print(f"Inbound:  0x{in_frame:08X}  [{_bits32(in_frame)}]")
    print(f"Outbound: 0x{out:08X}        [{_bits32(out)}]")
    print(f"  CE={ce}, STATUS={status} ({ui_status}), REG=0x{rout:02X}, OUT=0x{dout:04X}")
    return ce, status, dout

def _reg_write_uart(reg: int, data: int):
    _uart_open_if_needed()
    in_frame = build_inbound_frame(0, 1, reg & 0xFF, data & 0xFFFF)
    send_frame(ser, in_frame)
    out = recv_frame(ser)
    ce, status, rout, dout, _ = parse_outbound(out)
    print(f"Inbound:  0x{in_frame:08X}  [{_bits32(in_frame)}]")
    print(f"Outbound: 0x{out:08X}        [{_bits32(out)}]")
    print(f"  CE={ce}, STATUS={status} ({_status_name(status)}), REG=0x{rout:02X}, IN=0x{data & 0xFFFF:04X}, OUT=0x{dout:04X}")
    return ce, status, data & 0xFFFF, dout

# ---------------- Transport switch ----------------
def reg_read(reg: int):
    return _reg_read_uart(reg) if IS_UART else _reg_read_ctypes(reg)

def reg_write(reg: int, data: int):
    return _reg_write_uart(reg, data) if IS_UART else _reg_write_ctypes(reg, data)


def _reg_read_uart_quiet(reg: int):
    _uart_open_if_needed()
    in_frame = build_inbound_frame(0, 0, reg & 0xFF, 0)
    send_frame(ser, in_frame)
    out = recv_frame(ser)
    ce, status, _rout, dout, _ = parse_outbound(out)
    return ce, status, dout


def _reg_write_uart_quiet(reg: int, data: int):
    _uart_open_if_needed()
    in_frame = build_inbound_frame(0, 1, reg & 0xFF, data & 0xFFFF)
    send_frame(ser, in_frame)
    out = recv_frame(ser)
    ce, status, _rout, dout, _ = parse_outbound(out)
    return ce, status, (data & 0xFFFF), dout


def _reg_read_ctypes_quiet(reg: int):
    _ensure_ctypes_lib()
    ce = ctypes.c_uint8(0)
    status = ctypes.c_uint8(0)
    rout = ctypes.c_uint8(0)
    dout = ctypes.c_uint16(0)
    inframe = ctypes.c_uint32(0)
    LIB.itla_process(0, 0, reg & 0xFF, 0,
                     ctypes.byref(ce), ctypes.byref(status),
                     ctypes.byref(rout), ctypes.byref(dout),
                     ctypes.byref(inframe))
    return ce.value, status.value, dout.value


def _reg_write_ctypes_quiet(reg: int, data: int):
    _ensure_ctypes_lib()
    ce = ctypes.c_uint8(0)
    status = ctypes.c_uint8(0)
    rout = ctypes.c_uint8(0)
    dout = ctypes.c_uint16(0)
    inframe = ctypes.c_uint32(0)
    LIB.itla_process(0, 1, reg & 0xFF, data & 0xFFFF,
                     ctypes.byref(ce), ctypes.byref(status),
                     ctypes.byref(rout), ctypes.byref(dout),
                     ctypes.byref(inframe))
    return ce.value, status.value, (data & 0xFFFF), dout.value


def reg_read_quiet(reg: int):
    return _reg_read_uart_quiet(reg) if IS_UART else _reg_read_ctypes_quiet(reg)


def reg_write_quiet(reg: int, data: int):
    return _reg_write_uart_quiet(reg, data) if IS_UART else _reg_write_ctypes_quiet(reg, data)

# ===============================
# MSA register extensions helpers
# ===============================
# Manufacturer-specific extensions (match nanoITLA.c additions)
REG_TUNER_PHASE = 0x8C
REG_TUNER_RING1 = 0x8D
REG_TUNER_RING2 = 0x8E
REG_SOA         = 0x8F
REG_BIAS        = 0x90
REG_TEC         = 0x91

# PD aliases (already used in PD test / Web GUI)
REG_WL_PD     = 0x89
REG_ETALON_PD = 0x8A
REG_POWER_PD  = 0x8B


def _float_to_u16_fixed(val: float, scale: float) -> int:
    x = int(round(float(val) * float(scale)))
    if x < 0:
        x = 0
    if x > 0xFFFF:
        x = 0xFFFF
    return x

def _u16_to_float_fixed(u: int, scale: float) -> float:
    return float(int(u) & 0xFFFF) / float(scale)

def _s16_from_u16(u: int) -> int:
    u = int(u) & 0xFFFF
    return u - 0x10000 if (u & 0x8000) else u

def _u16_from_s16(s: int) -> int:
    s = int(s)
    if s < -32768:
        s = -32768
    if s > 32767:
        s = 32767
    return s & 0xFFFF


# ---- PD reads (scaled x10 -> float) ----
def read_wl_pd() -> float:
    _, _, dout = reg_read(REG_WL_PD)
    return float(dout) / 10.0

def read_etalon_pd() -> float:
    _, _, dout = reg_read(REG_ETALON_PD)
    return float(dout) / 10.0

def read_power_pd() -> float:
    _, _, dout = reg_read(REG_POWER_PD)
    return float(dout) / 10.0


# ---- Tuners (milli-units: value * 1000) ----
def write_tuner(which: str, val: float):
    w = str(which or "").strip().lower()
    if w == "phase":
        reg = REG_TUNER_PHASE
    elif w == "ring1":
        reg = REG_TUNER_RING1
    elif w == "ring2":
        reg = REG_TUNER_RING2
    else:
        raise ValueError("which must be 'phase', 'ring1', or 'ring2'")

    data = _float_to_u16_fixed(val, 1000.0)
    reg_write(reg, data)

def WriteTuners(phase_val: float, ring1_val: float, ring2_val: float):
    write_tuner("phase", phase_val)
    write_tuner("ring1", ring1_val)
    write_tuner("ring2", ring2_val)

def read_tuner(which: str) -> float:
    w = str(which or "").strip().lower()
    if w == "phase":
        reg = REG_TUNER_PHASE
    elif w == "ring1":
        reg = REG_TUNER_RING1
    elif w == "ring2":
        reg = REG_TUNER_RING2
    else:
        raise ValueError("which must be 'phase', 'ring1', or 'ring2'")

    _, _, dout = reg_read(reg)
    return _u16_to_float_fixed(dout, 1000.0)


# ---- SOA (centi-units: value * 100) ----
def WriteSOA(val: float):
    reg_write(REG_SOA, _float_to_u16_fixed(val, 100.0))

def read_soa() -> float:
    _, _, dout = reg_read(REG_SOA)
    return _u16_to_float_fixed(dout, 100.0)


# ---- Bias / TEC (signed raw int16) ----
def WriteBias(val: int):
    reg_write(REG_BIAS, _u16_from_s16(val))

def read_bias() -> int:
    _, _, dout = reg_read(REG_BIAS)
    return _s16_from_u16(dout)

def WriteTEC(val: int):
    reg_write(REG_TEC, _u16_from_s16(val))

def read_tec() -> int:
    _, _, dout = reg_read(REG_TEC)
    return _s16_from_u16(dout)
# ---------------- Logging ----------------
DVTCONhandler = "CONDVT"
DVTCONLogger = None
TesterName = ""
loggerName = DVTCONhandler

def init_CONlogger(name):
    logger = logging.getLogger(name + "C")
    log_format = '@%(asctime)s @%(message)s'
    handler = RotatingFileHandler(ConsoleLogPath, maxBytes=1_000_000, backupCount=4)
    print(f"[INFO] Console log: {ConsoleLogPath}")
    logger_format = logging.Formatter(fmt=log_format, datefmt="%d/%m/%Y-%H:%M:%S")
    handler.setFormatter(logger_format)
    logger.addHandler(handler)
    logger.setLevel(logging.DEBUG)
    return [logger]

def DVTCONLoggerInit():
    global DVTCONLogger
    if DVTCONLogger is None:
        DVTCONLogger = init_CONlogger(DVTCONhandler)
    return DVTCONLogger

def sprint(a, b):
    if b != 0: print(a, b)
    else:      print(a)

logger = DVTCONLoggerInit()

# ---------------- Expected-values helper + Supervisory Table A ----------------

def _expect(default_val: int, env_name: str) -> int:
    """
    Read an override from environment (hex or decimal), or fall back to default.
    Example: EXPECT_DEVTYPE=0x0008 or 8
    """
    raw = os.getenv(env_name)
    if not raw:
        return default_val
    try:
        return int(raw, 0)
    except ValueError:
        return default_val

# Table A = Supervisory identity reads (MSA §9.x)
# Defaults taken from nanoITLA.c string lengths (len+1).
SUPERVISORY_TESTS = [
    ("DeviceType",    0x01, _expect(0x0008, "EXPECT_DEVTYPE")),
    ("Manufacturer",  0x02, _expect(0x0010, "EXPECT_MFGR")),
    ("Model",         0x03, _expect(0x000A, "EXPECT_MODEL")),
    ("SerialNumber",  0x04, _expect(0x000A, "EXPECT_SERNO")),
    ("MfgDate",       0x05, _expect(0x000B, "EXPECT_MFGDATE")),
    ("Release",       0x06, _expect(0x0006, "EXPECT_RELEASE")),
    ("RelBack",       0x07, _expect(0x0006, "EXPECT_RELBACK")),
]

# Expected ASCII identity strings from nanoITLA.c (can be overridden via env)
EXPECTED_ID_STRINGS = {
    0x01: os.getenv("EXPECT_DEVTYPE_STR", "CW ITLA"),
    0x02: os.getenv("EXPECT_MFGR_STR",    "Pilot Photonics"),
    0x03: os.getenv("EXPECT_MODEL_STR",   "NYITLA-01"),
    0x04: os.getenv("EXPECT_SERNO_STR",   "PP-000123"),
    0x05: os.getenv("EXPECT_MFGDATE_STR", "2025-08-27"),
    0x06: os.getenv("EXPECT_RELEASE_STR", "1.0.0"),
    0x07: os.getenv("EXPECT_RELBACK_STR", "1.0.0"),
}

# Requirements tracker register map.
REQUIREMENT_SECTIONS = [
    ("9.5",  "Module status / alarms / triggers", [0x14, 0x15, 0x20, 0x21, 0x22, 0x23, 0x24, 0x25, 0x63, 0x64, 0x26, 0x27, 0x28, 0x29, 0x2A]),
    ("9.6",  "General module configuration",       [0x30, 0x65, 0x31, 0x32, 0x33, 0x34, 0x66, 0x35, 0x36, 0x67, 0x40, 0x41, 0x68, 0x42, 0x43]),
    ("9.7",  "Fine tune / limits",                 [0x4F, 0x50, 0x51, 0x52, 0x53, 0x69, 0x54, 0x55, 0x6A, 0x56, 0x6B]),
    ("9.8",  "Health / dither / age",              [0x57, 0x58, 0x59, 0x5A, 0x5B, 0x5C, 0x5D, 0x5E, 0x5F, 0x60, 0x61, 0x62]),
    ("9.9",  "Manufacturer specific (implemented subset)", [0x80, 0x81, 0x82, 0x83, 0x84, 0x85, 0x86, 0x87, 0x88, 0x89, 0x8A, 0x8B, 0x8C, 0x8D, 0x8E, 0x8F, 0x90, 0x91, 0x92]),
]

# Flat set of all registers the nanoITLA firmware is expected to implement.
# Base identity / NOP registers (MSA 9.3/9.4): 0x00–0x15
# Plus every address listed in REQUIREMENT_SECTIONS (9.5–9.9).
# Any register NOT in this set that returns XE is labelled "Not Implemented"
# (expected behaviour). Any register IN this set that returns XE is a real Fail.
IMPLEMENTED_REGISTERS: set = (
    set(range(0x00, 0x16))  # 0x00–0x15: identity, NOP, basic status
    | {reg for _, _, regs in REQUIREMENT_SECTIONS for reg in regs}
)

# ---------------- Extended Address (EA) helpers for identity strings ----------------

EAC_INC_ON_READ = 0x0001  # matches EAC_INC_ON_READ in nanoITLA.c for AEA_EAC (0x09)

def ea_enable_auto_increment():
    """
    Enable 'increment on read' for the AEA path (0x09 / 0x0A / 0x0B).
    This matches the st_AEA_EAC behaviour in nanoITLA.c.
    """
    # 0x09 = AEA_EAC
    reg_write(0x09, EAC_INC_ON_READ)


def read_identity_string(reg: int, max_bytes: int = 64) -> str:
    """
    Read an identity string (DevTyp/Mfgr/Model/SerNo/MfgDate/Release/RelBack)
    using the AEA extended-address path:

      1. Enable auto-increment on read (0x09).
      2. Read the identity register (0x01..0x07) to select the string
         and get its length (len+1, including '\0').
      3. Repeatedly read 0x0B; each read returns 2 bytes from the EA buffer.
      4. Stop at NUL (0x00) or max_bytes.

    Returns:
        decoded ASCII string (without the terminating NUL).
    """
    # 1) enable EA auto-increment
    ea_enable_auto_increment()

    # 2) prime AEA buffer by reading the ID register
    ce, status, length = reg_read(reg)
    if ce != 0 or status not in (0, 2):  # 0=OK, 2=AEA
        raise RuntimeError(
            f"Failed to prime EA for reg 0x{reg:02X}: CE={ce}, STATUS={status}"
        )

    # Optional: use length as an upper bound if it's sane
    if 0 < length < max_bytes:
        max_bytes = length

    # 3) stream bytes from 0x0B
    buf = []
    while len(buf) < max_bytes:
        ce2, status2, word = reg_read(0x0B)  # EAR data
        if ce2 != 0 or status2 not in (0, 2, 3):  # allow CP as informational
            break

        hi = (word >> 8) & 0xFF
        lo = word & 0xFF
        for b in (hi, lo):
            if b == 0x00:  # NUL terminator
                text = bytes(buf).decode("ascii", errors="ignore")
                return text
            buf.append(b)

    text = bytes(buf).decode("ascii", errors="ignore")
    return text


def decode_9_8_aea(reg: int, payload_len: int) -> str:
    """
    Human decoder for section 9.8 AEA payloads:
      - 0x57: currents array (4 signed words, mA*10)
      - 0x58: temperatures array (2 signed words, C*100)
    """
    if reg not in (0x57, 0x58):
        return ""
    if payload_len <= 0:
        return "AEA[empty]"

    # Bound reads to expected payload size for known regs.
    expected_words = 4 if reg == 0x57 else 2
    words_to_read = min(expected_words, max(1, payload_len // 2))

    try:
        ce, status, _din, _dout = reg_write_quiet(0x09, EAC_INC_ON_READ)
        if ce != 0 or status not in (0, 2):
            return f"AEA[setup_failed CE={ce} ST={status}]"

        ce, status, _din, _dout = reg_write_quiet(0x0A, 0x0000)
        if ce != 0 or status not in (0, 2):
            return f"AEA[ptr_failed CE={ce} ST={status}]"

        raw_words = []
        for _ in range(words_to_read):
            ce, status, dout = reg_read_quiet(0x0B)
            if ce != 0 or status not in (0, 2, 3):
                break
            raw_words.append(int(dout) & 0xFFFF)

        if not raw_words:
            return "AEA[no_data]"

        s16 = [_s16_from_u16(w) for w in raw_words]
        if reg == 0x57:
            labels = ("TEC", "DIODE", "MON", "SOA")
            fields = [f"{labels[i]}={s16[i] / 10.0:.1f}mA" for i in range(min(len(s16), 4))]
            return "Currents[" + ", ".join(fields) + "]"
        else:
            labels = ("DIODE", "CASE")
            fields = [f"{labels[i]}={s16[i] / 100.0:.2f}C" for i in range(min(len(s16), 2))]
            return "Temps[" + ", ".join(fields) + "]"
    except Exception as ex:
        return f"AEA[decode_error: {ex}]"


# ---------------- Table A: Supervisory identity reads ----------------

def run_supervisory_table():
    """
    Table A: Supervisory identity reads (DevType, Mfgr, Model, SerNo, MfgDate, Release, RelBack)

    For each of 0x01–0x07:
      - Executes a read via reg_read()
      - Verifies OUT (length) against expected
      - Uses EA (0x09/0x0B) to stream the identity string
      - Compares the string to EXPECTED_ID_STRINGS
      - Logs to Excel + console + rotating log
    """
    print("\n[Table A] Supervisory Identity Reads")
    ws = get_sheet("TableA")
    for name, reg, expected in SUPERVISORY_TESTS:
        # Defaults for Excel row
        dout = 0
        ce = 0
        status = -1
        status_name = "UNKNOWN"
        value_ok = False
        str_ok = True
        read_str = ""
        expected_str = EXPECTED_ID_STRINGS.get(reg, "")

        try:
            # --- raw register read ---
            ce, status, dout = reg_read(reg)
            status_name = _status_name(status)

            # Status rules: OK or AEA are acceptable here (depending on mode)
            if MODE == "uart":
                ok_status = (status == 0)
            else:
                ok_status = (status in (0, 2))  # 0=OK, 2=AEA

            # ---------- length check (OUT word) ----------
            if expected is None:
                value_ok = True
                length_part = f"len={dout}"
                len_check = ""
            else:
                value_ok = (dout == expected)
                len_check = "PASS" if value_ok else "MISMATCH"
                length_part = f"len={dout} (expected {expected}) [{len_check}]"

            # ---------- optional identity string check via EA ----------
            if expected_str:
                try:
                    read_str = read_identity_string(reg)
                    # tolerate substring match
                    str_ok = (expected_str in read_str)
                except Exception as se:
                    str_ok = False
                    read_str = f"<STR_ERROR: {se}>"
            else:
                # No expected string configured ⇒ we don't fail on it
                str_ok = True

            pass_fail = "Pass" if (ce == 0 and ok_status and value_ok and str_ok) else "Fail"

            # ---------- final “requirement-style” console line ----------
            extra_info = ""
            if read_str:
                extra_info += f"  STR='{read_str}'"
                if expected_str and not str_ok and expected_str not in read_str:
                    extra_info += f"  (expected contains '{expected_str}')"

            response = (
                f"{name:<12s} (Reg 0x{reg:02X}): "
                f"{length_part}  "
                f"CE={ce}, STATUS={status} ({status_name})"
                f"{extra_info}"
            )

        except Exception as e:
            pass_fail = "Fail"
            response = f"{name:<12s} (Reg 0x{reg:02X}): ERROR talking to device: {e}"

        # ---------- console print (Table A summary line) ----------
        print("  " + response)

        # ---------- Excel logging (structured columns) ----------
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Decide length check text for Excel
        if expected is None:
            excel_len_check = ""
        else:
            excel_len_check = "PASS" if value_ok else "MISMATCH"

        # String check text
        if expected_str:
            excel_str_check = "PASS" if str_ok else "FAIL"
        else:
            excel_str_check = ""

        ws.append([
            name,                          # 1: TestName
            f"0x{reg:02X}",                # 2: RegHex
            dout,                          # 3: ActualLen
            expected if expected is not None else "",  # 4: ExpectedLen
            excel_len_check,               # 5: LenCheck
            read_str,                      # 6: IdentityString (actual)
            expected_str,                  # 7: ExpectedSubstring
            excel_str_check,               # 8: StringCheck
            ce,                            # 9: CE
            status,                        # 10: StatusCode
            status_name,                   # 11: StatusName
            pass_fail,                     # 12: OverallResult
            TesterName,                    # 13: Tester
            current_datetime,              # 14: DateTime
            "TableA",                      # 15: TableName
            response                       # 16: FullResponseText
        ])
        # (saving deferred) 

        # ---------- Rotating log (reuses same response text) ----------
        logger_info = (
            f": Test: {name} Response: {response}, "
            f"Result: {pass_fail}, Tester: {TesterName}"
        )
        if pass_fail == "Pass":
            logger[0].info(loggerName + logger_info)
        else:
            logger[0].warning(loggerName + logger_info)

        # Flush handlers
        for h in logger[0].handlers:
            try:
                h.flush()
            except Exception:
                pass

    # Save once at the end of the table (faster than saving per row)
    _excel_save()


def run_9_5_status_table():
    """
    Table 9.5: Module status / alarms / triggers (starter implementation).

    - Verifies read path for status registers.
    - Verifies read/write echo for threshold/trigger/config registers.
    - Restores original values after RW checks.
    """
    print("\n[Table 9.5] Module Status / Trigger Registers")
    ws = get_sheet("Table_9_5_Status")

    # mode: "RO" = read only check, "RW" = write/readback/restore
    reg_plan = [
        ("DLConfig",       0x14, "RW"),
        ("DLStatus",       0x15, "RO"),
        ("StatusF",        0x20, "RO"),
        ("StatusW",        0x21, "RO"),
        ("FPowTh",         0x22, "RW"),
        ("WPowTh",         0x23, "RW"),
        ("FFreqTh",        0x24, "RW"),
        ("WFreqTh",        0x25, "RW"),
        ("FFreqTh2",       0x63, "RW"),
        ("WFreqTh2",       0x64, "RW"),
        ("FThermTh",       0x26, "RW"),
        ("WThermTh",       0x27, "RW"),
        ("SRQ_Triggers",   0x28, "RW"),
        ("FatalTriggers",  0x29, "RW"),
        ("ALMT",           0x2A, "RW"),
    ]

    # Keep writes deterministic and low-risk.
    rw_pattern = {
        0x14: 0x0001,
        0x22: 0x0011,
        0x23: 0x0012,
        0x24: 0x0013,
        0x25: 0x0014,
        0x63: 0x0017,
        0x64: 0x0018,
        0x26: 0x0015,
        0x27: 0x0016,
        0x28: 0x0001,
        0x29: 0x0001,
        0x2A: 0x0001,
    }

    for name, reg, mode in reg_plan:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            ce_r0, st_r0, v0 = reg_read(reg)
            st_name_r0 = _status_name(st_r0)

            if mode == "RO":
                ok = (ce_r0 == 0 and _status_ok_for_test(st_r0, allow_aea=True))
                d0 = _decode_9_5_value(reg, v0)
                msg = (
                    f"{name}: REG=0x{reg:02X}, mode=RO, "
                    f"OUT=0x{v0:04X}, CE={ce_r0}, STATUS={st_r0} ({st_name_r0})"
                )
                if d0:
                    msg += f", decode={d0}"
            else:
                test_val = rw_pattern.get(reg, 0x0001)

                ce_w, st_w, _, _ = reg_write(reg, test_val)
                ce_r1, st_r1, v1 = reg_read(reg)
                st_name_r1 = _status_name(st_r1)

                # Restore original value best-effort.
                try:
                    reg_write(reg, v0)
                except Exception:
                    pass

                write_ok = (ce_w == 0 and _status_ok_for_test(st_w, allow_aea=True))
                read_ok = (ce_r1 == 0 and _status_ok_for_test(st_r1, allow_aea=True))
                value_ok = (v1 == (test_val & 0xFFFF))
                ok = write_ok and read_ok and value_ok

                msg = (
                    f"{name}: REG=0x{reg:02X}, mode=RW, "
                    f"W=0x{test_val:04X}, R=0x{v1:04X}, "
                    f"CEw={ce_w}, STw={st_w} ({_status_name(st_w)}), "
                    f"CEr={ce_r1}, STr={st_r1} ({st_name_r1}), "
                    f"orig=0x{v0:04X}"
                )
                dw = _decode_9_5_value(reg, test_val)
                dr = _decode_9_5_value(reg, v1)
                if dw or dr:
                    msg += f", Wdec={dw or '-'}, Rdec={dr or '-'}"

            print("  " + msg)
            ws.append([name, f"0x{reg:02X}", mode, msg, "Pass" if ok else "Fail", TesterName, now, "Table9.5"])
            logger_info = f": Test: {name} Response: {msg}, Result: {'Pass' if ok else 'Fail'}, Tester: {TesterName}"
            if ok:
                logger[0].info(loggerName + logger_info)
            else:
                logger[0].warning(loggerName + logger_info)

        except Exception as e:
            msg = f"{name}: REG=0x{reg:02X}, ERROR: {e}"
            print("  " + msg)
            ws.append([name, f"0x{reg:02X}", mode, msg, "Fail", TesterName, now, "Table9.5"])
            logger[0].warning(loggerName + f": Test: {name} Response: {msg}, Result: Fail, Tester: {TesterName}")

        for h in logger[0].handlers:
            try:
                h.flush()
            except Exception:
                pass

    _excel_save()


def run_requirements_coverage():
    """
    Requirement coverage probe:
    - 9.5 / 9.6 / 9.7 / 9.8 / 9.9 are checked via live register reads.
    """
    print("\n[Requirements Coverage] 9.5 / 9.6 / 9.7 / 9.8 / 9.9")
    ws = get_sheet("ReqCoverage")

    for sec, desc, regs in REQUIREMENT_SECTIONS:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not regs:
            msg = f"Section {sec}: {desc} -> PENDING (register map required)"
            print("  " + msg)
            ws.append([sec, desc, "", "", "Pending", TesterName, now, "ReqCoverage"])
            logger[0].warning(loggerName + f": Test: Section {sec} Response: {msg}, Result: Pending, Tester: {TesterName}")
            continue

        total = len(regs)
        passed = 0
        failed_regs = []

        for reg in regs:
            try:
                ce, status, _ = reg_read(reg)
                ok = (ce == 0 and _status_ok_for_test(status, allow_aea=True))
                if ok:
                    passed += 1
                else:
                    failed_regs.append(f"0x{reg:02X}(CE={ce},ST={status})")
            except Exception as e:
                failed_regs.append(f"0x{reg:02X}(ERR={e})")

        result = "Pass" if passed == total else "Partial"
        fail_txt = ",".join(failed_regs) if failed_regs else ""
        msg = f"Section {sec}: {passed}/{total} regs responded"
        if fail_txt:
            msg += f" | issues: {fail_txt}"

        print("  " + msg)
        ws.append([sec, desc, passed, total, result, TesterName, now, fail_txt])

        logger_info = f": Test: Section {sec} Response: {msg}, Result: {result}, Tester: {TesterName}"
        if result == "Pass":
            logger[0].info(loggerName + logger_info)
        else:
            logger[0].warning(loggerName + logger_info)

        for h in logger[0].handlers:
            try:
                h.flush()
            except Exception:
                pass

    _excel_save()

def run_9_6_snapshot_table():
    """
    Table B: Optical Frequency Plan Snapshot (9.6 Passive Reads)

    Reads:
      - Channel (0x30)
      - Grid (0x34)
      - FCF1 (0x35)
      - FCF2 (0x36)
      - LF1 (0x40)
      - LF2 (0x41)
      - LastFreq THz/G10 (0x54, 0x55)
      - LGrid10 (0x56)
      - Fine Tune (0x62)
      - StatusW (0x21)
    """

    print("\n[Table B] 9.6 Optical Snapshot Reads")
    ws = get_sheet("TableB_Snapshot")

    REG_LIST = [
        ("Channel",     0x30),
        ("ChannelH",    0x65),  # new: high word
        ("Grid",        0x34),
        ("Grid2",       0x66),  # new: fine grid
        ("FCF1_THz",    0x35),
        ("FCF2_G10",    0x36),
        ("FCF3_MHz",    0x67),  # new: fine FCF
        ("LF1_THz",     0x40),
        ("LF2_G10",     0x41),
        ("LastF_THz",   0x54),
        ("LastF_G10",   0x55),
        ("LGrid10",     0x56),
        ("FineTuneFTF", 0x62),
        ("StatusW",     0x21),
    ]


    for name, reg in REG_LIST:
        try:
            ce, status, dout = reg_read(reg)
            status_name = _status_name(status)

            pass_fail = "Pass" if (ce == 0 and _status_ok_for_test(status, allow_aea=True)) else "Fail"

            response = (
                f"{name}: REG=0x{reg:02X}, OUT=0x{dout:04X}, "
                f"CE={ce}, STATUS={status} ({status_name})"
            )

        except Exception as e:
            response = f"{name}: ERROR talking to REG=0x{reg:02X}: {e}"
            pass_fail = "Fail"

        print("  " + response)

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([name, response, pass_fail, TesterName, current_datetime, "TableB"])
        # (saving deferred)

        logger_info = f": Test: {name} Response: {response}, Result: {pass_fail}, Tester: {TesterName}"
        if pass_fail == "Pass":
            logger[0].info(loggerName + logger_info)
        else:
            logger[0].warning(loggerName + logger_info)

        for h in logger[0].handlers:
            try: h.flush()
            except: pass

    _excel_save()

def run_9_6_channel_sweep(start_ch=1, end_ch=10, fine_tune_mhz=0):
    """
    Table B Active: 9.6 Channel Sweep Test

    - Programs fine tune.
    - Sweeps channels [start_ch .. end_ch] using 0x30 (Channel).
    - Uses current GRID/GRID2/FCF1/FCF2/FCF3 from the device.
    - Polls NOP/BUSY.
    - Reads LF1/LF2.
    - Compares against a spec-accurate model of the register representation.
    """

    global workbook, worksheet
    print("\n[Table B Active] 9.6 Channel Sweep Test")
    ws = get_sheet("TableB_Active")
    # --- Read configuration registers ---
    _, _, grid      = reg_read(0x34)  # GRID (GHz×10, signed)
    _, _, grid2     = reg_read(0x66)  # GRID2 (MHz, signed)
    _, _, fcf1_thz  = reg_read(0x35)  # FCF1 (THz)
    _, _, fcf2_g10  = reg_read(0x36)  # FCF2 (GHz×10, signed)
    _, _, fcf3_mhz  = reg_read(0x67)  # FCF3 (MHz, signed)

    # --- Ideal (float) view (for information only) ---
    grid_ghz   = (grid / 10.0) + (grid2 / 1000.0)           # GHz
    base_ghz   = (fcf1_thz * 1000.0) + (fcf2_g10 / 10.0) + (fcf3_mhz / 1000.0)
    ftf_ghz    = fine_tune_mhz / 1000.0

    # --- Fixed-point view (GHz × 1e4) to mirror nanoITLA.c ---
    grid_gx1e4 = grid * 1000 + grid2 * 10           # (GHz×10)*1000 + MHz*10
    base_gx1e4 = (
        fcf1_thz * 1000 * 10000 +   # THz → GHz, then ×1e4
        fcf2_g10 * 1000 +           # GHz×10 → ×1e4
        fcf3_mhz * 10               # MHz → ×1e4
    )
    ftf_gx1e4  = fine_tune_mhz * 10                 # MHz → ×1e4

    print(
        f"Using config: GRID={grid_ghz:.6f} GHz, "
        f"FCF={base_ghz:.6f} GHz, FTF={ftf_ghz:.6f} GHz"
    )

    for ch in range(start_ch, end_ch + 1):
        test_name = f"Channel_{ch}"

        try:
            # --- Program channel (MSA-correct 32-bit write: ChannelH then Channel) ---
            _set_channel_32(ch)

            # --- Wait for tune complete (prefer NOP pending bits, spec example) ---
            ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)

            if not ok:
                status_msg = f"CH={ch}: tune completion timeout"
                pass_fail = "Fail"
            else:
                # --- Read LF1/LF2 (and LF3 if implemented) from DUT ---
                _, _, lf1_thz = reg_read(0x40)
                _, _, lf2_g10 = reg_read(0x41)
                try:
                    ce3, st3, lf3_mhz = reg_read(0x68)
                    if ce3 != 0 or st3 != 0:
                        lf3_mhz = 0
                except Exception:
                    lf3_mhz = 0

                # Actual register-reported frequency (GHz, using spec resolution)
                actual_ghz = (lf1_thz * 1000.0) + (lf2_g10 / 10.0) + (lf3_mhz / 1000.0)

                # Ideal optical model (float) – for info only
                ideal_ghz = base_ghz + (ch - 1) * grid_ghz + ftf_ghz

                # Spec-accurate expected register representation (GHz, using fixed-point mirroring)
                total_gx1e4 = base_gx1e4 + (ch - 1) * grid_gx1e4 + ftf_gx1e4
                expected_ghz = float(total_gx1e4) / 10000.0

                # Error in MHz (compare full-resolution GHz values)
                err_mhz = abs(actual_ghz - expected_ghz) * 1000.0

                tol_mhz = 100.0
                pass_fail = "Pass" if err_mhz <= tol_mhz else "Fail"

                status_msg = (
                    f"CH={ch}, Model={ideal_ghz:.6f} GHz, "
                    f"Expected={expected_ghz:.6f} GHz, "
                    f"Actual={actual_ghz:.6f} GHz, "
                    f"Err={err_mhz:.2f} MHz, Tolerance={tol_mhz:.2f} MHz"
                )

        except Exception as e:
            status_msg = f"ERROR Channel={ch}: {e}"
            pass_fail = "Fail"

        # --- Console output ---
        print("  " + status_msg)

        # --- Excel logging (same 6-column layout as before) ---
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            test_name,
            status_msg,
            pass_fail,
            TesterName,
            current_datetime,
            "TableB_Active",
        ])
        # (saving deferred)

        # --- Rotating log ---
        logger_info = (
            f": Test: {test_name} Response: {status_msg}, "
            f"Result: {pass_fail}, Tester: {TesterName}"
        )
        if pass_fail == "Pass":
            logger[0].info(loggerName + logger_info)
        else:
            logger[0].warning(loggerName + logger_info)

        for h in logger[0].handlers:
            try:
                h.flush()
            except Exception:
                pass

    _excel_save()


# ---------------- Full Register Scan ----------------
def run_full_register_scan():
    """Full 0x00–0xFF register read scan.

    For each register:
      - Performs a read via reg_read(reg)
      - Logs CE/STATUS/OUT to console and Excel
      - Marks Pass if CE==0 and STATUS in (0, 2, 3) (OK/AEA/CP).
      - Marks "Not Implemented" if STATUS==XE and the register is not in
        IMPLEMENTED_REGISTERS (expected behaviour — firmware deliberately
        rejects unimplemented addresses).
      - Marks Fail if STATUS==XE on a register that IS in IMPLEMENTED_REGISTERS
        (real firmware defect).
    """
    print("\n[Full Scan] 0x00–0xFF Register Readback")
    ws = get_sheet("FullScan")

    for reg in range(0x00, 0x100):
        name = f"Reg_0x{reg:02X}"
        is_implemented = reg in IMPLEMENTED_REGISTERS
        implemented_str = "Yes" if is_implemented else "No"
        try:
            ce, status, dout = reg_read(reg)
            status_name = _status_name(status)
            response = (
                f"Reg 0x{reg:02X}: OUT=0x{dout:04X}, "
                f"CE={ce}, STATUS={status} ({status_name})"
            )
            # Determine pass/fail, distinguishing unimplemented XE from real failures.
            if IS_UART:
                if ce == 0 and status == 0:
                    pass_fail = "Pass"
                elif status == 1 and not is_implemented:  # XE on non-implemented reg
                    pass_fail = "Not Implemented"
                else:
                    pass_fail = "Fail"
            else:
                if ce == 0 and status in (0, 2):
                    pass_fail = "Pass"
                elif ce == 0 and status == 3:
                    pass_fail = "Info"
                elif status == 1 and not is_implemented:  # XE on non-implemented reg
                    pass_fail = "Not Implemented"
                else:
                    pass_fail = "Fail"
        except Exception as e:
            response = f"Reg 0x{reg:02X}: ERROR talking to device: {e}"
            pass_fail = "Fail"
            ce = ""
            status = ""
            status_name = "ERROR"

        print("  " + response)

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            name,                          # Test
            f"0x{reg:02X}",                # Reg
            ce,                            # CE
            status,                        # STATUS
            status_name,                   # STATUS_Text
            implemented_str,               # Implemented
            pass_fail,                     # OverallResult
            TesterName,                    # Tester
            current_datetime,              # DateTime
            "FullScan",                    # Test Suite
            response                       # Test Description
        ])
        # Periodic save (every 16 regs) to reduce IO but still keep progress.
        if (reg % 16) == 0:
            _excel_save()

        logger_info = f": Test: {name} Response: {response}, Implemented: {implemented_str}, Result: {pass_fail}, Tester: {TesterName}"
        if pass_fail == "Pass":
            logger[0].info(loggerName + logger_info)
        elif pass_fail in ("Info", "Not Implemented"):
            logger[0].info(loggerName + f"[{pass_fail.upper()}]" + logger_info)
        else:
            logger[0].warning(loggerName + logger_info)

        for h in logger[0].handlers:
            try:
                h.flush()
            except Exception:
                pass

    _excel_save()

def run_msa_extension_test():
    """Sanity test for manufacturer-specific MSA extension registers (0x8C–0x91)."""
    print("\n[MSA Extensions] Tuners / SOA / Bias / TEC R/W")
    ws = get_sheet("MSA_Extensions")

    phase_v = 12.12
    ring1_v = 9.99
    ring2_v = 5.550
    soa_v   = 1.23
    bias_v  = 123
    tec_v   = -45

    def _log(name: str, ok: bool, msg: str):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([name, msg, "Pass" if ok else "Fail", TesterName, now, "MSA_Extensions"])

    try:
        write_tuners(phase_v, ring1_v, ring2_v)
        rp = read_tuner("phase")
        r1 = read_tuner("ring1")
        r2 = read_tuner("ring2")
        ok = (abs(rp - phase_v) <= 0.01 and abs(r1 - ring1_v) <= 0.01 and abs(r2 - ring2_v) <= 0.01)
        _log("Tuners", ok, f"Wrote phase={phase_v},ring1={ring1_v},ring2={ring2_v} | Read phase={rp:.3f},ring1={r1:.3f},ring2={r2:.3f}")
        print("  Tuners:", "PASS" if ok else "FAIL")
    except Exception as e:
        _log("Tuners", False, f"ERROR: {e}")
        print("  Tuners: FAIL", e)

    try:
        write_soa(soa_v)
        rs = read_soa()
        ok = (abs(rs - soa_v) <= 0.02)
        _log("SOA", ok, f"Wrote soa={soa_v} | Read soa={rs:.2f}")
        print("  SOA:", "PASS" if ok else "FAIL")
    except Exception as e:
        _log("SOA", False, f"ERROR: {e}")
        print("  SOA: FAIL", e)

    try:
        write_bias(bias_v)
        rb = read_bias()
        ok = (rb == bias_v)
        _log("Bias", ok, f"Wrote bias={bias_v} | Read bias={rb}")
        print("  Bias:", "PASS" if ok else "FAIL")
    except Exception as e:
        _log("Bias", False, f"ERROR: {e}")
        print("  Bias: FAIL", e)

    try:
        write_tec(tec_v)
        rt = read_tec()
        ok = (rt == tec_v)
        _log("TEC", ok, f"Wrote tec={tec_v} | Read tec={rt}")
        print("  TEC:", "PASS" if ok else "FAIL")
    except Exception as e:
        _log("TEC", False, f"ERROR: {e}")
        print("  TEC: FAIL", e)

    _excel_save()
# ---------------- PD Register Test ----------------
def run_pd_test():
    """Simple PD register test.

    Reads the PD-related registers (current mapping uses the LUT debug
    window / PD slots) and logs the scaled float values.
    """
    print("\n[PD Test] MPD / WLPD / WMPD registers")
    ws = get_sheet("PD_Test")

    # Current mapping mirrors nanoITLA.c LUT debug window:
    #   0x86: MPD    (×10)
    #   0x87: WLPD   (×10)
    #   0x88: WMPD   (×10)
    PD_REGS = [
        ("MPD",   0x86, 10.0),
        ("WLPD",  0x87, 10.0),
        ("WMPD",  0x88, 10.0),
        ("WM_PD",     0x89, 10.0),
        ("ETALON_PD", 0x8A, 10.0),
        ("Power_PD",  0x8B, 10.0),
    ]

    for label, reg, scale in PD_REGS:
        try:
            ce, status, dout = reg_read(reg)
            status_name = _status_name(status)
            value = dout / scale
            response = (
                f"{label}: REG=0x{reg:02X}, OUT=0x{dout:04X} "
                f"({value:.2f}), CE={ce}, STATUS={status} ({status_name})"
            )
            if IS_UART:
                pass_fail = "Pass" if (ce == 0 and status == 0) else "Fail"
            else:
                if ce == 0 and status in (0, 2):
                    pass_fail = "Pass"
                elif ce == 0 and status == 3:
                    pass_fail = "Info"
                else:
                    pass_fail = "Fail"
        except Exception as e:
            response = f"{label}: ERROR talking to REG=0x{reg:02X}: {e}"
            pass_fail = "Fail"
            ce = ""
            status = ""
            status_name = "ERROR"

        print("  " + response)

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            label,                        # Test
            f"0x{reg:02X}",              # Reg
            "",                          # Length
            "",                          # ExpectedLen
            "",                          # LenResult
            "",                          # ID_String
            "",                          # Expected_ID_Substr
            "",                          # ID_Result
            ce,                           # CE
            status,                       # STATUS
            status_name,                  # STATUS_Text
            pass_fail,                    # OverallResult
            TesterName,                   # Tester
            current_datetime,             # DateTime
            "PD_Test",                   # Test Suite
            response                      # Test Description
        ])
        # (saving deferred)

        logger_info = f": Test: {label} Response: {response}, Result: {pass_fail}, Tester: {TesterName}"
        if pass_fail == "Pass":
            logger[0].info(loggerName + logger_info)
        elif pass_fail == "Info":
            logger[0].info(loggerName + "[INFO]" + logger_info)
        else:
            logger[0].warning(loggerName + logger_info)

        for h in logger[0].handlers:
            try:
                h.flush()
            except Exception:
                pass

    _excel_save()


# ---------------- OSA Channel Test ----------------
def run_osa_channel_test(start_ch=1, end_ch=10, tol_mhz=100.0):
    """Set Channel + OSA readPeaks() comparison.

    For each channel N in [start_ch..end_ch]:
      1. Write CHANNEL (0x30) = N
      2. Poll StatusW (0x21) BUSY bit clear
      3. Read LASTF (0x54 THz, 0x55 GHz×10) as expected frequency
      4. Call readPeaks() (if available) to get measured frequency
      5. Compare and log Pass/Fail based on tol_mhz.

    NOTE: readPeaks() may return either a GHz float OR a tuple
    (wavelength_nm, power_dBm, smsr_dB). This function will
    coerce either form into GHz.
    """
    print("\n[OSA Test] Set Channel + readPeaks() comparison")
    ws = get_sheet("OSA_Channel")

    # Choose external measurement source:
    #  - Preferred: lab-provided readPeaks() function (if present)
    #  - Fallback: HP wavemeter over GPIB (HP_wavemeter.read_peaks / scalar query)
    use_readpeaks = (readPeaks is not None)
    use_gpib = (Gpib is not None)

    if not use_readpeaks and not use_gpib:
        print("[ERROR] No external measurement source available.")
        print("[ERROR] Either provide a readPeaks() function/module, OR install linux-gpib Python bindings so Gpib is importable.")
        if READPEAKS_IMPORT_ERROR:
            print(f"[DEBUG] readPeaks import error: {READPEAKS_IMPORT_ERROR}")
        return

    if use_readpeaks:
        print("[INFO] Using readPeaks() for external OSA measurement")
    else:
        print("[INFO] Using HP wavemeter over GPIB for external measurement")

    for ch in range(start_ch, end_ch + 1):
        test_name = f"OSA_CH_{ch}"
        try:
            # 1) Program channel (MSA-correct: ChannelH then Channel)
            w65, w30 = _set_channel_32(ch)

            # Validate writes (in UART mode status==1 means XE)
            ce65, st65, _, _ = w65
            ce30, st30, _, _ = w30
            if ce65 != 0 or st65 != 0 or ce30 != 0 or st30 != 0:
                status_msg = (
                    f"CH={ch}: SetChannel write failed "
                    f"(0x65 CE={ce65} ST={st65}; 0x30 CE={ce30} ST={st30})"
                )
                pass_fail = "Fail"
                expected_ghz = None
                osa_ghz = None
                err_mhz = None
                print("  " + status_msg)
                current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([test_name, status_msg, pass_fail, TesterName, current_datetime, "OSA_Channel"])
                continue

            # 2) Wait for tune complete (NOP pending clear preferred)
            ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)

            if not ok:
                status_msg = f"CH={ch}: BUSY timeout before OSA read"
                pass_fail = "Fail"
                expected_ghz = None
                osa_ghz = None
                err_mhz = None
            else:
    # 3) Expected frequency from CURRENT channel frequency registers (LF1/LF2/LF3)
    # Spec: F(GHz) = LF1*10^3 + LF2*10^-1 + LF3*10^-3
                _, _, lf1_thz = reg_read(0x40)
                _, _, lf2_g10 = reg_read(0x41)
                # Optional high-resolution register (LF3 in MHz). If not implemented, treat as 0.
                try:
                    ce3, st3, lf3_mhz = reg_read(0x68)
                    # In UART mode, "not implemented" shows up as XE (status==1) but still returns OUT=0.
                    if ce3 != 0 or st3 != 0:
                        lf3_mhz = 0
                except Exception:
                    lf3_mhz = 0

                expected_ghz = (lf1_thz * 1000.0) + (lf2_g10 / 10.0) + (lf3_mhz / 1000.0)

                # 4) Measured frequency via OSA / wavemeter
                # readPeaks may return:
                #   - GHz float
                #   - (wavelength_nm, power_dBm, smsr)
                time.sleep(0.2)
                if use_readpeaks:
                    rp = readPeaks()
                    osa_ghz = _coerce_readpeaks_to_ghz(rp)
                else:
                    # GPIB fallback: read peak wavelength and convert to GHz
                    wm = _get_hp_wavemeter()
                    wl_nm, pw_dbm, smsr_db = _wm_read_peaks_stable(wm, discard_first=True, settle_s=0.0)
                    wl_nm = float(wl_nm)
                    osa_ghz = _nm_to_ghz(wl_nm)

                # 5) Error in MHz
                err_mhz = abs(osa_ghz - expected_ghz) * 1000.0
                pass_fail = "Pass" if err_mhz <= tol_mhz else "Fail"
                # Note: expected_ghz comes from CURRENT LF1/LF2/LF3; osa_ghz comes from readPeaks() (preferred) or wavemeter conversion.

                status_msg = (
                    f"CH={ch}, Expected={expected_ghz:.6f} GHz, "
                    f"OSA={osa_ghz:.6f} GHz, Err={err_mhz:.2f} MHz, "
                    f"Tolerance={tol_mhz:.2f} MHz"
                )

        except Exception as e:
            status_msg = f"ERROR CH={ch}: {e}"
            pass_fail = "Fail"
            expected_ghz = None
            osa_ghz = None
            err_mhz = None

        print("  " + status_msg)

        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            test_name,                 # Test
            status_msg,                # Test Description
            pass_fail,                 # OverallResult
            TesterName,                # Tester
            current_datetime,          # DateTime
            "OSA_Channel",            # Test Suite
        ])
        # (saving deferred)

        logger_info = (
            f": Test: {test_name} Response: {status_msg}, "
            f"Result: {pass_fail}, Tester: {TesterName}"
        )
        if pass_fail == "Pass":
            logger[0].info(loggerName + logger_info)
        else:
            logger[0].warning(loggerName + logger_info)

        for h in logger[0].handlers:
            try:
                h.flush()
            except Exception:
                pass

    _excel_save()


SETCHANNEL_COMPARE_HEADERS = [
    "Channel",
    "LF1 (0x40) THz",
    "LF2 (0x41) GHz*10",
    "LF3 (0x68) MHz",
    "Laser Frequency (GHz)",
    "WaveMeter PD (0x89)",
    "Etalon PD (0x8A)",
    "Optical Power PD (0x8B)",
    "External Wavemeter Wavelength (nm)",
    "External Wavemeter Frequency (GHz)",
    "External Wavemeter Power (dBm)",
    "External Wavemeter SMSR (dB)",
    "Frequency Error (MHz)",
    "Tolerance (MHz)",
    "Result",
    "Tester",
    "DateTime",
    "Notes",
]

def _reset_sheet_header_exact(ws, header_list):
    """If header row doesn't match EXACTLY, wipe sheet and recreate header."""
    cur = [ws.cell(1, i+1).value for i in range(ws.max_column)]
    expected = list(header_list)
    if cur[:len(expected)] != expected or any(x is not None for x in cur[len(expected):]):
        if ws.max_row >= 1:
            ws.delete_rows(1, ws.max_row)
        ws.append(expected)
        _excel_save()

def _ensure_sheet_setchannel_compare():
    ws = get_sheet("SetChannel_Compare")
    _reset_sheet_header_exact(ws, SETCHANNEL_COMPARE_HEADERS)
    return ws

def run_setchannel_compare_many(start_ch: int = 1, count: int = 20, tol_mhz: float = 100.0, settle_s: float = 0.3, notes: str = ""):
    """Run SetChannel+Compare for a sequence of channels and log each row to Excel."""
    start = int(start_ch)
    n = int(count)
    for ch in range(start, start + n):
        run_setchannel_compare(ch, tol_mhz=float(tol_mhz), settle_s=float(settle_s), notes=notes)
def _env_int(name: str, default: int) -> int:
    s = os.environ.get(name)
    if not s:
        return default
    try:
        return int(s, 0)
    except Exception:
        return default


def run_setchannel_compare_all(tol_mhz: float = 100.0, settle_s: float = 0.3, notes: str = ""):
    """Run Option 9 across ALL channels.

    Channel range is controlled by env vars (so we don't hardcode magic):
      - ITLA_CH_START (default 1)
      - ITLA_CH_COUNT (default 2000)

    Example (20 channels):
      ITLA_CH_START=1 ITLA_CH_COUNT=20 python3 dvt_test_script.py

    This function writes one Excel row per channel in the SetChannel_Compare sheet.
    """
    ch_start = _env_int("ITLA_CH_START", 1)
    ch_count = _env_int("ITLA_CH_COUNT", 52)

    if ch_start < 1:
        ch_start = 1
    if ch_count < 1:
        ch_count = 1

    print(f"\n[Option 9] SetChannel test: start={ch_start} count={ch_count} tol={float(tol_mhz):.2f} MHz")
    run_setchannel_compare_many(
        start_ch=int(ch_start),
        count=int(ch_count),
        tol_mhz=float(tol_mhz),
        settle_s=float(settle_s),
        notes=str(notes or ""),
    )
def _read_pd_triplet():
    ce1, st1, wm_raw = reg_read(0x89)
    ce2, st2, et_raw = reg_read(0x8A)
    ce3, st3, pw_raw = reg_read(0x8B)

    ok1 = (ce1 == 0 and _status_ok_for_test(st1, allow_aea=True))
    ok2 = (ce2 == 0 and _status_ok_for_test(st2, allow_aea=True))
    ok3 = (ce3 == 0 and _status_ok_for_test(st3, allow_aea=True))
    if not (ok1 and ok2 and ok3):
        raise RuntimeError(
            f"PD read failed: WM_PD(ce={ce1},st={st1}) ETALON_PD(ce={ce2},st={st2}) Power_PD(ce={ce3},st={st3})"
        )

    return (
        wm_raw, et_raw, pw_raw,
        float(wm_raw) / 10.0, float(et_raw) / 10.0, float(pw_raw) / 10.0
    )


def run_setchannel_compare(channel: int, tol_mhz: float = 100.0, settle_s: float = 0.3, notes: str = ""):
    """SetChannel + internal PD regs + external wavemeter compare (single channel).

    Flow:
      1) SetChannel(channel) using 32-bit method (0x65 then 0x30)
      2) Wait tune complete
      3) Read DUT Laser Frequency registers: LF1 (0x40), LF2 (0x41), LF3 (0x68)
      4) Read internal PD registers: WaveMeter (0x89), Etalon (0x8A), Optical Power (0x8B)
      5) Read external wavemeter peak (max-power) wavelength (nm), power (dBm), SMSR (dB)
      6) Convert wl->GHz, compare to DUT laser frequency, compute error and Pass/Fail

    Notes:
      - No offsets, no target/biasing. Peak selection is max-power.
      - Discard-first external WM read is allowed to avoid stale/transient sweep buffers.
    """
    ch = int(channel)
    ws = _ensure_sheet_setchannel_compare()
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Defaults for excel row
    lf1_out = lf2_out = lf3_out = ""
    dut_lf_ghz = 0.0
    wm_pd = et_pd = pw_pd = ""
    ext_wl_nm = 0.0
    ext_wm_ghz = 0.0
    ext_p_dbm = -99.0
    ext_smsr_db = 0.0
    err_mhz = 0.0
    result = "Fail"

    # 1) Set channel (MSA-correct: 0x65 then 0x30)
    w65, w30 = _set_channel_32(ch)
    ce65, st65, _, _ = w65
    ce30, st30, _, _ = w30
    if ce65 != 0 or st65 != 0 or ce30 != 0 or st30 != 0:
        msg = f"SetChannel failed (0x65 CE={ce65} ST={st65}; 0x30 CE={ce30} ST={st30})"
        print(f"[SetChannel+Compare] CH={ch}: {msg}")
        ws.append([
            ch,
            lf1_out, lf2_out, lf3_out,
            dut_lf_ghz,
            wm_pd, et_pd, pw_pd,
            ext_wl_nm, ext_wm_ghz, ext_p_dbm, ext_smsr_db,
            err_mhz, float(tol_mhz),
            "Fail",
            TesterName,
            current_datetime,
            (notes or msg),
        ])
        _excel_save()
        return

    # 2) Wait tune complete
    ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)
    if not ok:
        msg = "BUSY timeout after SetChannel"
        print(f"[SetChannel+Compare] CH={ch}: {msg}")
        ws.append([
            ch,
            lf1_out, lf2_out, lf3_out,
            dut_lf_ghz,
            wm_pd, et_pd, pw_pd,
            ext_wl_nm, ext_wm_ghz, ext_p_dbm, ext_smsr_db,
            err_mhz, float(tol_mhz),
            "Fail",
            TesterName,
            current_datetime,
            (notes or msg),
        ])
        _excel_save()
        return

    # Allow a short settle time after tuning
    time.sleep(float(settle_s))

    # 3) DUT reference frequency from LF regs (GHz)
    _, _, lf1_thz = reg_read(0x40)
    _, _, lf2_g10 = reg_read(0x41)
    try:
        ce3, st3, lf3_mhz = reg_read(0x68)
        if ce3 != 0 or st3 != 0:
            lf3_mhz = 0
    except Exception:
        lf3_mhz = 0

    lf1_out = int(lf1_thz)
    lf2_out = int(lf2_g10)
    lf3_out = int(lf3_mhz)
    dut_lf_ghz = (lf1_out * 1000.0) + (lf2_out / 10.0) + (lf3_out / 1000.0)

    # 4) PD registers (logged only)
    try:
        _wm_raw, _et_raw, _pw_raw, wm_pd, et_pd, pw_pd = _read_pd_triplet()
    except Exception as e:
        # Keep running even if PD registers fail; record the error in Notes
        wm_pd = et_pd = pw_pd = ""
        if not notes:
            notes = f"PD read error: {e}"

    # 5) External wavemeter: read max-power peak wavelength (nm), power (dBm), SMSR (dB)
    wm = _get_hp_wavemeter()
    ext_wl_nm, ext_p_dbm, ext_smsr_db = _wm_read_peaks_stable(wm, discard_first=True, settle_s=0.0)
    # Apply lab power correction (attenuator) to external wavemeter power
    # Apply lab power correction (attenuator) to external wavemeter power
    try:
        ext_p_dbm = float(ext_p_dbm) + float(EXT_WM_POWER_OFFSET_DB)
    except Exception:
        pass
    ATTENUATOR_DB = 6.0  # fixed external attenuator at wavemeter input
    ext_wl_nm = float(ext_wl_nm)
    ext_p_dbm = float(ext_p_dbm)
    ext_smsr_db = float(ext_smsr_db)

    # Apply attenuator correction so reported power matches the actual optical power
    ext_p_dbm = ext_p_dbm + ATTENUATOR_DB

    # 6) Convert wl->GHz and compare
    ext_wm_ghz = _nm_to_ghz(ext_wl_nm)
    err_mhz = (ext_wm_ghz - dut_lf_ghz) * 1000.0
    result = "Pass" if abs(err_mhz) <= float(tol_mhz) else "Fail"

    # --- Peak snapshot logging (debug only; no bias / no offsets) ---
    peak_snapshot_str = ""
    try:
        # We are using the HP wavemeter over GPIB in this test.
        # Grab a snapshot of ALL peaks so we can debug wrong-peak selection.
        snap = wm.read_peaks_snapshot()
        peak_snapshot_str = _format_peak_snapshot(snap, selected_wl_nm=ext_wl_nm, max_n=5)
    except Exception as se:
        peak_snapshot_str = f"peaks=<gpib_snapshot_error:{se}>"

    if peak_snapshot_str:
        print(f"[DEBUG] CH={ch} {peak_snapshot_str}")
    # Console summary (include wavelength)
    print(
        f"CH={ch} DUT={dut_lf_ghz:.6f} GHz | ExtWM={ext_wm_ghz:.6f} GHz "
        f"({ext_wl_nm:.6f} nm) | Err={err_mhz:.2f} MHz | Tol={float(tol_mhz):.2f} MHz | Result={result}"
    )

    # Excel row (no RAW PD columns)
    ws.append([
        ch,
        lf1_out,
        lf2_out,
        lf3_out,
        dut_lf_ghz,
        wm_pd,
        et_pd,
        pw_pd,
        ext_wl_nm,
        ext_wm_ghz,
        ext_p_dbm,
        ext_smsr_db,
        err_mhz,
        tol_mhz,
        result,
        TesterName,
        current_datetime,
        (str(notes or "") + (" | " if (notes and peak_snapshot_str) else "") + str(peak_snapshot_str or "")).strip(),
    ])
    _excel_save()

# ---------------- Requirement Test: SetChannel + Internal PD regs + External HP WM ----------------

def _dut_ghz_from_lf_regs() -> float:
    """Read LF1/LF2/LF3 and return DUT frequency in GHz.

    Spec form used in this project:
      DUT_GHz = LF1*1000 + LF2/10 + LF3/1000
      - LF1 (0x40): THz
      - LF2 (0x41): GHz*10
      - LF3 (0x68): MHz
    """
    ce, st, lf1_thz = reg_read(0x40)
    if ce != 0 or st != 0:
        raise RuntimeError(f"LF1 read failed: CE={ce} ST={st} OUT={lf1_thz}")

    ce, st, lf2_g10 = reg_read(0x41)
    if ce != 0 or st != 0:
        raise RuntimeError(f"LF2 read failed: CE={ce} ST={st} OUT={lf2_g10}")

    ce, st, lf3_mhz = reg_read(0x68)
    if ce != 0 or st != 0:
        raise RuntimeError(f"LF3 read failed: CE={ce} ST={st} OUT={lf3_mhz}")

    return (lf1_thz * 1000.0) + (lf2_g10 / 10.0) + (lf3_mhz / 1000.0)


def _ensure_sheet_requirement_setchannel():
    """Ensure Requirement_SetChannel sheet exists and has the correct 14-column header.

    IMPORTANT:
      - Do NOT call get_sheet("Requirement_SetChannel") here, otherwise you can recurse.
      - Work directly with the global openpyxl `workbook`.
    """
    global workbook

    sheet_name = "Requirement_SetChannel"

    desired = [
        "Channel",
        "Target Frequency (THz)",
        "Real Frequency (THz)",
        "External Wavemeter Frequency (THz)",
        "Real Wavelength (nm)",
        "External Wavemeter Wavelength (nm)",
        "Frequency Error (GHz)",
        "Tolerance (GHz)",
        "External Wavemeter Power (dBm)",
        "External Wavemeter SMSR (dB)",
        "Result",
        "Tester",
        "DateTime",
        "Notes",
    ]

    # Ensure workbook exists
    if workbook is None:
        try:
            if os.path.exists(ExcelFileName):
                workbook = load_workbook(ExcelFileName)
            else:
                workbook = Workbook()
        except Exception:
            workbook = Workbook()

    # Create/get sheet directly (NO get_sheet here)
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(title=sheet_name)

    # Compare existing header row to desired; if mismatch, wipe sheet and rewrite header.
    try:
        current = [ws.cell(1, c).value for c in range(1, len(desired) + 1)]
    except Exception:
        current = []

    if current != desired:
        try:
            if ws.max_row >= 1:
                ws.delete_rows(1, ws.max_row)
        except Exception:
            pass
        ws.append(desired)
        _excel_save()

    return ws


def run_requirement_setchannel_test(channels, tol_mhz: float = 1250.0, settle_s: float = 0.3):
    """Requirement SetChannel validation (Option 9) — Excel output is the agreed minimal 14 columns.

    Per channel:
      0) Look up LUT CSV row first.
         - If row missing OR operating point missing (Gain==0 and SOA==0), do NOT touch hardware; log FAIL.
      1) SetChannel (0x65 then 0x30) and wait tune complete
      2) Read DUT Target Frequency (THz) from the frequency plan registers (FCF/GRID/FTF)
      3) Read CSV Real Wavelength (nm) and compute Real Frequency (THz) = 299792.458 / nm
      4) Read external HP wavemeter peak (nm, power, SMSR) and convert to frequency (THz)
      5) Compute Frequency Error (GHz) = (ExternalTHz - RealTHz) * 1000
      6) Apply tolerance in GHz and log Pass/Fail

    Notes:
      - External power correction uses env vars (first match wins):
          EXT_WM_POWER_OFFSET_DB, EXT_WM_ATTEN_DB, ATTENUATOR_DB  (default 10.0)
      - Channel mapping offset: WM_CHANNEL_OFFSET (default 0)
      - Excel sheet: Requirement_SetChannel (14 columns, includes Notes)
    """

    ws = _ensure_sheet_requirement_setchannel()

    # Channel mapping between DUT and CSV
    try:
        ch_offset = int(os.environ.get("WM_CHANNEL_OFFSET", "0"), 0)
    except Exception:
        ch_offset = 0

    # External attenuator/power correction — default +10 dB
    try:
        atten_db = float(
            os.environ.get(
                "EXT_WM_POWER_OFFSET_DB",
                os.environ.get("EXT_WM_ATTEN_DB", os.environ.get("ATTENUATOR_DB", "10"))
            )
        )
    except Exception:
        atten_db = 10.0

    # Load LUT rows (if LUT_CSV_PATH is not set / load fails, we still run but will fail rows with missing LUT)
    try:
        lut_rows = _get_lut_rows()
    except Exception:
        lut_rows = {}

    # Normalize channels input
    if isinstance(channels, int):
        ch_list = [int(channels)]
    else:
        ch_list = [int(c) for c in channels]

    def _lut_lookup(rows: dict, key: int):
        if not rows:
            return None
        if key in rows:
            return rows.get(key)
        sk = str(key)
        if sk in rows:
            return rows.get(sk)
        return None

    try:
        sanity_n = int(os.environ.get("WM_SANITY_COUNT", "5"), 0)
    except Exception:
        sanity_n = 5
    lut_csv_path = os.environ.get("LUT_CSV_PATH", "").strip()
    print(f"[INFO] Option 9: Running channels {ch_list[0]}..{ch_list[-1]} (n={len(ch_list)}) tol={float(tol_mhz) / 1000.0:.3f}GHz settle={float(settle_s):.2f}s")
    if lut_csv_path:
        print(f"[INFO] Option 9: LUT_CSV_PATH={lut_csv_path}")
    else:
        print("[WARN] Option 9: LUT_CSV_PATH is not set")
    if sanity_n > 0:
        print(f"[INFO] Option 9: Sanity preview for first {sanity_n} channels (CSV mapping + expected WL)")

    select_closest = _env_bool("WM_SELECT_CLOSEST_PEAK_DIAG", False)
    log_peaks = _env_bool("WM_LOG_PEAKS", False)
    peak_diag = _env_bool("WM_PEAK_DIAG", False)
    if select_closest:
        print("[INFO] Option 9: WM_SELECT_CLOSEST_PEAK_DIAG=1 (select closest peak to expected)")
    if log_peaks:
        print("[INFO] Option 9: WM_LOG_PEAKS=1 (print peak snapshot per channel)")
    if peak_diag:
        print("[INFO] Option 9: WM_PEAK_DIAG=1 (log strongest/closest/selected peak info)")

    # Summary counters
    total_count = 0
    pass_count = 0
    fail_count = 0
    missing_lut_count = 0
    missing_op_count = 0
    error_count = 0
    max_abs_err_ghz = 0.0
    max_err_ch = None
    max_err_val_ghz = 0.0

    for idx, ch in enumerate(ch_list):
        total_count += 1
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Defaults for Excel
        target_thz = ""
        real_thz = ""
        ext_thz = ""

        real_wl_nm = ""
        ext_wl_nm = ""
        err_thz = ""
        tol_ghz = float(tol_mhz) / 1000.0    # MHz -> THz (for internal compare)
        tol_ghz_str = float(tol_ghz)         # GHz value for Excel display
        ext_p_dbm = ""
        ext_smsr_db = ""
        notes = ""
        result = "Fail"

        csv_ch = int(ch) + int(ch_offset)

        # --- Step 0: LUT row lookup FIRST (skip hardware for missing/zero rows) ---
        lut_row = _lut_lookup(lut_rows, csv_ch)
        if not lut_row:
            missing_lut_count += 1
            fail_count += 1
            ws.append([
                ch,
                target_thz,
                real_thz,
                ext_thz,
                real_wl_nm,
                ext_wl_nm,
                err_thz,
                tol_ghz_str,
                ext_p_dbm,
                ext_smsr_db,
                result,
                TesterName,
                current_datetime,
                notes,
            ])
            _excel_save()
            continue

        # Auto-skip/fail missing operating points (Gain==0 and SOA==0)
        if _lut_operating_point_missing(lut_row):
            missing_op_count += 1
            fail_count += 1
            ws.append([
                ch,
                target_thz,
                real_thz,
                ext_thz,
                real_wl_nm,
                ext_wl_nm,
                err_thz,
                tol_ghz_str,
                ext_p_dbm,
                ext_smsr_db,
                result,
                TesterName,
                current_datetime,
                notes,
            ])
            _excel_save()
            continue
        if sanity_n > 0 and idx < sanity_n:
            try:
                exp_wl_nm = float(lut_row.get("real_wl_nm", 0.0) or 0.0)
                exp_thz = float(lut_row.get("real_freq_thz", 0.0) or 0.0)
                if exp_thz == 0.0 and exp_wl_nm > 0.0:
                    exp_thz = 299792.458 / exp_wl_nm
            except Exception:
                exp_wl_nm = 0.0
                exp_thz = 0.0
            print(
                f"[SANITY] CH={ch} -> CSV_CH={csv_ch} expected_wl={exp_wl_nm:.6f} nm "
                f"expected_freq={exp_thz:.9f} THz"
            )

        try:
            # 1) SetChannel (MSA-correct: 0x65 then 0x30)
            w65, w30 = _set_channel_32(ch)
            ce65, st65, _, _ = w65
            ce30, st30, _, _ = w30
            if ce65 != 0 or st65 != 0 or ce30 != 0 or st30 != 0:
                raise RuntimeError(
                    f"SetChannel write failed (0x65 CE={ce65} ST={st65}; 0x30 CE={ce30} ST={st30})"
                )

            ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)
            if not ok:
                raise RuntimeError("Tune did not complete (BUSY timeout)")

            time.sleep(float(settle_s))

            # 2) Target frequency from DUT frequency-plan regs (THz)
            _target_ghz = _compute_target_frequency_ghz_for_channel(ch)
            _target_thz = float(_target_ghz) / 1000.0

            # 3) Real wavelength from LUT CSV, Real frequency derived from wavelength (THz)
            _real_wl_nm = float(lut_row.get("real_wl_nm", 0.0) or 0.0)
            if _real_wl_nm > 0.0:
                _real_thz = 299792.458 / _real_wl_nm
                _real_ghz = _real_thz * 1000.0
            else:
                _real_thz = 0.0
                _real_ghz = 0.0

            # 4) External wavemeter read (stable helper preferred)
            wm = _get_hp_wavemeter()
            peaks = None
            if select_closest:
                peaks = wm.read_peaks_snapshot()
                if not peaks:
                    raise RuntimeError("No peaks returned from wavemeter snapshot")
                exp_wl_nm = float(_real_wl_nm)
                # Choose closest peak to expected wavelength for diagnostics
                closest = min(peaks, key=lambda x: abs(float(x.get("wl_nm", 0.0)) - exp_wl_nm))
                _wl = float(closest.get("wl_nm", 0.0))
                _p = float(closest.get("pwr_dbm", -99.0))
                # SMSR estimate from snapshot
                if len(peaks) >= 2:
                    sorted_pw = sorted([float(x.get("pwr_dbm", -99.0)) for x in peaks], reverse=True)
                    _smsr = float(sorted_pw[0]) - float(sorted_pw[1])
                else:
                    _smsr = 40.0
                if log_peaks:
                    peak_snapshot_str = _format_peak_snapshot(peaks, selected_wl_nm=_wl, max_n=5)
                    print(f"[PEAKS] CH={ch} " + peak_snapshot_str)
                    notes = (notes + " | " if notes else "") + peak_snapshot_str
            else:
                if "_wm_read_peaks_stable" in globals():
                    _wl, _p, _smsr = _wm_read_peaks_stable(wm, discard_first=True, settle_s=0.0)
                else:
                    try:
                        _wl, _p, _smsr = wm.read_peaks(target_wl_nm=None)
                    except TypeError:
                        _wl, _p, _smsr = wm.read_peaks()
                if log_peaks or peak_diag:
                    try:
                        peaks = wm.read_peaks_snapshot()
                        if log_peaks:
                            peak_snapshot_str = _format_peak_snapshot(peaks, selected_wl_nm=_wl, max_n=5)
                            print(f"[PEAKS] CH={ch} " + peak_snapshot_str)
                            notes = (notes + " | " if notes else "") + peak_snapshot_str
                    except Exception as e:
                        if log_peaks:
                            notes = (notes + " | " if notes else "") + f"peaks=<snapshot_error:{e}>"

            if peak_diag and peaks:
                try:
                    strongest = max(peaks, key=lambda x: float(x.get("pwr_dbm", -99.0)))
                    strongest_wl = float(strongest.get("wl_nm", 0.0))
                    strongest_p = float(strongest.get("pwr_dbm", -99.0))
                    if _real_wl_nm > 0.0:
                        closest = min(peaks, key=lambda x: abs(float(x.get("wl_nm", 0.0)) - float(_real_wl_nm)))
                        closest_wl = float(closest.get("wl_nm", 0.0))
                        closest_p = float(closest.get("pwr_dbm", -99.0))
                        closest_txt = f"{closest_wl:.3f}nm@{closest_p:.2f}dBm"
                    else:
                        closest_txt = "n/a"
                    selected_txt = f"{_wl:.3f}nm@{_p:.2f}dBm"
                    diag_note = (
                        f"diag: strongest={strongest_wl:.3f}nm@{strongest_p:.2f}dBm "
                        f"closest={closest_txt} selected={selected_txt}"
                    )
                    print(f"[DIAG] CH={ch} {diag_note}")
                    notes = (notes + " | " if notes else "") + diag_note
                except Exception as e:
                    print(f"[DIAG] CH={ch} peak diag error: {e}")

            _wl = float(_wl)
            _p = float(_p)
            _smsr = float(_smsr)

            # Apply attenuator/power correction
            _p = _p + float(atten_db)

            # Validate external signal
            if _wl <= 0.0 or _p <= -90.0:
                raise RuntimeError(
                    f"External wavemeter has no valid signal (wl={_wl}, p={_p:.2f}dBm after +{atten_db}dB)"
                )

            _ext_ghz = _nm_to_ghz(_wl)
            _ext_thz = float(_ext_ghz) / 1000.0
            if sanity_n > 0 and idx < sanity_n:
                print(
                    f"[SANITY] CH={ch} measured_wl={_wl:.6f} nm "
                    f"measured_freq={_ext_thz:.9f} THz"
                )

            # 5) Frequency error in THz vs Real frequency (derived from CSV wavelength)
            _err_ghz = (_ext_thz - _real_thz) * 1000.0 
            # Tolerance is specified in MHz on input, but requirements are expressed in GHz.
            # We keep the compare in THz to match the rest of the sheet math.
            #tol_ghz = float(tol_mhz) / 1000.0  # 1250 MHz -> 0.00125 THz (i.e., 1.25 GHz)
            freq_ok = abs(_err_ghz) <= tol_ghz
            result = "Pass" if freq_ok else "Fail"
            if result == "Pass":
                pass_count += 1
            else:
                fail_count += 1

            abs_err_ghz = abs(_err_ghz)
            if abs_err_ghz > max_abs_err_ghz:
                max_abs_err_ghz = abs_err_ghz
                max_err_ch = ch
                max_err_val_ghz = _err_ghz

            # Format for Excel (store as numbers so Excel formatting applies)
            target_thz = float(_target_thz)
            real_thz = float(_real_thz)
            real_wl_nm = float(_real_wl_nm) if _real_wl_nm > 0 else 0.0
            ext_thz = float(_ext_thz)
            ext_wl_nm = float(_wl)
            err_ghz = float(_err_ghz)
            tol_ghz_str = float(tol_ghz)
            ext_p_dbm = float(_p)
            ext_smsr_db = float(_smsr)

            print(
                f"CH={ch} Target={_target_thz:.7f} THz | Real={_real_thz:.7f} THz ({_real_wl_nm:.6f} nm) | "
                f"ExtWM={_ext_thz:.9f} THz (wl={_wl:.6f} nm) | Err={_err_ghz:.3f} GHz | "
                f"Tol={(float(tol_mhz) / 1000.0):.6f} GHz | Result={result} P={_p:.2f}dBm SMSR={_smsr:.2f}dB"
            )

        except Exception as e:
            # Keep defaults and print error; row logs as Fail
            print(f"CH={ch} Result=Fail | ERROR: {e}")
            notes = (notes + " | " if notes else "") + f"ERROR: {e}"
            error_count += 1
            fail_count += 1

        # Excel row (agreed 14 columns)
        ws.append([
            ch,
            target_thz,
            real_thz,
            ext_thz,
            real_wl_nm,
            ext_wl_nm,
            err_ghz,
            tol_ghz_str,
            ext_p_dbm,
            ext_smsr_db,
            result,
            TesterName,
            current_datetime,
            notes,
        ])

        _excel_save()
    if total_count > 0:
        print(
            "[SUMMARY] Option 9: "
            f"total={total_count} pass={pass_count} fail={fail_count} "
            f"missing_lut={missing_lut_count} missing_op={missing_op_count} errors={error_count} "
            f"tol={float(tol_mhz) / 1000.0:.3f}GHz"
        )
        if max_err_ch is not None:
            print(
                "[SUMMARY] Option 9: "
                f"max_error={max_err_val_ghz:.3f} GHz (abs {max_abs_err_ghz:.3f} GHz) at CH={max_err_ch}"
            )

def run_setchannel_compare_sweep(start_ch: int = 1, end_ch: int = 10, tol_mhz: float = 100.0, settle_s: float = 0.3, notes: str = ""):
    for ch in range(int(start_ch), int(end_ch) + 1):
        run_setchannel_compare(ch, tol_mhz=tol_mhz, settle_s=settle_s, notes=notes)
def read_96_model_registers():
    regs = {
        "LASER_EN": 0x65,
        "CH":       0x30,
        "GRID":     0x34,
        "FCF1":     0x35,
        "FCF2":     0x36,
        "GRID2":    0x66,
        "FCF3":     0x67,
        "FTF":      0x62,
    }
    out = {}
    for name, addr in regs.items():
        ce, status, val = reg_read(addr)   # your reg_read returns (CE, STATUS, OUT)
        out[name] = val
    return out

def _ghz_to_nm(ghz: float) -> float:
    # Invert _nm_to_ghz (c = 299792.458 km/s)
    if ghz <= 0:
        return 0.0
    return (299792.458 * 1000.0) / ghz
def _compute_target_frequency_ghz_for_channel(ch: int) -> float:
    """Compute target/model frequency for a channel from DUT frequency plan regs.

    Uses:
      GRID 0x34 (GHz×10, signed)
      GRID2 0x66 (MHz, signed)
      FCF1 0x35 (THz)
      FCF2 0x36 (GHz×10, signed)
      FCF3 0x67 (MHz, signed)
      FTF  0x62 (MHz, signed)  (fine tune)

    Returns:
        float GHz
    """
    try:
        _, _, grid = reg_read(0x34)
        _, _, grid2 = reg_read(0x66)
        _, _, fcf1_thz = reg_read(0x35)
        _, _, fcf2_g10 = reg_read(0x36)
        _, _, fcf3_mhz = reg_read(0x67)
        _, _, ftf_mhz_u16 = reg_read(0x62)

        # signed conversions where applicable
        grid_s = _s16_from_u16(grid)
        grid2_s = _s16_from_u16(grid2)
        fcf2_s = _s16_from_u16(fcf2_g10)
        fcf3_s = _s16_from_u16(fcf3_mhz)
        ftf_mhz = float(_s16_from_u16(ftf_mhz_u16))

        grid_ghz = (grid_s / 10.0) + (grid2_s / 1000.0)
        base_ghz = (float(fcf1_thz) * 1000.0) + (fcf2_s / 10.0) + (fcf3_s / 1000.0)
        ftf_ghz = ftf_mhz / 1000.0

        return base_ghz + (int(ch) - 1) * grid_ghz + ftf_ghz
    except Exception:
        return 0.0

def calibrate_unit_offset_ghz(wm: "HP_wavemeter", expected_map: dict, cal_ch: int, settle_s: float) -> float:
    """Calibrate a per-unit frequency offset using one reference channel.

    We set the DUT to `cal_ch`, measure wavelength, compare to CSV expected,
    and return offset in GHz (measured - expected).

    Env knobs used elsewhere:
      - WM_APPLY_UNIT_OFFSET (default 1)
      - WM_CAL_CH (default 1)
    """
    cal_ch = int(cal_ch)
    if cal_ch not in expected_map:
        raise KeyError(f"Calibration channel {cal_ch} missing from expected_map")

    # 1) Set channel (MSA-correct: 0x65 then 0x30)
    w65, w30 = _set_channel_32(cal_ch)
    ce65, st65, _, _ = w65
    ce30, st30, _, _ = w30
    if ce65 != 0 or st65 != 0 or ce30 != 0 or st30 != 0:
        raise RuntimeError(
            f"SetChannel failed for CAL_CH={cal_ch} (0x65 CE={ce65} ST={st65}; 0x30 CE={ce30} ST={st30})"
        )

    # 2) Wait tune complete
    ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)
    if not ok:
        raise RuntimeError(f"BUSY timeout during calibration for CAL_CH={cal_ch}")

    time.sleep(float(settle_s))

    exp_nm = float(expected_map[cal_ch])

    # 3) One wavemeter snapshot; pick the peak closest to expected
    peaks = wm.read_peaks_snapshot()
    if not peaks:
        raise RuntimeError("No peaks returned from wavemeter during calibration")

    closest = min(peaks, key=lambda x: abs(float(x.get("wl_nm", 0.0)) - exp_nm))
    meas_nm = float(closest["wl_nm"])
  # Guardrails: if the closest peak is still far away or too weak, abort calibration.
    try:
        max_delta_nm = float(os.environ.get("WM_CAL_MAX_DELTA_NM", "0.5"))  # default ±0.5 nm
    except Exception:
        max_delta_nm = 0.5

    try:
        min_power_dbm = float(os.environ.get("WM_MIN_POWER_DBM", "-50"))    # default -50 dBm
    except Exception:
        min_power_dbm = -50.0

    delta_nm = abs(meas_nm - exp_nm)
    meas_p_dbm = float(closest.get("pwr_dbm", -99.0))

    if delta_nm > max_delta_nm:
        raise RuntimeError(
            f"Calibration rejected: closest peak is too far from expected. "
            f"exp={exp_nm:.3f}nm meas={meas_nm:.3f}nm Δ={delta_nm:.3f}nm (max {max_delta_nm:.3f}nm). "
            "Check optical routing / wavemeter input / address / wrong source."
        )

    if meas_p_dbm < min_power_dbm:
        raise RuntimeError(
            f"Calibration rejected: peak power too low ({meas_p_dbm:.2f} dBm < {min_power_dbm:.2f} dBm). "
            "Likely no light at the wavemeter input."
        )
    exp_ghz = _nm_to_ghz(exp_nm)
    meas_ghz = _nm_to_ghz(meas_nm)

    offset_ghz = meas_ghz - exp_ghz
    print(f"[CAL] CH={cal_ch} exp={exp_nm:.3f}nm meas={meas_nm:.3f}nm Δ={delta_nm:.3f}nm p={meas_p_dbm:.2f}dBm offset={offset_ghz:.6f}GHz")
    return float(offset_ghz)

def run_wm_compare_to_csv(start_ch: int, end_ch: int, tol_mhz: float = 100.0, settle_s: float = 0.3):
    """Set Channel + HP wavemeter (GPIB) compare to CSV expected 'Real Wavelength'.

    Uses:
      - Expected wavelength per channel loaded from LUT_CSV_PATH
      - Measured wavelength via HP_wavemeter.read_peaks() (nm)

    Comparison:
      - Converts both expected and measured nm -> GHz and computes MHz error.

    Notes:
      - This test validates *optical output* against your golden table.
      - It does NOT require adding expected values into nano_lut_50ghz.c.
    """
    print("\n[WM Compare] Set Channel + HP 86120C compare to CSV expected")
    if STRICT_NO_MAGIC:
        # Force strict behaviour regardless of other env knobs
        if OPTICAL_TUNING_MODE:
            print("[STRICT] STRICT_NO_MAGIC=1: ignoring OPTICAL_TUNING_MODE")
        print("[STRICT] No unit offsets. No closest-peak bias. Direct WM vs CSV comparison.")


    # Tolerance precedence:
    #   1) TUNING_TOL_MHZ (if set)
    #   2) WM_TOL_MHZ
    #   3) function argument
    try:
        tol_mhz = float(os.environ.get("TUNING_TOL_MHZ",
                      os.environ.get("WM_TOL_MHZ", str(tol_mhz))))
    except Exception:
        tol_mhz = float(tol_mhz)

    # If OPTICAL_TUNING_MODE is enabled, keep the same tolerance value (already covered by env)
    if OPTICAL_TUNING_MODE:
        try:
            tol_mhz = float(os.environ.get("TUNING_TOL_MHZ", str(tol_mhz)))
        except Exception:
            tol_mhz = float(tol_mhz)

    try:
        settle_s = float(os.environ.get("WM_SETTLE_S", str(settle_s)))
    except Exception:
        settle_s = float(settle_s)

    def _want_closest_peak() -> bool:
        if STRICT_NO_MAGIC:
            return False
        return _env_bool("WM_SELECT_CLOSEST_PEAK", False)
    # CSV channel mapping can differ from DUT channel numbering.
    # Example: if DUT CH=1 corresponds to CSV CH=4, set WM_CHANNEL_OFFSET=3.
    try:
        ch_offset = int(os.environ.get("WM_CHANNEL_OFFSET", "0"), 0)
    except Exception:
        ch_offset = 0
    lut_rows = _get_lut_rows()
    ws = _ensure_sheet_wm_compare()

    # Open the HP wavemeter once for the whole run
    if Gpib is None:
        raise RuntimeError(
            "Gpib module not available. Install linux-gpib python bindings to use WM compare."
        )

    # Allow either a VISA-like string or board+address
    gpib_str = os.environ.get("WM_GPIB", "GPIB0::18::INSTR").strip()
    try:
        if gpib_str:
            wm = HP_wavemeter(gpib=gpib_str)
        else:
            gpib_addr = _env_int("WM_GPIB_ADDR", 18)
            gpib_board = _env_int("WM_GPIB_BOARD", 0)
            wm = HP_wavemeter(gpib_addr=gpib_addr, board=gpib_board)
    except Exception as e:
        msg = f"Wavemeter GPIB init failed: {e}"
        print(f"[ERROR] {msg}")
        now0 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            ws.append(["", "", "", "", "", "", tol_mhz, "", "", "ERROR", TesterName, now0, msg])
            _excel_save()
        except Exception:
            pass
        print("[HINT] Verify linux-gpib sees the interface/device:")
        print("       ls -l /dev/gpib*    (should show /dev/gpib0)")
        print("       gpib_config -l      (list configured boards/devices)")
        print("       ibtest              (try talking to address 18 on board 0)")
        print("[HINT] If the address is wrong, set WM_GPIB='GPIB0::<ADDR>::INSTR' or WM_GPIB_ADDR=<ADDR>.")
        return

    model = read_96_model_registers()
    print("[9.6 PRE-FLIGHT] ", model)

    # Log the 9.6 register snapshot once into the WM_Compare_CSV sheet
    now0 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(["", "", "", "", "", "", tol_mhz, "", "", "INFO", TesterName, now0, f"9.6 preflight regs: {model}"])

    # Optional: per-unit offset calibration (measured - expected) in GHz
    apply_offset = os.environ.get("WM_APPLY_UNIT_OFFSET", "1").strip() not in ("0", "false", "no", "off")
    offset_ghz = 0.0
    if apply_offset:
        try:
            cal_ch = int(os.environ.get("WM_CAL_CH", "1"), 0)
        except Exception:
            cal_ch = 1
        try:
            offset_ghz = calibrate_unit_offset_ghz(wm, expected_map, cal_ch=cal_ch, settle_s=settle_s)
            ws.append(["", "", "", "", "", "", tol_mhz, "", "", "INFO", TesterName, now0, f"Applied unit offset: {offset_ghz:.6f} GHz (from CAL_CH={cal_ch})"])
        except Exception as e:
            offset_ghz = 0.0
            ws.append(["", "", "", "", "", "", tol_mhz, "", "", "WARN", TesterName, now0, f"Unit offset calibration skipped/failed: {e}"])

    for ch in range(int(start_ch), int(end_ch) + 1):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        notes = ""

        # Apply CSV channel offset mapping
        csv_ch = int(ch) + int(ch_offset)

        lut_row = lut_rows.get(csv_ch)

        # Missing LUT row → auto FAIL
        if not lut_row:
            notes = f"No LUT row for CSV_CH={csv_ch} (DUT_CH={ch}, offset={ch_offset})"
            ws.append([ch, "", "", "", "", "", tol_mhz, "", "", "Fail", TesterName, now, notes])
            continue

        # All-zero / missing operating point → auto FAIL
        if _lut_operating_point_missing(lut_row):
            notes = (
                f"LUT operating point missing (Gain={lut_row.get('gain',0)}, "
                f"SOA={lut_row.get('soa',0)}) — skipped"
            )
            ws.append([ch, "", "", "", "", "", tol_mhz, "", "", "Fail", TesterName, now, notes])
            continue

        # Expected wavelength from LUT CSV
        exp_wl_nm = float(lut_row.get("real_wl_nm", 0.0))
        # Apply calibrated per-unit offset (measured - expected) in GHz by shifting expected
        if offset_ghz != 0.0:
            exp_ghz_base = _nm_to_ghz(exp_wl_nm)
            exp_ghz_adj = exp_ghz_base + float(offset_ghz)
            exp_wl_nm_adj = _ghz_to_nm(exp_ghz_adj)
        else:
            exp_wl_nm_adj = exp_wl_nm

        # 1) Set channel (MSA correct: 0x65 then 0x30)
        try:
            w65, w30 = _set_channel_32(ch)
            ce65, st65, _, _ = w65
            ce30, st30, _, _ = w30
            if ce65 != 0 or st65 != 0 or ce30 != 0 or st30 != 0:
                notes = f"SetChannel write failed (0x65 CE={ce65} ST={st65}; 0x30 CE={ce30} ST={st30})"
                print(f"  CH={ch}: FAIL ({notes})")
                ws.append([ch, exp_wl_nm, "", _nm_to_ghz(exp_wl_nm), "", "", tol_mhz, "", "", "Fail", TesterName, now, notes])
                continue

            # 2) Wait tune complete
            ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)
            if not ok:
                notes = "BUSY timeout before wavemeter read"
                print(f"  CH={ch}: FAIL ({notes})")
                ws.append([ch, exp_wl_nm, "", _nm_to_ghz(exp_wl_nm), "", "", tol_mhz, "", "", "Fail", TesterName, now, notes])
                continue

            # Let things settle
            time.sleep(float(settle_s))

            # --- One wavemeter snapshot; pick peak ---
            peaks = wm.read_peaks_snapshot()
            if not peaks:
                raise RuntimeError("No peaks returned from wavemeter")

            if _want_closest_peak():
                # Choose the peak closest to the CSV expected wavelength (tuning aid)
                closest = min(peaks, key=lambda x: abs(float(x.get("wl_nm", 0.0)) - float(exp_nm)))
                meas_nm = float(closest.get("wl_nm", 0.0))
                meas_p_dbm = float(closest.get("pwr_dbm", -99.0))
            else:
                # Strict/validation: choose the highest-power peak (no bias)
                best = max(peaks, key=lambda x: float(x.get("pwr_dbm", -99.0)))
                meas_nm = float(best.get("wl_nm", 0.0))
                meas_p_dbm = float(best.get("pwr_dbm", -99.0))

            # Strongest peak (max power) from snapshot
            strongest = max(peaks, key=lambda x: x["pwr_dbm"])
            max_wl_nm = float(strongest["wl_nm"])
            max_pw_dbm = float(strongest["pwr_dbm"])

            # Closest-to-expected peak from snapshot
            closest = min(peaks, key=lambda x: abs(float(x["wl_nm"]) - float(exp_wl_nm_adj)))
            pick_wl_nm = float(closest["wl_nm"])
            pick_pw_dbm = float(closest["pwr_dbm"])

            # Peak selection policy:
            # - TUNING: choose the peak closest to expected wavelength
            # - VALIDATION: choose the strongest peak (max power)
            if OPTICAL_TUNING_MODE:
                chosen_wl_nm = float(pick_wl_nm)
                chosen_pw_dbm = float(pick_pw_dbm)
            else:
                chosen_wl_nm = float(max_wl_nm)
                chosen_pw_dbm = float(max_pw_dbm)

            # SMSR estimate from snapshot: strongest minus runner-up
            if len(peaks) >= 2:
                sorted_pw = sorted([float(x["pwr_dbm"]) for x in peaks], reverse=True)
                pick_smsr_db = float(sorted_pw[0]) - float(sorted_pw[1])
            else:
                pick_smsr_db = 40.0

            # Reject if chosen (closest) peak is much weaker than max-power peak
            try:
                max_delta_db = float(os.environ.get("WM_PEAK_MAX_DELTA_DB", "8.0"))
            except Exception:
                max_delta_db = 8.0

            delta_db = float(max_pw_dbm) - float(pick_pw_dbm)
            if delta_db > max_delta_db:
                notes = (notes + " | " if notes else "") + (
                    f"Multi-peak: strongest={max_wl_nm:.3f}nm@{max_pw_dbm:.2f}dBm; "
                    f"picked={pick_wl_nm:.3f}nm@{pick_pw_dbm:.2f}dBm; Δ={delta_db:.2f}dB>thr{max_delta_db:.1f}"
                )

            # --- DEBUG: compare DUT internal LF vs CSV expected ---
            try:
                _, _, lf1 = reg_read(0x40)
                _, _, lf2 = reg_read(0x41)
                lf_thz = lf1 + (lf2 / 10.0) / 1000.0
                lf_nm = 299792.458 / lf_thz
                notes = (notes + " | " if notes else "") + f"LF_nm={lf_nm:.3f}"
            except Exception:
                pass

            # --- CSV vs wavemeter comparison ---
            # Use the chosen peak for error calculation against the CSV expected wavelength
            exp_ghz = _nm_to_ghz(exp_wl_nm)
            meas_ghz = _nm_to_ghz(float(chosen_wl_nm))
            err_mhz = abs(meas_ghz - exp_ghz) * 1000.0

            # Keep existing variable names for downstream logging
            meas_wl_nm, meas_pw_dbm, meas_smsr_db = chosen_wl_nm, chosen_pw_dbm, pick_smsr_db

            result = "Pass" if err_mhz <= float(tol_mhz) else "Fail"

            # Only force-fail multi-peak if the chosen peak is ambiguous AND already failing
            # (prevents false FAIL when closest peak is correct but not strongest)
            if "Multi-peak:" in (notes or "") and result != "Pass":
                result = "Fail"

            # Include CSV channel mapping info in notes for traceability
            map_note = f"CSV_CH={csv_ch} (offset={ch_offset})"
            notes = (notes + " | " if notes else "") + map_note

            mode = "TUNING" if OPTICAL_TUNING_MODE else "VALIDATION"
            print(
                f"  [{mode}] CH={ch}: Expected={exp_wl_nm:.3f} nm ({exp_ghz:.6f} GHz) | "
                f"Measured={float(meas_wl_nm):.3f} nm ({meas_ghz:.6f} GHz) | "
                f"Err={err_mhz:.2f} MHz (Tol={float(tol_mhz):.2f}) => {result}"
            )

            ws.append([
                ch,
                round(exp_wl_nm, 6),
                round(float(meas_wl_nm), 6),
                round(exp_ghz, 6),
                round(meas_ghz, 6),
                round(err_mhz, 2),
                float(tol_mhz),
                float(meas_pw_dbm),
                float(meas_smsr_db),
                result,
                TesterName,
                now,
                notes,
            ])

        except Exception as e:
            notes = f"Exception: {e} | CSV_CH={csv_ch} (offset={ch_offset})"
            print(f"  CH={ch}: FAIL ({notes})")
            ws.append([
                ch,
                round(exp_wl_nm, 6) if 'exp_wl_nm' in locals() else "",
                "",
                round(_nm_to_ghz(exp_wl_nm), 6) if 'exp_wl_nm' in locals() else "",
                "",
                "",
                float(tol_mhz),
                "",
                "",
                "Fail",
                TesterName,
                now,
                notes,
            ])

        # periodic save to keep progress
        if (ch % 5) == 0:
            _excel_save()

    _excel_save()


def interactive_reg_io():
    """
    Single-session REPL (type 'q' to quit)

    Commands:
      r <reg>                 # read (hex or decimal)
      <reg>                   # read (shortcut)
      w <reg> <data>          # write
      <reg> <data>            # write (shortcut)
    """
    print("\nInteractive Register I/O (type 'q' to quit)")
    print("Examples: r 0x30  |  0x30  |  w 0x34 50  |  0x34 0x0032")

    # Labels for our temporary LUT debug window (0x80–0x88)
    lut_labels = {
        0x80: "V1",
        0x81: "V2",
        0x82: "V3",
        0x83: "Gain",
        0x84: "SOA",
        0x85: "Temp",
        0x86: "MPD",
        0x87: "WLPD",
        0x88: "WMPD",
        # New explicit PD regs (after 0x88)
        0x89: "WM_PD",
        0x8A: "ETALON_PD",
        0x8B: "Power_PD",
    }


    while True:
        cmd = input("> ").strip()
        if not cmd:
            continue
        if cmd.lower() == 'q':
            break

        parts = cmd.split()
        op = parts[0].lower()

        # READ forms
        if (op == 'r' and len(parts) == 2) or (len(parts) == 1 and op != 'w'):
            reg_s = parts[1] if op == 'r' else parts[0]
            try:
                reg = int(reg_s, 0)
                if not (0 <= reg <= 0xFF):
                    raise ValueError("reg out of range (0..0xFF)")
            except Exception as e:
                print(f"Invalid reg: {e}")
                continue

            ce, status, dout = reg_read(reg)

            # Base response
            extra = ""
            # If this is one of our LUT debug registers, decode the scaled value
            if reg in lut_labels:
                label = lut_labels[reg]

                # 0x80–0x84: voltages / gain / SOA in centi-units (×100)
                # 0x85–0x88: temperature & PDs in deci-units (×10)
                if reg == 0x85:  # Temp (°C × 10)
                    value = dout / 10.0
                    extra = f" ({label}={value:.1f} °C)"
                elif reg in (0x86, 0x87, 0x88, 0x89, 0x8A, 0x8B):  # MPD / WLPD / WMPD (×10)
                    # PD regs (×10): MPD/WLPD/WMPD and WM_PD/ETALON_PD/Power_PD
                    value = dout / 10.0
                    extra = f" ({label}={value:.1f})"
                else:  # V1/V2/V3/Gain/SOA (×100)
                    value = dout / 100.0
                    extra = f" ({label}={value:.2f})"

            # 9.5-friendly decode for status/trigger/threshold registers.
            d95 = _decode_9_5_value(reg, dout)
            if d95:
                extra += f" ({d95})"
            if reg in (0x57, 0x58) and ce == 0 and status == 0:
                d98 = decode_9_8_aea(reg, dout)
                if d98:
                    extra += f" ({d98})"

            response = (
                f"Read REG=0x{reg:02X}, OUT=0x{dout:04X}{extra}, "
                f"CE={ce}, STATUS={status} ({_status_name(status)})"
            )

            # Pass/Info/Fail decision
            if IS_UART:
                pass_fail = "Pass" if ce == 0 and status == 0 else "Fail"
            else:
                pass_fail = "Pass" if ce == 0 and status in (0, 2) else ("Info" if status == 3 else "Fail")

            print(f"Response: {response}")

            current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws = get_sheet("Manual")
            ws.append([
                f"Read 0x{reg:02X}",   # Test
                f"0x{reg:02X}",        # Reg
                "",                    # Length
                "",                    # ExpectedLen
                "",                    # LenResult
                "",                    # ID_String
                "",                    # Expected_ID_Substr
                "",                    # ID_Result
                ce,                    # CE
                status,                # STATUS
                _status_name(status),  # STATUS_Text
                pass_fail,             # OverallResult
                TesterName,            # Tester
                current_datetime,      # DateTime
                "Manual",              # Test Suite
                response               # Test Description
            ])
            workbook.save(ExcelFileName)

            logger_info = f": Test : Read 0x{reg:02X} Response : {response}, Result : {pass_fail}, Tester : {TesterName}"
            if pass_fail == "Pass":
                logger[0].info(loggerName + logger_info)
            elif pass_fail == "Info":
                logger[0].info(loggerName + "[INFO]" + logger_info)
            else:
                logger[0].warning(loggerName + logger_info)
            for h in logger[0].handlers:
                try:
                    h.flush()
                except Exception:
                    pass
            continue

        # WRITE forms
        if (op == 'w' and len(parts) == 3) or (len(parts) == 2 and op != 'r'):
            if op == 'w':
                if len(parts) != 3:
                    print("Usage: w <reg> <data>   (hex or decimal)")
                    continue
                reg_s, data_s = parts[1], parts[2]
            else:
                if len(parts) != 2:
                    print("Usage: <reg> <data>   (hex or decimal)   OR   w <reg> <data>")
                    continue
                reg_s, data_s = parts[0], parts[1]
            try:
                reg  = int(reg_s, 0)
                data = int(data_s, 0)
                if not (0 <= reg <= 0xFF):    raise ValueError("reg out of range (0..0xFF)")
                if not (0 <= data <= 0xFFFF): raise ValueError("data out of range (0..0xFFFF)")
            except Exception as e:
                print(f"Invalid reg/data: {e}")
                continue

            ce, status, din, dout = reg_write(reg, data)
            extra_in = ""
            if reg == 0x30:
                extra_in = f" (Channel={din})"
            d95_in = _decode_9_5_value(reg, din)
            d95_out = _decode_9_5_value(reg, dout)
            if d95_in:
                extra_in += f" (W:{d95_in})"
            extra_out = f" (R:{d95_out})" if d95_out else ""

            response = (
                f"Write REG=0x{reg:02X}, IN=0x{din:04X}{extra_in}, OUT=0x{dout:04X}{extra_out}, "
                f"CE={ce}, STATUS={status} ({_status_name(status)})"
            )
            if IS_UART:
                pass_fail = "Pass" if ce == 0 and status == 0 else "Fail"
            else:
                pass_fail = "Pass" if ce == 0 and status in (0, 2) else ("Info" if status == 3 else "Fail")
            print(f"Response: {response}")

            current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws = get_sheet("Manual")
            ws.append([
                f"Write 0x{reg:02X}",  # Test
                f"0x{reg:02X}",        # Reg
                "",                    # Length
                "",                    # ExpectedLen
                "",                    # LenResult
                "",                    # ID_String
                "",                    # Expected_ID_Substr
                "",                    # ID_Result
                ce,                    # CE
                status,                # STATUS
                _status_name(status),  # STATUS_Text
                pass_fail,             # OverallResult
                TesterName,            # Tester
                current_datetime,      # DateTime
                "Manual",              # Test Suite
                response               # Test Description
            ])
            workbook.save(ExcelFileName)

            logger_info = f": Test : Write 0x{reg:02X} Response : {response}, Result : {pass_fail}, Tester : {TesterName}"
            if pass_fail == "Pass":
                logger[0].info(loggerName + logger_info)
            elif pass_fail == "Info":
                logger[0].info(loggerName + "[INFO]" + logger_info)
            else:
                logger[0].warning(loggerName + logger_info)
            for h in logger[0].handlers:
                try:
                    h.flush()
                except Exception:
                    pass
            continue

        print("Usage: r <reg> | <reg> | w <reg> <data> | <reg> <data>   (hex or decimal)")


# ---------------- Batch table runner (legacy DVT_TableA/B harness) ----------------
def checkDVTTable(DVT_Table, test_suite):
    ws = get_sheet(test_suite if test_suite else "LegacyB")
    global TesterName
    lenTable = len(DVT_Table)
    i = 0
    while i < lenTable:
        test = DVT_Table[i]
        arg1 = DVT_Table[i + 1]
        arg2 = DVT_Table[i + 2]
        arg3 = DVT_Table[i + 3]
        if arg3 == "Enable":
            if test == "test1":
                response, err = test1(arg1, arg2)
            elif test == "test2":
                response, err = test2(arg1, arg2)
            elif test == "test3":
                response, err = test3(arg1, arg2)
            elif test == "test4":
                response, err = test4(arg1, arg2)
            elif test == "test5":
                response, err = test5(arg1, arg2)
            elif test == "test6":
                response, err = test6(arg1, arg2)
            elif test == "test7":
                response, err = test7(arg1, arg2)
            elif test == "test8":
                response, err = test8(arg1, arg2)
            pass_fail = "Pass" if err == 0 else "Fail"
            logger_info = f": Test: {test} Response: {response}, Result: {pass_fail}, Tester: {TesterName}"
            logger[0].info(loggerName + logger_info)
            sprint(f"Test: {test} Response: {response}, Result: {pass_fail}, Tester: {TesterName}", 0)
        else:
            response = "Test not enabled"
            pass_fail = "Fail"
            logger_info = f": Test: {test}, Response: {response}, Result: {pass_fail}, Tester: {TesterName}"
            sprint(f"Test: {test}, Response: {response}", 0)
            logger[0].warning(loggerName + logger_info)
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([
            test,              # Test
            "",                # Reg
            "",                # Length
            "",                # ExpectedLen
            "",                # LenResult
            "",                # ID_String
            "",                # Expected_ID_Substr
            "",                # ID_Result
            "",                # CE
            "",                # STATUS
            "",                # STATUS_Text
            pass_fail,         # OverallResult
            TesterName,        # Tester
            current_datetime,  # DateTime
            test_suite,        # Test Suite (e.g. "B")
            response           # Test Description
        ])
        i += 4
    _excel_save()

# ---------------- Pilot Laboratory Backend Hooks ----------------
# These hooks allow Test 9 to run on the Pi without Windows/Keil.
# We use the HP/Agilent 86120C wavemeter (confirmed via *IDN?) as the external instrument.
# NOTE: The instrument provides wavelength + optical power (+ derived SMSR from peak list).
# The names below are kept for historical reasons, but we clearly document what is returned.

_WM_SINGLETON = None


def _get_hp_wavemeter() -> "HP_wavemeter":
    """Return a cached HP_wavemeter instance."""
    global _WM_SINGLETON
    if _WM_SINGLETON is not None:
        return _WM_SINGLETON

    if Gpib is None:
        raise NotImplementedError("Gpib module not available; cannot use HP wavemeter over GPIB")

    addr = _env_int("WAVEMETER_GPIB_ADDR", 18)
    board = _env_int("WAVEMETER_GPIB_BOARD", 0)
    # Helpful diagnostic: if /dev/gpib0 is not accessible, advise the user.
    try:
        if os.path.exists("/dev/gpib0") and not os.access("/dev/gpib0", os.R_OK | os.W_OK):
            print("[WARN] /dev/gpib0 is not accessible without elevated permissions. Use 'sudo -E' or fix udev/group perms.")
    except Exception:
        pass
    _WM_SINGLETON = HP_wavemeter(gpib_addr=addr, board=board)
    return _WM_SINGLETON


def read_wavemeter_reg():
    """External wavemeter reading used by Test 9.

    Returns:
        float: Peak wavelength in nm (from HP 86120C scalar query or peak list).

    Why wavelength_nm?
      - The 86120C is fundamentally a wavemeter; wavelength is the most direct measurement.
      - Test 9 also separately reads the external frequency (GHz) via read_external_wavemeter().
    """
    wm = _get_hp_wavemeter()
    try:
        return float(wm.read_wavelength_nm_max())
    except Exception:
        wl_nm, pw_dbm, smsr_db = wm.read_peaks()
        return float(wl_nm)


def read_power_reg():
    """External optical power reading proxy for Test 9.

    Returns:
        float: Peak optical power in dBm (from HP 86120C peak list).
    """
    wm = _get_hp_wavemeter()
    try:
        return float(wm.read_power_dbm_max())
    except Exception:
        wl_nm, pw_dbm, smsr_db = wm.read_peaks()
        return float(pw_dbm)


def read_etalon_reg():
    """External quality reading proxy for Test 9.

    The bench may have a separate etalon monitor; on the Pi we don't.
    As a practical proxy, return SMSR in dB derived from the peak list.

    Returns:
        float: SMSR in dB.
    """
    wm = _get_hp_wavemeter()
    try:
        wl_nm, pw_dbm, smsr_db = wm.read_peaks()
        return float(smsr_db)
    except Exception:
        return 0.0

def read_external_wavemeter():
    """Read external wavemeter frequency in GHz.

    Priority:
      1) If readPeaks() is available, use it (and coerce to GHz).
      2) Else, if Gpib is available, use HP_wavemeter over GPIB.

    Configure:
      - WAVEMETER_GPIB_ADDR (default 18)
      - WAVEMETER_GPIB_BOARD (default 0)
    """
    if readPeaks is not None:
        return _coerce_readpeaks_to_ghz(readPeaks())

    if Gpib is None:
        raise NotImplementedError("External wavemeter not available: no readPeaks and no Gpib module")

    wm = _get_hp_wavemeter()
    try:
        wl_nm = wm.read_wavelength_nm_max()
    except Exception:
        wl_nm, pw_dbm, smsr_db = wm.read_peaks()
    return _nm_to_ghz(float(wl_nm))
    
def run_full_setchannel_test(start_ch=1, end_ch=10, tol_mhz=100.0):
    print("\n[Full SetChannel Test] Internal PD + External WM Comparison")
    ws = get_sheet("SetChannelFull")

    for ch in range(start_ch, end_ch + 1):
        test_name = f"SCH_{ch}"
        pass_fail = "Fail"
        result = ""

        try:
            # 1) Write channel (MSA-correct: ChannelH then Channel)
            w65, w30 = _set_channel_32(ch)
            ce65, st65, _, _ = w65
            ce30, st30, _, _ = w30
            if ce65 != 0 or st65 != 0 or ce30 != 0 or st30 != 0:
                result = f"CH={ch}: SetChannel write failed (0x65 CE={ce65} ST={st65}; 0x30 CE={ce30} ST={st30})"
                pass_fail = "Fail"
            else:
                # 2) Wait for tune complete
                ok = _wait_tune_complete(timeout_s=30.0, poll_s=0.1)

            if 'ok' in locals() and not ok:
                result = f"CH={ch}: BUSY timeout"
                pass_fail = "Fail"
            elif 'ok' in locals() and ok:
                # 3) Read internal PD registers (scaled ×10)
                # New explicit regs (0x89–0x8B)
                _, _, wm_pd_raw_reg     = reg_read(0x89)
                _, _, etalon_pd_raw_reg = reg_read(0x8A)
                _, _, power_pd_raw_reg  = reg_read(0x8B)

                wm_pd_reg     = wm_pd_raw_reg / 10.0
                etalon_pd_reg = etalon_pd_raw_reg / 10.0
                power_pd_reg  = power_pd_raw_reg / 10.0

                # Legacy LUT slots (0x86–0x88) for cross-check
                _, _, mpd_raw  = reg_read(0x86)
                _, _, wlpd_raw = reg_read(0x87)
                _, _, wmpd_raw = reg_read(0x88)

                mpd_reg  = mpd_raw / 10.0
                wlpd_reg = wlpd_raw / 10.0
                wmpd_reg = wmpd_raw / 10.0

                # 4) Read expected frequency from CURRENT channel frequency registers (LF1/LF2/LF3)
                # Spec: F(GHz) = LF1*10^3 + LF2*10^-1 + LF3*10^-3
                _, _, lf1 = reg_read(0x40)
                _, _, lf2 = reg_read(0x41)
                # Optional high-resolution register (LF3 in MHz). If not implemented, treat as 0.
                try:
                    _, _, lf3 = reg_read(0x68)
                except Exception:
                    lf3 = 0
                expected_ghz = (lf1 * 1000.0) + (lf2 / 10.0) + (lf3 / 1000.0)

                # 5) External measurements (Pilot backend hooks)
                # NOTE: these are placeholders until Jorge/Chris provides implementations.
                # External OSA measurement (preferred): use readPeaks() on the Pi
                if readPeaks is not None:
                    rp = readPeaks()
                    # rp can be GHz OR (wl_nm, pw_dbm, smsr)
                    if isinstance(rp, (tuple, list)) and len(rp) >= 3:
                        ext_wl_nm = float(rp[0])
                        power_pd_ext = float(rp[1])
                        etalon_pd_ext = float(rp[2])
                        ext_wm_ghz = _nm_to_ghz(ext_wl_nm) if 1000.0 <= ext_wl_nm <= 2000.0 else float(rp[0])
                    else:
                        ext_wm_ghz = _coerce_readpeaks_to_ghz(rp)
                        ext_wl_nm = (299792.458 * 1000.0) / ext_wm_ghz if ext_wm_ghz > 0 else 0.0
                        power_pd_ext = 0.0
                        etalon_pd_ext = 0.0
                else:
                    # Fallback to HP wavemeter over GPIB
                    ext_wl_nm      = read_wavemeter_reg()   # wavelength in nm
                    etalon_pd_ext  = read_etalon_reg()      # proxy metric (SMSR dB)
                    power_pd_ext   = read_power_reg()       # dBm
                    ext_wm_ghz     = read_external_wavemeter()  # GHz

                # 6) Frequency comparison
                err_mhz = abs(ext_wm_ghz - expected_ghz) * 1000.0
                pass_fail = "Pass" if err_mhz <= tol_mhz else "Fail"

                result = (
                    f"CH={ch}, Expected={expected_ghz:.6f} GHz, ExtWM={ext_wm_ghz:.6f} GHz, "
                    f"Err={err_mhz:.2f} MHz, Tol={tol_mhz:.2f} MHz, "
                    f"WM_PD reg(0x89)={wm_pd_reg:.2f} | ext_wl_nm={ext_wl_nm:.3f} nm, "
                    f"ETALON_PD reg(0x8A)={etalon_pd_reg:.2f} | ext_smsr_dB={etalon_pd_ext:.2f}, "
                    f"Power_PD reg(0x8B)={power_pd_reg:.2f} | ext_power_dBm={power_pd_ext:.2f}, "
                    f"Legacy MPD(0x86)={mpd_reg:.2f}, WLPD(0x87)={wlpd_reg:.2f}, WMPD(0x88)={wmpd_reg:.2f}"
                )

        except NotImplementedError as missing:
            # Backends not wired yet: treat as informational rather than a functional failure.
            result = f"CH={ch}: BACKEND MISSING → {missing}"
            pass_fail = "Info"
        except Exception as e:
            result = f"CH={ch}: ERROR → {e}"
            pass_fail = "Fail"

        print("  " + result)

        ws.append([
            test_name,
            result,
            pass_fail,
            TesterName,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "SetChannelFull",
        ])
        # (saving deferred)
    _excel_save()

# ---------------- Main ----------------
if __name__ == "__main__":
    # Prefer Python path (only if LIB exists)
    if LIB and hasattr(LIB, "itla_use_python"):
        try:
            LIB.itla_use_python(1)  # 1=prefer Python, 0=pure C fallback
        except Exception:
            pass


    dvt_menu = {
        1: "Register Read & Write",
        2: "Table A (Supervisory Reads)",
        3: "Table 9.5 (Status / Triggers)",
        4: "9.6 Channel Sweep (Active)",
        5: "Full 0x00–0xFF Register Scan",
        6: "Set Channel + OSA readPeaks() Test",
        7: "Full SetChannel Test (Internal PD + External WM)",
        8 : "Set Channel + HP WM compare to CSV expected (Real Wavelength)",
        9 : "SetChannel Test (Internal WM/Etalon/Power + External HP WM)",
        10: "Table B (9.6 Snapshot)",
        11: "Requirements Coverage (9.5/9.6/9.7/9.8/9.9)",
    }

    for menu in dvt_menu:
        print(f" {menu} : {dvt_menu[menu]}")

    # Open UART if requested
    if IS_UART:
        if not SER_AVAILABLE:
            print("[ERROR] UART mode selected but pyserial not present.")
            sys.exit(2)
        dev = os.environ.get("ITLA_TTY")
        if not dev:
            if sys.platform.startswith("win"):
                dev = "COM3"   # adjust if your adapter is a different COM port
            else:
                dev = "/dev/ttyUSB0"
        try:
            ser = serial.Serial(dev, baudrate=115200, timeout=0.5)
            print(f"[INFO] Opened {dev} @115200")
        except Exception as e:
            print(f"[ERROR] Failed to open {dev}: {e}")
            print("       Tip: Close RealTerm/Arduino Serial Monitor or change ITLA_TTY.")
            sys.exit(3)

        # Quick wavemeter connectivity check (optional):
        #   export WAVEMETER_GPIB_ADDR=18
        #   python3 dvt_test_script.py
        # then in Python you can do:
        #   wm = HP_wavemeter(gpib_addr=_env_int("WAVEMETER_GPIB_ADDR", 18), board=_env_int("WAVEMETER_GPIB_BOARD", 0))
        #   print(wm.idn())
        # If you see Permission denied from /dev/gpib0, either run with:
        #   sudo -E ITLA_MODE=uart WAVEMETER_GPIB_ADDR=18 WAVEMETER_GPIB_BOARD=0 python3 dvt_test_script.py
        # or fix /dev/gpib0 permissions via udev rules / gpib group.

    try:
        dvt_choice = int(input("Select an option => "))
    except Exception:
        print("Invalid choice")
        sys.exit(1)

    TesterName = input("Please enter tester name => ").title()

    ran_non_gui = False

    if dvt_choice in dvt_menu:
        if dvt_choice == 1:
            interactive_reg_io()
            ran_non_gui = True
        elif dvt_choice == 2:
            run_supervisory_table()
            ran_non_gui = True
        elif dvt_choice == 3:
            run_9_5_status_table()
            ran_non_gui = True
        elif dvt_choice == 4:
            # default sweep: channels 1..10, fine_tune=0
            run_9_6_channel_sweep(start_ch=1, end_ch=10, fine_tune_mhz=0)
            ran_non_gui = True
        elif dvt_choice == 5:
            run_full_register_scan()
            ran_non_gui = True
        elif dvt_choice == 6:
            # Default: channels 1..10 and 100 MHz tolerance
            run_osa_channel_test(start_ch=1, end_ch=10, tol_mhz=100.0)
            ran_non_gui = True
        elif dvt_choice == 7:
            run_full_setchannel_test(start_ch=1, end_ch=10, tol_mhz=100.0)
            ran_non_gui = True
        elif dvt_choice == 8:
            start_ch = int(os.environ.get("WM_START_CH", "1"))
            end_ch   = int(os.environ.get("WM_END_CH",   "10"))

            tol_mhz  = float(os.environ.get("TUNING_TOL_MHZ",
                             os.environ.get("WM_TOL_MHZ", "100")))

            settle_s = float(os.environ.get("WM_SETTLE_S", "0.3"))

            run_wm_compare_to_csv(
                start_ch=start_ch,
                end_ch=end_ch,
                tol_mhz=tol_mhz,
                settle_s=settle_s
            )
            ran_non_gui = True
        elif dvt_choice == 9:
            # Option 9: Run SetChannel requirement test across ALL channels.
            # Range is controlled by env vars ITLA_CH_START / ITLA_CH_COUNT.
            #notes = ""
            #try:
                #notes = input("Notes (optional) => ").strip()
            #except Exception:
                #notes = ""

            tol = 1250.0
            try:
                s_tol = input("Tolerance GHz (default 1.25) => ").strip()
                if s_tol:
                    tol = float(s_tol) * 1000.0
            except Exception:
                tol = 1250.0

            settle = 0.3
            try:
                s_settle = input("Settle seconds after tune (default 0.3) => ").strip()
                if s_settle:
                    settle = float(s_settle)
            except Exception:
                settle = 0.3
            ch_offset = 0
            try:
                s_off = input("Channel offset (CSV_ch = DUT_ch + offset, default 0) => ").strip()
                if s_off:
                    ch_offset = int(s_off, 0)
            except Exception:
                ch_offset = 0
            os.environ["WM_CHANNEL_OFFSET"] = str(ch_offset)
            print(f"[Option 9] Using channel offset: {ch_offset} (CSV_ch = DUT_ch + {ch_offset})")
            channels = list(range(1, 102))
            run_requirement_setchannel_test(channels=channels, tol_mhz=tol, settle_s=settle)
            ran_non_gui = True
        elif dvt_choice == 10:
            run_9_6_snapshot_table()
            ran_non_gui = True
        elif dvt_choice == 11:
            run_requirements_coverage()
            ran_non_gui = True
    else:
        print("Invalid choice")
        sys.exit(1)

    # Always flush Excel to disk after completing any non-GUI option
    if ran_non_gui:
        _excel_save()
        print(f"[INFO] Excel saved: {ExcelFileName}")


    if IS_UART and ser:
        try:
            ser.close()
        except Exception:
            pass
