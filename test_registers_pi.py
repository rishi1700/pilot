#!/usr/bin/env python3
"""
test_registers_pi.py – nanoITLA UART register test for Raspberry Pi
=====================================================================
Walks every implemented ITLA register (standard + manufacturer-specific),
reads each one, optionally writes a known value and reads back, then
prints a formatted table showing pass/fail for each register.

Hardware setup
--------------
  nanoITLA UART TX  ->  Pi RX  (e.g. /dev/ttyUSB0)
  nanoITLA UART RX  ->  Pi TX
  GND               ->  GND

Usage
-----
  pip install pyserial
  sudo python3 test_registers_pi.py
  sudo python3 test_registers_pi.py --port /dev/ttyAMA0
"""

import serial
import time
import argparse
import sys

# ---------------------------------------------------------------------------
# LUT coverage — channels whose LUT row is all-zeros (no valid operating point
# found during calibration).  Per team: auto-mark these as FAIL, skip hardware.
# Source: nano_lut_50ghz.c  (101-channel 50 GHz LUT for the current device)
# ---------------------------------------------------------------------------
LUT_ZERO_CHANNELS = {
    1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,
    21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,
    41,42,43,44,45,46,47,48,49,50,51,52,
    54, 56, 57, 78, 84,
}

# ---------------------------------------------------------------------------
# ITLA frame helpers (mirrors nanoITLA.c build_inbound_frame logic)
# ---------------------------------------------------------------------------

def _bip4(b0, b1, b2, b3):
    """4-bit BIP checksum used by ITLA protocol."""
    bip8 = (b0 & 0x0F) ^ b1 ^ b2 ^ b3
    return ((bip8 >> 4) & 0x0F) ^ (bip8 & 0x0F)


def build_inbound_frame(is_write: bool, reg: int, data: int, lst_rsp: int = 0) -> int:
    """Pack a 32-bit ITLA inbound frame."""
    isw = 1 if is_write else 0
    app = (isw << 26) | ((reg & 0xFF) << 18) | (data & 0xFFFF)
    tmp = ((lst_rsp & 1) << 27) | (app & 0x07FFFFFF)
    b0 = (tmp >> 24) & 0xFF
    b1 = (tmp >> 16) & 0xFF
    b2 = (tmp >> 8) & 0xFF
    b3 = tmp & 0xFF
    csum = _bip4(b0 & 0x0F, b1, b2, b3) & 0x0F
    return tmp | (csum << 28)


def parse_outbound_frame(frame: int):
    """
    Decode a 32-bit ITLA outbound (device → host) frame.
    Returns (ce, xe, reg, data).
      ce  – checksum error from device perspective
      xe  – execution error (register not implemented / read-only violation)
      reg – echoed register number
      data – register value
    """
    csum = (frame >> 28) & 0x0F
    b0 = (frame >> 24) & 0xFF
    b1 = (frame >> 16) & 0xFF
    b2 = (frame >> 8) & 0xFF
    b3 = frame & 0xFF
    rc = _bip4(b0 & 0x0F, b1, b2, b3) & 0x0F
    our_ce = 1 if rc != csum else 0     # CE we compute locally
    ce = (frame >> 27) & 1              # CE bit set by device
    xe = (frame >> 25) & 1
    reg = (frame >> 17) & 0xFF
    data = (frame >> 1) & 0xFFFF
    return ce | our_ce, xe, reg, data


def frame_to_bytes(frame: int):
    """Convert 32-bit frame to 4-byte list (big-endian)."""
    return [(frame >> 24) & 0xFF,
            (frame >> 16) & 0xFF,
            (frame >> 8) & 0xFF,
            frame & 0xFF]


def bytes_to_frame(b):
    """Convert 4-byte list (big-endian) back to 32-bit int."""
    return (b[0] << 24) | (b[1] << 16) | (b[2] << 8) | b[3]


# ---------------------------------------------------------------------------
# UART transaction helper
# ---------------------------------------------------------------------------

def uart_transact(ser, is_write: bool, reg: int, data: int = 0, lst_rsp: int = 0):
    """
    Send one 4-byte ITLA frame over UART and return (ce, xe, reg_echo, data_back).
    """
    frame = build_inbound_frame(is_write, reg, data, lst_rsp)
    tx = frame_to_bytes(frame)
    ser.reset_input_buffer()
    ser.write(bytes(tx))
    rx = ser.read(4)
    if len(rx) < 4:
        raise TimeoutError(f"UART timeout: got {len(rx)}/4 bytes")
    rframe = bytes_to_frame(list(rx))
    return parse_outbound_frame(rframe)


def spi_read(ser, reg: int, lst_rsp: int = 0):
    return uart_transact(ser, False, reg, 0, lst_rsp)


def spi_write(ser, reg: int, data: int, lst_rsp: int = 0):
    return uart_transact(ser, True, reg, data, lst_rsp)


# ---------------------------------------------------------------------------
# Register table
# Each entry: (reg_hex, name, access, test_write_val_or_None, description)
#   access: 'RO' | 'WO' | 'RW' | 'AEA'
#   test_write_val: value to write and read-back verify (None = read-only test)
# ---------------------------------------------------------------------------

REGISTERS = [
    # reg    name              access  test_wr   description
    (0x00, "NOP",             "RO",   None,     "No-operation / poll status"),
    (0x01, "DevTyp",          "AEA",  None,     "Device type string (AEA)"),
    (0x02, "Mfgr",            "AEA",  None,     "Manufacturer string (AEA)"),
    (0x03, "Model",           "AEA",  None,     "Model string (AEA)"),
    (0x04, "SerNo",           "AEA",  None,     "Serial number string (AEA)"),
    (0x05, "MfgDate",         "AEA",  None,     "Manufacturing date string (AEA)"),
    (0x06, "Release",         "AEA",  None,     "Firmware release string (AEA)"),
    (0x07, "RelBack",         "AEA",  None,     "Firmware back-level string (AEA)"),
    (0x08, "GenCfg",          "RW",   0x0001,   "General configuration"),
    (0x09, "AEA_EAC",         "RW",   0x0003,   "AEA extension access control"),
    (0x0A, "AEA_EA",          "RW",   0x0000,   "AEA extension address"),
    (0x0D, "IOCap",           "RW",   0x0000,   "I/O capability"),
    (0x0E, "EAC",             "RW",   0x0000,   "Extended access control"),
    (0x0F, "EA",              "RO",   None,     "Extended access data"),
    (0x10, "NOOP",            "RW",   0x0000,   "No-operation (writable NOP)"),
    (0x14, "DLConfig",        "RW",   0x0000,   "Dither/lock config"),
    (0x15, "DLStatus",        "RO",   None,     "Dither/lock status"),
    (0x20, "StatusF",         "RO",   None,     "Fatal status (write-to-clear; read-only in test)"),
    (0x21, "StatusW",         "RO",   None,     "Warning status (write-to-clear; read-only in test)"),
    (0x22, "FPowTh",          "RW",   0x0064,   "Fatal power threshold"),
    (0x23, "WPowTh",          "RW",   0x0032,   "Warning power threshold"),
    (0x24, "FFreqTh",         "RW",   0x0064,   "Fatal frequency threshold"),
    (0x25, "WFreqTh",         "RW",   0x0032,   "Warning frequency threshold"),
    (0x26, "FThermTh",        "RW",   0x0064,   "Fatal thermal threshold"),
    (0x27, "WThermTh",        "RW",   0x0032,   "Warning thermal threshold"),
    (0x28, "SRQ_MASK",        "RW",   0x0000,   "SRQ mask"),
    (0x29, "FatalT",          "RW",   0x0000,   "Fatal trigger config"),
    (0x2A, "ALMT",            "RW",   0x0000,   "Alarm mask"),
    (0x30, "Channel",         "RO",   None,     "Laser channel – read-only in test (write via 0x65+0x30 sequence)"),
    (0x31, "PWR",             "RW",   0x0064,   "Laser output power (0.01 dBm)"),
    (0x32, "ResEna",          "RW",   0x0000,   "Resource enable"),
    (0x33, "MCB",             "RW",   0x0000,   "Module control byte"),
    (0x34, "Grid",            "RW",   0x0019,   "Channel grid spacing (GHz×10)"),
    (0x35, "FCF1_THz",        "RW",   0x00C1,   "First channel freq – THz part (193)"),
    (0x36, "FCF2_G10",        "RW",   0x1194,   "First channel freq – GHz×10 (450.0)"),
    (0x40, "LF1",             "RO",   None,     "Laser frequency – THz part"),
    (0x41, "LF2",             "RO",   None,     "Laser frequency – GHz×10 part"),
    (0x42, "LF1Min_THz",      "RO",   None,     "Min lasing freq – THz (Table A)"),
    (0x43, "LF1Max_THz",      "RO",   None,     "Max lasing freq – THz (Table A)"),
    (0x4F, "FTFR_MHz",        "RO",   None,     "Fine tune freq range (MHz)"),
    (0x50, "MinFreq_THz",     "RO",   None,     "Min lasing freq – THz (cap97)"),
    (0x51, "MinFreq_G10",     "RO",   None,     "Min lasing freq – GHz×10 (cap97)"),
    (0x52, "MaxFreq_THz",     "RO",   None,     "Max lasing freq – THz (cap97)"),
    (0x53, "MaxFreq_G10",     "RO",   None,     "Max lasing freq – GHz×10 (cap97)"),
    (0x54, "LastFreq_THz",    "RO",   None,     "Last channel freq – THz"),
    (0x55, "LastFreq_G10",    "RO",   None,     "Last channel freq – GHz×10 (note: uses current GRID register)"),
    (0x56, "LGrid10",         "RO",   None,     "Laser grid step (GHz×10)"),
    (0x59, "DitherCtrl",      "RO",   None,     "Dither waveform/enable control"),
    (0x62, "FTF",             "RW",   0x0000,   "Fine tune frequency (MHz, signed)"),
    (0x65, "ChannelH",        "RW",   0x0000,   "Laser channel (high word)"),
    (0x66, "ChannelL",        "RW",   0x003C,   "Laser channel – ch60 (first populated LUT row with real cal data)"),
    (0x67, "FCF3_MHz",        "RW",   0x0000,   "First channel freq – MHz part"),
    (0x68, "Grid2_MHz",       "RO",   None,     "Grid 2 – MHz offset (read-only in test)"),
    # ── Manufacturer-specific LUT / PD window (0x80–0x8B) ──────────────────
    (0x80, "V1_lut",          "RO",   None,     "Mfr: Ring-1 V – LUT setpoint (÷100 → V)"),
    (0x81, "V2_lut",          "RO",   None,     "Mfr: Ring-2 V – LUT setpoint (÷100 → V)"),
    (0x82, "V3_lut",          "RO",   None,     "Mfr: Phase V – LUT setpoint (÷100 → V)"),
    (0x83, "Gain_lut",        "RO",   None,     "Mfr: Gain bias – LUT setpoint (÷100)"),
    (0x84, "SOA_lut",         "RO",   None,     "Mfr: SOA current – LUT setpoint (÷100)"),
    (0x85, "Temp_lut",        "RO",   None,     "Mfr: TEC temp – LUT setpoint (raw)"),
    (0x86, "MPD_lut",         "RO",   None,     "Mfr: Main PD – LUT setpoint"),
    (0x87, "WLPD_lut",        "RO",   None,     "Mfr: Etalon PD – LUT setpoint"),
    (0x88, "WMPD_lut",        "RO",   None,     "Mfr: WM PD – LUT setpoint"),
    (0x89, "WMPD_adc",        "RO",   None,     "Mfr: WM PD – live ADC (÷10 = mV)"),
    (0x8A, "WLPD_adc",        "RO",   None,     "Mfr: Etalon PD – live ADC (÷10 = mV)"),
    (0x8B, "MPD_adc",         "RO",   None,     "Mfr: Main PD – live ADC (÷10 = mV)"),
    # ── Manufacturer-specific tuner R/W (0x8C–0x91) ─────────────────────────
    (0x8C, "PHASE_tuner",     "RW",   0x0064,   "Mfr: Phase tuner (÷100 → V, e.g. 100=1.00 V)"),
    (0x8D, "RING1_tuner",     "RW",   0x00C8,   "Mfr: Ring-1 tuner (÷100 → V, e.g. 200=2.00 V)"),
    (0x8E, "RING2_tuner",     "RW",   0x00C8,   "Mfr: Ring-2 tuner (÷100 → V, e.g. 200=2.00 V)"),
    (0x8F, "SOA_tuner",       "RW",   0x0064,   "Mfr: SOA current tuner (÷100 → V)"),
    (0x90, "GainBias_tuner",  "RW",   0x0064,   "Mfr: Gain bias tuner (÷100 → V)"),
    (0x91, "TEC_raw",         "RO",   None,     "Mfr: TEC current – live ADC readback (overwritten by itla_update_hw_telemetry)"),
]

# Implemented register addresses (for full-scan N/I filtering)
_IMPLEMENTED = frozenset(r for r, *_ in REGISTERS)

# All 256 addresses — unimplemented ones added as N/I stubs for --full mode
def _all_256_registers():
    full = list(REGISTERS)
    for addr in range(0x100):
        if addr not in _IMPLEMENTED:
            full.append((addr, "Reserved", "N/I", None, "Not implemented / reserved"))
    full.sort(key=lambda x: x[0])
    return full

# ---------------------------------------------------------------------------
# Test runner
# ---------------------------------------------------------------------------

PASS = "PASS"
FAIL = "FAIL"
SKIP = "SKIP"
WARN = "WARN"
NI   = "N/I"

COL_WIDTHS = (4, 16, 6, 7, 7, 7, 40)
HEADER = ("Reg", "Name", "Access", "Read", "Write", "Verify", "Notes")


def fmt_row(*fields):
    return "  ".join(str(f).ljust(w) for f, w in zip(fields, COL_WIDTHS))


def run_tests(spi, verbose=False, full=False):
    reg_list = _all_256_registers() if full else REGISTERS
    print()
    print("=" * 95)
    title = "nano-ITLA Full 256-Register Scan" if full else "nano-ITLA Register Test"
    print(f"  {title}  —  Pilot Photonics nanoITLA-01")
    print("=" * 95)
    print(fmt_row(*HEADER))
    print("-" * 95)

    results = []
    passed = failed = skipped = warned = ni_count = 0

    for reg, name, access, test_val, desc in reg_list:
        read_result = "—"
        write_result = "—"
        verify_result = "—"
        notes = ""

        # ── N/I (reserved/unimplemented) ─────────────────────────────────────
        if access == "N/I":
            row = fmt_row(f"0x{reg:02X}", name, access, "—", "—", NI, desc[:COL_WIDTHS[-1]])
            print(row)
            results.append((reg, name, access, "—", "—", "—", desc[:COL_WIDTHS[-1]], NI))
            ni_count += 1
            continue

        # ── READ ────────────────────────────────────────────────────────────
        try:
            ce, xe, reg_echo, val = spi_read(spi, reg)
            if ce:
                read_result = "CE-ERR"
                notes = "checksum error on read"
            elif xe:
                read_result = f"XE({val:#06x})"
                notes = "register returned execution error"
            else:
                read_result = f"{val:#06x}"
        except Exception as ex:
            read_result = "EXC"
            notes = str(ex)

        # ── WRITE + READ-BACK ────────────────────────────────────────────────
        if test_val is not None and access in ("RW",):
            try:
                wce, wxe, _, _ = spi_write(spi, reg, test_val)
                time.sleep(0.002)   # 2 ms settle

                if wce:
                    write_result = "CE-ERR"
                    notes = "checksum error on write"
                elif wxe:
                    write_result = f"XE"
                    notes = "write returned execution error"
                else:
                    write_result = f"{test_val:#06x}"

                    # read back
                    rce, rxe, _, rval = spi_read(spi, reg)
                    if rce:
                        verify_result = "CE-ERR"
                    elif rxe:
                        verify_result = "XE"
                    elif rval == test_val:
                        verify_result = PASS
                    else:
                        verify_result = f"MISMATCH({rval:#06x})"
                        notes = f"expected {test_val:#06x} got {rval:#06x}"

            except Exception as ex:
                write_result = "EXC"
                notes = str(ex)

        # ── Score ───────────────────────────────────────────────────────────
        if access == "RO" or access == "AEA":
            # Just reading — pass if no CE/XE or AEA response (status=2)
            if "CE-ERR" in read_result or "EXC" in read_result:
                overall = FAIL
                failed += 1
            else:
                overall = PASS
                passed += 1
        else:
            if verify_result == PASS:
                overall = PASS
                passed += 1
            elif verify_result == "—":
                # write wasn't attempted (AEA, WO, etc.)
                overall = SKIP
                skipped += 1
            elif "MISMATCH" in verify_result:
                overall = FAIL
                failed += 1
            elif "EXC" in (write_result, verify_result) or "CE-ERR" in (write_result, verify_result):
                overall = FAIL
                failed += 1
            else:
                overall = WARN
                warned += 1

        # Truncate notes to fit column
        notes_col = (desc if not notes else notes)[:COL_WIDTHS[-1]]

        row = fmt_row(f"0x{reg:02X}", name, access, read_result, write_result, verify_result, notes_col)
        print(row)
        results.append((reg, name, access, read_result, write_result, verify_result, notes_col, overall))

        if verbose and notes:
            print(f"      note: {notes}")

    print("-" * 95)
    tested = len(results) - ni_count
    ni_str = f"   N/I: {ni_count}" if ni_count else ""
    print(f"  TOTAL: {len(results)}   TESTED: {tested}   PASS: {passed}   FAIL: {failed}   WARN: {warned}   SKIP: {skipped}{ni_str}")
    print("=" * 95)
    print()
    return results, failed == 0


# ---------------------------------------------------------------------------
# Channel sweep test
# ---------------------------------------------------------------------------

def run_channel_sweep(ser, verbose=False):
    """
    Walk all 101 LUT channels (50 GHz grid).

    Zero-LUT channels: auto-FAIL without touching hardware — no valid operating
    point was found during calibration (per team instruction).

    Non-zero channels: write channel to 0x65/0x66, read back LUT-setpoint
    registers 0x80-0x84 and verify at least one drive is non-zero.

    Returns (results, all_pass) where results is a list of 8-tuples:
      (ch, lut_ok, v1, v2, v3, gain, soa, overall, notes)
    """
    NUM_CHANNELS = 101
    CH_COL_WIDTHS = (4, 8, 8, 8, 8, 8, 8, 40)
    CH_HEADER     = ("Ch", "V1_lut", "V2_lut", "V3_lut", "Gain_lut", "SOA_lut", "Result", "Notes")

    print()
    print("=" * 95)
    print("  nano-ITLA Channel Sweep Test  —  101 channels × 50 GHz")
    print("=" * 95)
    print("  \u25ba  Channels with all-zero LUT rows are auto-marked FAIL (no valid operating point)")
    print("-" * 95)
    print("  " + "  ".join(str(h).ljust(w) for h, w in zip(CH_HEADER, CH_COL_WIDTHS)))
    print("-" * 95)

    results = []
    n_pass = n_fail = n_zero = 0

    for ch in range(1, NUM_CHANNELS + 1):
        v1_s = v2_s = v3_s = gain_s = soa_s = "—"
        notes = ""

        if ch in LUT_ZERO_CHANNELS:
            # Auto-fail: zero LUT row
            overall = FAIL
            notes   = "LUT row all-zeros: no valid operating point"
            n_fail += 1
            n_zero += 1
        else:
            try:
                # Write channel number into ChannelH (0x65=0) and ChannelL (0x66=ch)
                spi_write(ser, 0x65, 0)
                spi_write(ser, 0x66, ch)
                time.sleep(0.02)   # 20 ms — let firmware apply LUT in main loop

                # Read LUT setpoint mirrors
                def _rd(reg):
                    ce, xe, _, val = spi_read(ser, reg)
                    if ce: return "CE-ERR"
                    if xe: return "XE"
                    return val

                v1   = _rd(0x80)
                v2   = _rd(0x81)
                v3   = _rd(0x82)
                gain = _rd(0x83)
                soa  = _rd(0x84)

                v1_s   = f"0x{v1:04X}"   if isinstance(v1,   int) else v1
                v2_s   = f"0x{v2:04X}"   if isinstance(v2,   int) else v2
                v3_s   = f"0x{v3:04X}"   if isinstance(v3,   int) else v3
                gain_s = f"0x{gain:04X}" if isinstance(gain, int) else gain
                soa_s  = f"0x{soa:04X}"  if isinstance(soa,  int) else soa

                # PASS if at least one of v1/v2/v3/gain/soa is non-zero
                numeric = [x for x in (v1, v2, v3, gain, soa) if isinstance(x, int)]
                if any(x != 0 for x in numeric):
                    overall = PASS
                    n_pass += 1
                else:
                    overall = FAIL
                    notes   = "all drive registers read zero after channel select"
                    n_fail += 1

            except Exception as ex:
                overall = FAIL
                notes   = str(ex)
                n_fail += 1

        row_fields = (str(ch), v1_s, v2_s, v3_s, gain_s, soa_s, overall, notes[:CH_COL_WIDTHS[-1]])
        print("  " + "  ".join(str(f).ljust(w) for f, w in zip(row_fields, CH_COL_WIDTHS)))
        results.append((ch, v1_s, v2_s, v3_s, gain_s, soa_s, overall, notes))

        if verbose and notes:
            print(f"      note: {notes}")

    print("-" * 95)
    print(f"  TOTAL: {NUM_CHANNELS}   PASS: {n_pass}   FAIL: {n_fail}  "
          f"  (of which {n_zero} auto-fail: zero LUT rows)")
    print("=" * 95)
    print()
    return results, n_fail == 0


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def save_excel(results, port, out_path, channel_results=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed — skipping Excel export  (pip install openpyxl)")
        return

    import datetime

    def _fill(hex6):
        return PatternFill("solid", fgColor=hex6)

    def _font(bold=False, color="000000", sz=10):
        return Font(name="Calibri", size=sz, bold=bold, color=color)

    def _bd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    CTR = Alignment(horizontal="center", vertical="center")
    LFT = Alignment(horizontal="left",   vertical="center")

    C_PASS  = "C6EFCE"; C_PASS_F  = "276221"
    C_FAIL  = "FFC7CE"; C_FAIL_F  = "9C0006"
    C_WARN  = "FFEB9C"; C_WARN_F  = "9C6500"
    C_NI    = "F2F2F2"; C_NI_F    = "AAAAAA"
    C_SKIP  = "DDEBF7"; C_SKIP_F  = "2E4057"
    C_HDR   = "1F3864"; C_HDR_F   = "FFFFFF"
    C_ROW1  = "FFFFFF"; C_ROW2    = "F7F9FC"

    STATUS_STYLE = {
        "PASS": (_fill(C_PASS), _font(bold=True, color=C_PASS_F)),
        "FAIL": (_fill(C_FAIL), _font(bold=True, color=C_FAIL_F)),
        "WARN": (_fill(C_WARN), _font(bold=True, color=C_WARN_F)),
        "N/I":  (_fill(C_NI),   Font(name="Calibri", size=9, italic=True, color=C_NI_F)),
        "SKIP": (_fill(C_SKIP), _font(color=C_SKIP_F)),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Register Scan"

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "nano-ITLA  Register Test Report  —  Pilot Photonics"
    t.font  = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.fill  = _fill(C_HDR)
    t.alignment = CTR
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = (f"Port: {port}    "
                 f"Date: {datetime.datetime.now().strftime('%Y-%m-%d  %H:%M')}    "
                 f"Registers: {len(results)}    "
                 f"PASS: {sum(1 for *_,s in results if s=='PASS')}    "
                 f"FAIL: {sum(1 for *_,s in results if s=='FAIL')}    "
                 f"N/I: {sum(1 for *_,s in results if s=='N/I')}")
    sub.font  = Font(name="Calibri", size=9, italic=True, color="FFFFFF")
    sub.fill  = _fill("2E4057")
    sub.alignment = CTR
    ws.row_dimensions[2].height = 14

    # ── Column headers ───────────────────────────────────────────────────────
    headers = ["Addr", "Dec", "Name", "Access", "Read", "Write", "Verify", "Notes / Description"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font      = _font(bold=True, color=C_HDR_F, sz=10)
        cell.fill      = _fill(C_HDR)
        cell.alignment = CTR
        cell.border    = _bd()
    ws.row_dimensions[3].height = 16

    # ── Data rows ────────────────────────────────────────────────────────────
    for r_idx, (reg, name, access, read_v, write_v, verify_v, notes, overall) in enumerate(results, 4):
        bg = _fill(C_ROW1 if r_idx % 2 == 0 else C_ROW2)
        vals = [f"0x{reg:02X}", str(reg), name, access, read_v, write_v, verify_v, notes]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = _bd()
            cell.alignment = LFT if c == 8 else CTR
            if c == 7 and overall in STATUS_STYLE:          # Verify col gets status colour
                fill_, font_ = STATUS_STYLE[overall]
                cell.fill = fill_; cell.font = font_
            elif overall == "N/I":
                cell.fill = _fill(C_NI)
                cell.font = Font(name="Calibri", size=9, italic=True, color=C_NI_F)
            else:
                cell.fill = bg
                cell.font = _font(sz=9)
        ws.row_dimensions[r_idx].height = 13

    # ── Column widths ────────────────────────────────────────────────────────
    for col, w in zip("ABCDEFGH", [7, 5, 18, 8, 10, 10, 8, 46]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # ── Channel Sweep sheet (optional) ───────────────────────────────────────
    if channel_results:
        cs = wb.create_sheet("Channel Sweep")

        cs.merge_cells("A1:I1")
        t2 = cs["A1"]
        t2.value = "nano-ITLA  Channel Sweep  —  101 channels × 50 GHz"
        t2.font  = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        t2.fill  = _fill(C_HDR); t2.alignment = CTR
        cs.row_dimensions[1].height = 22

        cs.merge_cells("A2:I2")
        sub2 = cs["A2"]
        n_pass_ch = sum(1 for *_, s, _ in channel_results if s == PASS)
        n_fail_ch = sum(1 for *_, s, _ in channel_results if s == FAIL)
        n_zero_ch = sum(1 for *_, s, note in channel_results if s == FAIL and "all-zeros" in note)
        sub2.value = (f"Date: {__import__('datetime').datetime.now().strftime('%Y-%m-%d  %H:%M')}    "
                      f"Total: {len(channel_results)}    PASS: {n_pass_ch}    FAIL: {n_fail_ch}"
                      f"  (of which {n_zero_ch} auto-fail: zero LUT rows)")
        sub2.font  = Font(name="Calibri", size=9, italic=True, color="FFFFFF")
        sub2.fill  = _fill("2E4057"); sub2.alignment = CTR
        cs.row_dimensions[2].height = 14

        ch_headers = ["Ch", "LUT", "V1_lut", "V2_lut", "V3_lut", "Gain_lut", "SOA_lut", "Result", "Notes"]
        for c, h in enumerate(ch_headers, 1):
            cell = cs.cell(row=3, column=c, value=h)
            cell.font = _font(bold=True, color=C_HDR_F); cell.fill = _fill(C_HDR)
            cell.alignment = CTR; cell.border = _bd()
        cs.row_dimensions[3].height = 16

        for r_idx, (ch, v1, v2, v3, gain, soa, overall, notes) in enumerate(channel_results, 4):
            lut_tag  = "ZERO" if ch in LUT_ZERO_CHANNELS else "OK"
            bg = _fill(C_ROW1 if r_idx % 2 == 0 else C_ROW2)
            row_vals = [str(ch), lut_tag, v1, v2, v3, gain, soa, overall, notes]
            for c, v in enumerate(row_vals, 1):
                cell = cs.cell(row=r_idx, column=c, value=v)
                cell.border = _bd()
                cell.alignment = LFT if c == 9 else CTR
                if c == 8 and overall in STATUS_STYLE:
                    fill_, font_ = STATUS_STYLE[overall]
                    cell.fill = fill_; cell.font = font_
                elif c == 2 and lut_tag == "ZERO":
                    cell.fill = _fill(C_NI)
                    cell.font = Font(name="Calibri", size=9, italic=True, color=C_NI_F)
                else:
                    cell.fill = bg; cell.font = _font(sz=9)
            cs.row_dimensions[r_idx].height = 13

        for col, w in zip("ABCDEFGHI", [5, 6, 10, 10, 10, 10, 10, 8, 46]):
            cs.column_dimensions[col].width = w
        cs.freeze_panes = "A4"

    wb.save(out_path)
    print(f"Excel saved: {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="nanoITLA UART register test (Raspberry Pi)")
    parser.add_argument("--port",    default="/dev/ttyUSB0", help="Serial port (default /dev/ttyUSB0)")
    parser.add_argument("--baud",    type=int, default=115200, help="Baud rate (default 115200)")
    parser.add_argument("--timeout", type=float, default=0.5,  help="Read timeout seconds (default 0.5)")
    parser.add_argument("-v", "--verbose", action="store_true", help="Print extra detail on failures")
    parser.add_argument("--full",     action="store_true", help="Scan all 256 registers (marks unimplemented as N/I)")
    parser.add_argument("--channels", action="store_true", help="Run 101-channel LUT sweep (auto-fails zero-LUT rows)")
    parser.add_argument("--excel",    metavar="FILE",      help="Save results to Excel .xlsx (e.g. --excel report.xlsx)")
    args = parser.parse_args()

    import time
    ser = serial.Serial(args.port, args.baud, timeout=args.timeout)
    time.sleep(0.1)
    ser.reset_input_buffer()

    print(f"UART {args.port}  baud={args.baud}  timeout={args.timeout}s")

    channel_results = None
    try:
        results, ok = run_tests(ser, verbose=args.verbose, full=args.full)
        if args.channels:
            channel_results, ch_ok = run_channel_sweep(ser, verbose=args.verbose)
            ok = ok and ch_ok
    finally:
        ser.close()

    if args.excel:
        save_excel(results, args.port, args.excel, channel_results=channel_results)

    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
