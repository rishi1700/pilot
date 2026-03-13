#!/usr/bin/env python3
"""
nano-ITLA Register Scanner
Generates a full register scan report as an .xlsx file.

Wire protocol (4-byte big-endian frames, 115200 baud):

  Inbound frame (host → device):
    bits[31:28]  BIP4 checksum (4 bits)
    bit [27]     lstRsp
    bit [26]     R/W  (1 = write)
    bits[25:18]  register address (8 bits)
    bits[17:16]  reserved (0)
    bits[15:0]   data (16 bits)

  Outbound frame (device → host):
    bits[31:28]  BIP4 checksum (4 bits)
    bit [27]     CE  (checksum error)
    bit [26]     1   (fixed)
    bit [25]     XE  (execution error)
    bits[24:17]  register address (8 bits)
    bits[16:1]   data (16 bits, shifted left by 1 in wire encoding)
    bit [0]      0   (padding)

Usage:
    python3 register_scan.py [--port /dev/ttyUSB0] [--out register_scan.xlsx]
"""

import argparse
import datetime
import struct
import time
import serial
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Register map
# Each entry: (hex_addr, name, access, description)
#   access: 'RO'  – read only, no write attempted
#           'RW'  – read then write-back and verify
#           'AEA' – async extended access (string via EA registers)
#           'N/I' – not implemented (probe still sent, XE expected)
# ---------------------------------------------------------------------------
REGISTER_MAP = [
    # ── Identity / general ──────────────────────────────────────────────────
    (0x00, 'NOP',           'RO',  'No-operation / poll status'),
    (0x01, 'DevTyp',        'AEA', 'Device type string (AEA)'),
    (0x02, 'Mfgr',          'AEA', 'Manufacturer string (AEA)'),
    (0x03, 'Model',         'AEA', 'Model string (AEA)'),
    (0x04, 'SerNo',         'AEA', 'Serial number string (AEA)'),
    (0x05, 'MfgDate',       'AEA', 'Manufacturing date string (AEA)'),
    (0x06, 'Release',       'AEA', 'Firmware release string (AEA)'),
    (0x07, 'RelBack',       'AEA', 'Firmware back-level string (AEA)'),
    (0x08, 'GenCfg',        'RW',  'General configuration'),
    (0x09, 'AEA_EAC',       'RW',  'AEA extension access control'),
    (0x0A, 'AEA_EA',        'RW',  'AEA extension address'),
    (0x0B, 'Reserved',      'N/I', 'Reserved'),
    (0x0C, 'Reserved',      'N/I', 'Reserved'),
    (0x0D, 'IOCap',         'RW',  'I/O capability'),
    (0x0E, 'EAC',           'RW',  'Extended access control'),
    (0x0F, 'EA',            'RO',  'Extended access data'),
    (0x10, 'NOOP',          'RW',  'Write returns execution error by design'),
    (0x11, 'Reserved',      'N/I', 'Reserved'),
    (0x12, 'Reserved',      'N/I', 'Reserved'),
    (0x13, 'Reserved',      'N/I', 'LstResp (deprecated, not implemented)'),
    # ── §9.4  Download control / status ─────────────────────────────────────
    (0x14, 'DLConfig',      'RW',  'Download configuration'),
    (0x15, 'DLStatus',      'RO',  'Download status'),
    (0x16, 'Reserved',      'N/I', 'Reserved'),
    (0x17, 'Reserved',      'N/I', 'Reserved'),
    (0x18, 'Reserved',      'N/I', 'Reserved'),
    (0x19, 'Reserved',      'N/I', 'Reserved'),
    (0x1A, 'Reserved',      'N/I', 'Reserved'),
    (0x1B, 'Reserved',      'N/I', 'Reserved'),
    (0x1C, 'Reserved',      'N/I', 'Reserved'),
    (0x1D, 'Reserved',      'N/I', 'Reserved'),
    (0x1E, 'Reserved',      'N/I', 'Reserved'),
    (0x1F, 'Reserved',      'N/I', 'Reserved'),
    # ── §9.5  Status & alarms ────────────────────────────────────────────────
    (0x20, 'StatusF',       'RO',  'Fatal status (write-to-clear; read-only in scan)'),
    (0x21, 'StatusW',       'RO',  'Warning status (write-to-clear; read-only in scan)'),
    (0x22, 'FPowTh',        'RW',  'Fatal power threshold'),
    (0x23, 'WPowTh',        'RW',  'Warning power threshold'),
    (0x24, 'FFreqTh',       'RW',  'Fatal frequency threshold'),
    (0x25, 'WFreqTh',       'RW',  'Warning frequency threshold'),
    (0x26, 'FThermTh',      'RW',  'Fatal thermal threshold'),
    (0x27, 'WThermTh',      'RW',  'Warning thermal threshold'),
    (0x28, 'SRQ_MASK',      'RW',  'SRQ mask'),
    (0x29, 'FatalT',        'RW',  'Fatal trigger config'),
    (0x2A, 'ALMT',          'RW',  'Alarm mask'),
    (0x2B, 'Reserved',      'N/I', 'Reserved'),
    (0x2C, 'Reserved',      'N/I', 'Reserved'),
    (0x2D, 'Reserved',      'N/I', 'Reserved'),
    (0x2E, 'Reserved',      'N/I', 'Reserved'),
    (0x2F, 'Reserved',      'N/I', 'Reserved'),
    # ── §9.6  Channel / grid / power ─────────────────────────────────────────
    (0x30, 'Channel',       'RO',  'Laser channel – read-only in scan (write triggers retune)'),
    (0x31, 'PWR',           'RW',  'Laser output power (0.01 dBm)'),
    (0x32, 'ResEna',        'RW',  'Resource enable'),
    (0x33, 'MCB',           'RW',  'Module control byte'),
    (0x34, 'Grid',          'RW',  'Channel grid spacing (GHz×10)'),
    (0x35, 'FCF1_THz',      'RW',  'First channel freq – THz part (193)'),
    (0x36, 'FCF2_G10',      'RW',  'First channel freq – GHz×10 (450.0)'),
    (0x37, 'Reserved',      'N/I', 'Reserved'),
    (0x38, 'Reserved',      'N/I', 'Reserved'),
    (0x39, 'Reserved',      'N/I', 'Reserved'),
    (0x3A, 'Reserved',      'N/I', 'Reserved'),
    (0x3B, 'Reserved',      'N/I', 'Reserved'),
    (0x3C, 'Reserved',      'N/I', 'Reserved'),
    (0x3D, 'Reserved',      'N/I', 'Reserved'),
    (0x3E, 'Reserved',      'N/I', 'Reserved'),
    (0x3F, 'Reserved',      'N/I', 'Reserved'),
    (0x40, 'LF1',           'RO',  'Laser frequency – THz part'),
    (0x41, 'LF2',           'RO',  'Laser frequency – GHz×10 part'),
    (0x42, 'LF1Min_THz',    'RO',  'Min lasing freq – THz (Table A)'),
    (0x43, 'LF1Max_THz',    'RO',  'Max lasing freq – THz (Table A)'),
    (0x44, 'Reserved',      'N/I', 'Reserved'),
    (0x45, 'Reserved',      'N/I', 'Reserved'),
    (0x46, 'Reserved',      'N/I', 'Reserved'),
    (0x47, 'Reserved',      'N/I', 'Reserved'),
    (0x48, 'Reserved',      'N/I', 'Reserved'),
    (0x49, 'Reserved',      'N/I', 'Reserved'),
    (0x4A, 'Reserved',      'N/I', 'Reserved'),
    (0x4B, 'Reserved',      'N/I', 'Reserved'),
    (0x4C, 'Reserved',      'N/I', 'Reserved'),
    (0x4D, 'Reserved',      'N/I', 'Reserved'),
    (0x4E, 'Reserved',      'N/I', 'Reserved'),
    # ── §9.7  Module capabilities ────────────────────────────────────────────
    (0x4F, 'FTFR_MHz',      'RO',  'Fine tune freq range (MHz)'),
    (0x50, 'MinFreq_THz',   'RO',  'Min lasing freq – THz (cap97)'),
    (0x51, 'MinFreq_G10',   'RO',  'Min lasing freq – GHz×10 (cap97)'),
    (0x52, 'MaxFreq_THz',   'RO',  'Max lasing freq – THz (cap97)'),
    (0x53, 'MaxFreq_G10',   'RO',  'Max lasing freq – GHz×10 (cap97)'),
    (0x54, 'LastFreq_THz',  'RO',  'Last channel freq – THz'),
    (0x55, 'LastFreq_G10',  'RO',  'Last channel freq – GHz×10'),
    (0x56, 'LGrid10',       'RO',  'Laser grid step (GHz×10)'),
    # ── §9.8  Command block ──────────────────────────────────────────────────
    (0x57, 'Currents',      'AEA', '9.8.1: Module currents array (mA×10) via AEA'),
    (0x58, 'Temps',         'AEA', '9.8.2: Diode and case temperatures (°C×100) via AEA'),
    (0x59, 'DitherCtrl',    'RW',  '9.8.3: Dither ctrl – bits[5:4]=waveform, bit[1]=enable'),
    (0x5A, 'DitherRate',    'RW',  '9.8.3: Dither rate (valid range 10–200, default 100)'),
    (0x5B, 'DitherFreq',    'RW',  '9.8.3: Dither frequency'),
    (0x5C, 'DitherAmp',     'RW',  '9.8.3: Dither amplitude (0–1000)'),
    (0x5D, 'TBTFL',         'RW',  '9.8.4: Thermal boundary low limit (°C×100, signed, default -500)'),
    (0x5E, 'TBTFH',         'RW',  '9.8.4: Thermal boundary high limit (°C×100, signed, default 7000)'),
    (0x5F, 'FAgeTh',        'RW',  '9.8.6: Fatal age threshold (%EOL, 0–100; 0=disabled)'),
    (0x60, 'WAgeTh',        'RW',  '9.8.6: Warning age threshold (%EOL, 0–100; 0=disabled)'),
    (0x61, 'Age',           'RO',  '9.8.6: Module age (%EOL, read-only)'),
    # ── §9.6 ext  Fine-tune & channel extensions ─────────────────────────────
    (0x62, 'FTF',           'RW',  'Fine tune frequency (MHz, signed, ±12500)'),
    # ── §9.5 ext  Frequency threshold extensions ─────────────────────────────
    (0x63, 'FFreqTh2',      'RW',  '9.5 ext: Fatal frequency threshold 2'),
    (0x64, 'WFreqTh2',      'RW',  '9.5 ext: Warning frequency threshold 2'),
    # ── §9.6 ext  Channel / FCF extensions ───────────────────────────────────
    (0x65, 'ChannelH',      'RW',  'Laser channel high word'),
    (0x66, 'ChannelL',      'RW',  'Laser channel low word (current channel index)'),
    (0x67, 'FCF3_MHz',      'RW',  'First channel freq – MHz part'),
    (0x68, 'Grid2_MHz',     'RO',  'Grid 2 – fine grid part (MHz, read-only in scan)'),
    (0x69, 'Reserved',      'N/I', 'Reserved'),
    (0x6A, 'Reserved',      'N/I', 'Reserved'),
    (0x6B, 'Reserved',      'N/I', 'Reserved'),
    (0x6C, 'Reserved',      'N/I', 'Reserved'),
    (0x6D, 'Reserved',      'N/I', 'Reserved'),
    (0x6E, 'Reserved',      'N/I', 'Reserved'),
    (0x6F, 'Reserved',      'N/I', 'Reserved'),
    (0x70, 'Reserved',      'N/I', 'Reserved'),
    (0x71, 'Reserved',      'N/I', 'Reserved'),
    (0x72, 'Reserved',      'N/I', 'Reserved'),
    (0x73, 'Reserved',      'N/I', 'Reserved'),
    (0x74, 'Reserved',      'N/I', 'Reserved'),
    (0x75, 'Reserved',      'N/I', 'Reserved'),
    (0x76, 'Reserved',      'N/I', 'Reserved'),
    (0x77, 'Reserved',      'N/I', 'Reserved'),
    (0x78, 'Reserved',      'N/I', 'Reserved'),
    (0x79, 'Reserved',      'N/I', 'Reserved'),
    (0x7A, 'Reserved',      'N/I', 'Reserved'),
    (0x7B, 'Reserved',      'N/I', 'Reserved'),
    (0x7C, 'Reserved',      'N/I', 'Reserved'),
    (0x7D, 'Reserved',      'N/I', 'Reserved'),
    (0x7E, 'Reserved',      'N/I', 'Reserved'),
    (0x7F, 'Reserved',      'N/I', 'Reserved'),
    # ── Manufacturer debug window ────────────────────────────────────────────
    (0x80, 'V1_lut',        'RO',  'Mfr: Ring-1 V – LUT setpoint (÷100 → V)'),
    (0x81, 'V2_lut',        'RO',  'Mfr: Ring-2 V – LUT setpoint (÷100 → V)'),
    (0x82, 'V3_lut',        'RO',  'Mfr: Phase V – LUT setpoint (÷100 → V)'),
    (0x83, 'Gain_lut',      'RO',  'Mfr: Gain bias – LUT setpoint (÷100)'),
    (0x84, 'SOA_lut',       'RO',  'Mfr: SOA current – LUT setpoint (÷100)'),
    (0x85, 'Temp_lut',      'RO',  'Mfr: TEC temp – LUT setpoint (raw)'),
    (0x86, 'MPD_lut',       'RO',  'Mfr: Main PD – LUT setpoint'),
    (0x87, 'WLPD_lut',      'RO',  'Mfr: Etalon PD – LUT setpoint'),
    (0x88, 'WMPD_lut',      'RO',  'Mfr: WM PD – LUT setpoint'),
    (0x89, 'WMPD_adc',      'RO',  'Mfr: WM PD – live ADC (÷10 = mV)'),
    (0x8A, 'WLPD_adc',      'RO',  'Mfr: Etalon PD – live ADC (÷10 = mV)'),
    (0x8B, 'MPD_adc',       'RO',  'Mfr: Main PD – live ADC (÷10 = mV)'),
    (0x8C, 'PHASE_tuner',   'RW',  'Mfr: Phase tuner (÷100 → V, e.g. 100=1.00 V)'),
    (0x8D, 'RING1_tuner',   'RW',  'Mfr: Ring-1 tuner (÷100 → V, e.g. 200=2.00 V)'),
    (0x8E, 'RING2_tuner',   'RW',  'Mfr: Ring-2 tuner (÷100 → V, e.g. 200=2.00 V)'),
    (0x8F, 'SOA_tuner',     'RW',  'Mfr: SOA current tuner (÷100 → V)'),
    (0x90, 'GainBias_tuner','RW',  'Mfr: Gain bias tuner (÷100 → V)'),
    (0x91, 'TEC_raw',       'RO',  'Mfr: TEC current – live ADC readback (overloaded write: mode switch)'),
]

# Pad remaining addresses 0x92–0xFF as Reserved N/I
_defined = {addr for addr, *_ in REGISTER_MAP}
for _addr in range(0x92, 0x100):
    REGISTER_MAP.append((_addr, 'Reserved', 'N/I', 'Reserved'))

assert len(REGISTER_MAP) == 256, f'Expected 256 entries, got {len(REGISTER_MAP)}'
assert len({r[0] for r in REGISTER_MAP}) == 256, 'Duplicate address in register map'


# ---------------------------------------------------------------------------
# Frame encoding / decoding
# ---------------------------------------------------------------------------

def _bip4(b0, b1, b2, b3):
    bip8 = (b0 & 0x0F) ^ b1 ^ b2 ^ b3
    return ((bip8 >> 4) & 0x0F) ^ (bip8 & 0x0F)


def build_frame(is_write, reg, data, lstrsp=0):
    """Encode a 4-byte inbound command frame."""
    app = ((is_write & 1) << 26) | ((reg & 0xFF) << 18) | (data & 0xFFFF)
    tmp = ((lstrsp & 1) << 27) | (app & 0x07FFFFFF)
    b0 = (tmp >> 24) & 0xFF
    b1 = (tmp >> 16) & 0xFF
    b2 = (tmp >> 8) & 0xFF
    b3 = tmp & 0xFF
    csum = _bip4(b0 & 0x0F, b1, b2, b3) & 0x0F
    frame = tmp | (csum << 28)
    return struct.pack('>I', frame)


def parse_response(raw4):
    """Decode a 4-byte outbound response frame.

    Returns (ce, xe, reg_echo, data).
    """
    frame = struct.unpack('>I', raw4)[0]
    csum = (frame >> 28) & 0x0F
    b0 = (frame >> 24) & 0xFF
    b1 = (frame >> 16) & 0xFF
    b2 = (frame >> 8) & 0xFF
    b3 = frame & 0xFF
    calc = _bip4(b0 & 0x0F, b1, b2, b3) & 0x0F
    ce = 1 if calc != csum else 0
    xe = (frame >> 25) & 1
    reg = (frame >> 17) & 0xFF
    data = (frame >> 1) & 0xFFFF
    return ce, xe, reg, data


# ---------------------------------------------------------------------------
# Scanner
# ---------------------------------------------------------------------------

def _transact(ser, is_write, reg, data):
    ser.write(build_frame(is_write, reg, data))
    raw = ser.read(4)
    if len(raw) != 4:
        raise IOError(f'Short read on reg 0x{reg:02X}: got {len(raw)} bytes')
    return parse_response(raw)


def read_reg(ser, reg):
    return _transact(ser, 0, reg, 0)


def write_reg(ser, reg, data):
    return _transact(ser, 1, reg, data)


def fetch_aea_string(ser, reg):
    """Read an AEA string register. Returns decoded string or None on error."""
    ce, xe, _, length = read_reg(ser, reg)
    if ce or xe or length == 0:
        return None
    # Enable auto-increment on read
    write_reg(ser, 0x0E, 0x0001)
    write_reg(ser, 0x0A, 0x0000)
    chars = []
    for _ in range(length):
        ce2, xe2, _, word = read_reg(ser, 0x0F)
        if ce2 or xe2:
            break
        hi = (word >> 8) & 0xFF
        lo = word & 0xFF
        if hi:
            chars.append(chr(hi))
        if lo:
            chars.append(chr(lo))
    write_reg(ser, 0x0E, 0x0000)
    return ''.join(chars).strip('\x00').strip()


def scan(port, baud=115200, timeout=1.0):
    results = []
    with serial.Serial(port, baud, timeout=timeout) as ser:
        time.sleep(0.1)
        ser.reset_input_buffer()

        for addr, name, access, desc in REGISTER_MAP:
            row = {
                'addr': addr,
                'name': name,
                'access': access,
                'desc': desc,
                'read_val': None,
                'write_val': None,
                'verify': '—',
                'aea_str': None,
                'xe_on_read': False,
                'xe_on_write': False,
                'ce': False,
            }

            if access == 'N/I':
                ce, xe, _, val = read_reg(ser, addr)
                row['ce'] = bool(ce)
                row['xe_on_read'] = bool(xe)
                # N/I registers should return XE; record value if unexpectedly responds
                if not xe:
                    row['read_val'] = val

            elif access == 'RO':
                ce, xe, _, val = read_reg(ser, addr)
                row['ce'] = bool(ce)
                row['xe_on_read'] = bool(xe)
                if not xe:
                    row['read_val'] = val

            elif access == 'AEA':
                ce, xe, _, length = read_reg(ser, addr)
                row['ce'] = bool(ce)
                row['xe_on_read'] = bool(xe)
                row['read_val'] = length if not xe else None
                if not xe:
                    row['aea_str'] = fetch_aea_string(ser, addr)

            elif access == 'RW':
                ce, xe, _, val = read_reg(ser, addr)
                row['ce'] = bool(ce)
                row['xe_on_read'] = bool(xe)
                if xe:
                    # Unexpectedly not implemented
                    row['verify'] = 'N/I'
                else:
                    row['read_val'] = val
                    wce, wxe, _, _ = write_reg(ser, addr, val)
                    row['xe_on_write'] = bool(wxe)
                    row['write_val'] = val

                    if wxe:
                        # Write rejected in current state (e.g. dither enabled, or by design)
                        row['write_val'] = 'XE'
                        row['verify'] = '—'
                    else:
                        vce, vxe, _, vval = read_reg(ser, addr)
                        if vxe:
                            row['verify'] = '—'
                        elif vval == val:
                            row['verify'] = 'PASS'
                        else:
                            row['verify'] = 'FAIL'

            results.append(row)
            time.sleep(0.005)

    return results


# ---------------------------------------------------------------------------
# xlsx report generation
# ---------------------------------------------------------------------------

PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
NI_FILL   = PatternFill('solid', fgColor='E0E0E0')
HDRF_FILL = PatternFill('solid', fgColor='2E4A7A')
HDRB_FILL = PatternFill('solid', fgColor='4472C4')
SEC_FILL  = PatternFill('solid', fgColor='D9E1F2')
DASH = '—'

# Section sub-headers keyed by first register address in that section
SECTIONS = {
    0x00: '§ Identity / General',
    0x14: '§ 9.4  Download Control / Status',
    0x20: '§ 9.5  Status & Alarms',
    0x30: '§ 9.6  Channel / Grid / Power Control',
    0x40: '§ 9.6  Laser Frequency Readback',
    0x4F: '§ 9.7  Module Capabilities',
    0x57: '§ 9.8  Command Block — Currents & Temps (AEA)',
    0x59: '§ 9.8  Dither Control',
    0x5D: '§ 9.8  Thermal Boundary (TBTF)',
    0x5F: '§ 9.8  Age Model',
    0x62: '§ 9.6 / 9.5 ext  Fine-tune, Channel & Frequency Extensions',
    0x69: '§ Reserved (0x69–0x7F)',
    0x80: '§ Manufacturer Debug Window (0x80–0x91)',
    0x92: '§ Reserved (0x92–0xFF)',
}


def _section_row(ws, row_idx, label):
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    cell = ws.cell(row=row_idx, column=1, value=label)
    cell.fill = SEC_FILL
    cell.font = Font(bold=True, color='1F3864')
    cell.alignment = Alignment(horizontal='left', indent=1)


def write_xlsx(results, port, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Register Scan'

    now = datetime.datetime.now()
    date_str = now.strftime('%Y-%m-%d  %H:%M')
    n_pass = sum(1 for r in results if r['verify'] == 'PASS')
    n_fail = sum(1 for r in results if r['verify'] == 'FAIL')
    n_ni   = sum(1 for r in results if r['access'] == 'N/I')

    # Row 1 – title
    ws.merge_cells('A1:H1')
    ws['A1'] = 'nano-ITLA  Register Test Report  —  Pilot Photonics'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = HDRF_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    # Row 2 – summary
    ws.merge_cells('A2:H2')
    summary = (
        f'Port: {port}    Date: {date_str}    '
        f'Registers: {len(results)}    PASS: {n_pass}    FAIL: {n_fail}    N/I: {n_ni}'
    )
    ws['A2'] = summary
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='left', indent=1)

    # Row 3 – column headers
    headers = ['Addr', 'Dec', 'Name', 'Access', 'Read', 'Write', 'Verify', 'Notes / Description']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = HDRB_FILL
        c.alignment = Alignment(horizontal='center')

    col_widths = [8, 6, 22, 9, 10, 10, 9, 65]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'

    data_row = 4
    last_section = None

    for r in results:
        addr = r['addr']

        # Insert section sub-header when section boundary reached
        if addr in SECTIONS and addr != last_section:
            _section_row(ws, data_row, SECTIONS[addr])
            data_row += 1
            last_section = addr

        access    = r['access']
        read_v    = r['read_val']
        write_v   = r['write_val']
        verify    = r['verify']

        if r['aea_str'] is not None:
            read_str = f'"{r["aea_str"]}"'
        elif isinstance(read_v, int):
            read_str = f'0x{read_v:04X}'
        else:
            read_str = DASH

        if isinstance(write_v, int):
            write_str = f'0x{write_v:04X}'
        elif write_v == 'XE':
            write_str = 'XE'
        else:
            write_str = DASH

        notes = r['desc']
        if r['ce']:
            notes = '[CE] ' + notes

        row_vals = [
            f'0x{addr:02X}',
            addr,
            r['name'],
            access,
            read_str,
            write_str,
            verify,
            notes,
        ]

        if access == 'N/I':
            fill = NI_FILL
        elif verify == 'PASS':
            fill = PASS_FILL
        elif verify == 'FAIL':
            fill = FAIL_FILL
        else:
            fill = None

        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            if fill:
                c.fill = fill
            c.alignment = Alignment(vertical='center')
            if col == 7 and verify in ('PASS', 'FAIL'):
                c.font = Font(bold=True)

        data_row += 1

    wb.save(out_path)
    print(f'Saved: {out_path}  ({n_pass} PASS, {n_fail} FAIL, {n_ni} N/I)')


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='nano-ITLA register scanner')
    parser.add_argument('--port',    default='/dev/ttyUSB0', help='Serial port (default: /dev/ttyUSB0)')
    parser.add_argument('--baud',    type=int, default=115200, help='Baud rate (default: 115200)')
    parser.add_argument('--timeout', type=float, default=1.0, help='Per-register read timeout in seconds (default: 1.0)')
    parser.add_argument('--out',     default=None, help='Output .xlsx path (default: register_scan_YYYYMMDD.xlsx)')
    args = parser.parse_args()

    if args.out is None:
        stamp = datetime.datetime.now().strftime('%Y%m%d')
        args.out = f'register_scan_{stamp}.xlsx'

    print(f'Scanning {len(REGISTER_MAP)} registers on {args.port} …')
    results = scan(args.port, args.baud, args.timeout)
    write_xlsx(results, args.port, args.out)


if __name__ == '__main__':
    main()
